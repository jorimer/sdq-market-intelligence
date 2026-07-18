"""Tests for the SIPEN pension sync (fixture-backed, no network)."""
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from shared.database.base import Base
from shared.data.sipen_client import afp_catalog, sipen_client
from modules.pension_intel.sipen_sync import sipen_pension_sync
from modules.pension_intel.models.models import (
    PensionEntity,
    PensionSeries,
    PensionSnapshot,
)


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    # autoflush=False mirrors the shared SessionLocal (guards the snapshot-visibility bug).
    Session = sessionmaker(bind=engine, autoflush=False)
    session = Session()
    try:
        yield session
    finally:
        session.close()


def test_sync_seeds_entities_series_and_snapshot(db):
    res = sipen_pension_sync(db)

    # Every AFP in the catalog is seeded.
    assert res["entities_created"] == len(afp_catalog())
    assert db.query(PensionEntity).count() == len(afp_catalog())

    # System series (entity_slug NULL) and per-AFP series both land.
    system = db.query(PensionSeries).filter(PensionSeries.entity_slug.is_(None)).all()
    per_afp = db.query(PensionSeries).filter(PensionSeries.entity_slug.isnot(None)).all()
    assert {s.series_code for s in system} >= {
        "sipen.rentabilidad.cci_nominal_anual",
        "sipen.comisiones.total_anual",
    }
    assert per_afp, "expected per-AFP series"

    # A real, cited figure round-trips intact (no interpolation).
    siembra_ret = (
        db.query(PensionSeries)
        .filter(
            PensionSeries.entity_slug == "afp_siembra",
            PensionSeries.series_code == "rentabilidad_nominal_anual",
            PensionSeries.period == "2025-02",
        )
        .one()
    )
    assert siembra_ret.value == 10.27
    assert siembra_ret.source == "SIPEN"

    # One snapshot per system period (backfill); the LATEST one carries the headline.
    # Chronological latest = "2025" (the bare annual period sorts after its months).
    from shared.products.periods import period_sort_key
    snaps = sorted(db.query(PensionSnapshot).all(),
                   key=lambda s: period_sort_key(s.period))
    assert res["snapshot_period"] == snaps[-1].period == "2025"
    assert snaps[-1].headline.get("sipen.rentabilidad.cci_nominal_anual") == 9.4
    assert snaps[-1].entity_count == len(afp_catalog())


def test_sync_is_idempotent(db):
    first = sipen_pension_sync(db)
    n_after_first = db.query(PensionSeries).count()
    n_snaps_first = db.query(PensionSnapshot).count()

    second = sipen_pension_sync(db)
    n_after_second = db.query(PensionSeries).count()

    # No duplicate rows, entities, or snapshots on a second run.
    assert n_after_first == n_after_second
    assert second["entities_created"] == 0
    assert db.query(PensionEntity).count() == len(afp_catalog())
    assert db.query(PensionSnapshot).count() == n_snaps_first


def test_fixture_has_no_invented_precision():
    """Sanity: the fixture only carries values we can cite (no fabricated numbers)."""
    system = sipen_client.fetch()
    codes = {r.series for r in system}
    assert "sipen.rentabilidad.cci_nominal_anual" in codes
    # Per-AFP rentabilidad present for all 7 administrators.
    ent = sipen_client.fetch_entities(series="rentabilidad_nominal_anual")
    assert {r.dimension for r in ent} == {slug for slug, _ in afp_catalog()}


def test_parse_mes_ano_xlsx_maps_spanish_months():
    """The CKAN 'Mes | Año | valor' workbook → [(YYYY-MM, value)]; junk rows skipped."""
    import io

    import openpyxl

    from shared.data.sipen_client import parse_mes_ano_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Mes", "Año", "Afiliados"])      # header (no month) → skipped
    ws.append(["Septiembre", "2003", 927082])
    ws.append(["Diciembre", "2025", 5600000])
    ws.append(["Totales", "", "x"])             # junk → skipped
    buf = io.BytesIO()
    wb.save(buf)

    rows = dict(parse_mes_ano_xlsx(buf.getvalue()))
    assert rows == {"2003-09": 927082.0, "2025-12": 5600000.0}


def test_previsional_xlsx_links_keys_on_aria_label():
    """The Estadística Previsional page maps each .xlsx to its document label."""
    from shared.data.sipen_client import previsional_xlsx_links

    html = (
        '<a href="https://sipen.gob.do/descarga/x_2026_06_111.xlsx" target="_blank" '
        'aria-label="Descargar documento Rentabilidad"><div>…</div></a>'
        '<a href="https://sipen.gob.do/descarga/x_2026_06_222.pdf" '
        'aria-label="Descargar documento Rentabilidad"></a>'  # pdf ignored
        '<a href="https://sipen.gob.do/descarga/x_2026_06_333.xlsx" '
        'aria-label="Descargar documento Patrimonio"></a>'
    )
    links = previsional_xlsx_links(html)
    assert links["rentabilidad"].endswith("111.xlsx")
    assert links["patrimonio"].endswith("333.xlsx")
    assert all(u.endswith(".xlsx") for u in links.values())


def test_parse_rentabilidad_xlsx_system_and_per_afp():
    """Fraction → percent; per-AFP column→slug; discontinued/0 cells skipped."""
    import datetime
    import io

    import openpyxl

    from shared.data.sipen_client import parse_rentabilidad_xlsx

    wb = openpyxl.Workbook()
    wsys = wb.active
    wsys.title = "Por SISTEMA"
    wsys.append([None, "Rentabilidad Nominal por Sistema"])
    wsys.append([None, "Mes", "Promedio CCI", "Promedio Sistema"])
    wsys.append([None, datetime.datetime(2025, 4, 30), 0.09398, 0.09604])
    wsys.append([None, datetime.datetime(2026, 5, 31), 0.0775, 0.0798])

    wf = wb.create_sheet("Por FONDO")
    wf.append([None, "Mes", "AFP"])
    wf.append([None, None, "Atlántico", "Camino1", "Crecer", "Popular", "Siembra"])
    wf.append([None, datetime.datetime(2026, 5, 31), 0.0599, 0, 0.0766, 0.0769, 0.0798])

    buf = io.BytesIO()
    wb.save(buf)
    system, per_afp = parse_rentabilidad_xlsx(buf.getvalue())

    assert system == [("2025-04", 9.4, 9.6), ("2026-05", 7.75, 7.98)]
    # Discontinued "Camino1" (0) dropped; the 4 live AFPs map to slugs.
    assert per_afp["afp_atlantico"] == [("2026-05", 5.99)]
    assert per_afp["afp_crecer"] == [("2026-05", 7.66)]
    assert per_afp["afp_popular"] == [("2026-05", 7.69)]
    assert per_afp["afp_siembra"] == [("2026-05", 7.98)]
    assert "afp_camino" not in per_afp


def test_sync_ingests_rentabilidad_live(db, monkeypatch):
    """Live rentabilidad Records (system + per-AFP) land and extend the fixture: the
    snapshot headline and an AFP's return both move to the newer live period."""
    from datetime import date

    from shared.data.base_client import Record
    from shared.data.lineage import Lineage

    lin = Lineage(source="SIPEN", license="x", fetched_at=date.today())
    sys_recs = [
        Record(series="sipen.rentabilidad.cci_nominal_anual", period="2026-05",
               value=7.75, lineage=lin, unit="%"),
        Record(series="sipen.rentabilidad.sdp_nominal_anual", period="2026-05",
               value=7.98, lineage=lin, unit="%"),
    ]
    ent_recs = [
        Record(series="rentabilidad_nominal_anual", period="2026-05", value=7.98,
               lineage=lin, unit="%", dimension="afp_siembra"),
    ]
    monkeypatch.setattr(
        "modules.pension_intel.sipen_sync.fetch_sipen_rentabilidad",
        lambda period=None: (sys_recs, ent_recs),
    )

    res = sipen_pension_sync(db)
    assert res["rentabilidad_rows"] == 3

    # The newer live value lands for the AFP (alongside the fixture's 2025-02).
    siembra_live = (db.query(PensionSeries)
                    .filter_by(entity_slug="afp_siembra",
                               series_code="rentabilidad_nominal_anual", period="2026-05").one())
    assert siembra_live.value == 7.98

    # Snapshot headline picks up the live 2026-05 CCI (latest per code).
    snap = db.query(PensionSnapshot).order_by(PensionSnapshot.period.desc()).first()
    assert snap.period == "2026-05"
    assert snap.headline.get("sipen.rentabilidad.cci_nominal_anual") == 7.75
    assert snap.headline.get("sipen.rentabilidad.sdp_nominal_anual") == 7.98


def test_sync_ingests_ckan_series_and_snapshot_uses_latest_per_code(db, monkeypatch):
    """Live CKAN records land as system series, and the snapshot keeps each indicator's
    latest value (fresh afiliados doesn't drop the older fixture rentabilidad)."""
    from datetime import date

    from shared.data.lineage import Lineage
    from shared.data.base_client import Record

    lin = Lineage(source="SIPEN", license="x", fetched_at=date.today())
    fake = [
        Record(series="sipen.afiliados.total", period="2026-05", value=5649211.0, lineage=lin, unit="personas"),
        Record(series="sipen.cotizantes.total", period="2026-05", value=2213058.0, lineage=lin, unit="personas"),
    ]
    monkeypatch.setattr("modules.pension_intel.sipen_sync.fetch_sipen_ckan", lambda period=None: fake)

    res = sipen_pension_sync(db)
    assert res["ckan_rows"] == 2

    afi = (db.query(PensionSeries)
           .filter_by(entity_slug=None, series_code="sipen.afiliados.total", period="2026-05").one())
    assert afi.value == 5649211.0 and afi.source == "SIPEN"

    # Snapshot is at the newest period but still carries the older fixture rentabilidad.
    snap = db.query(PensionSnapshot).order_by(PensionSnapshot.period.desc()).first()
    assert snap.period == "2026-05"
    assert snap.headline.get("sipen.afiliados.total") == 5649211.0
    assert "sipen.rentabilidad.cci_nominal_anual" in snap.headline  # older series retained


# ── Backfill de snapshots por período + pulso as-of (2026-07-18) ─────────────
# Regresión (espejo del PR #552 de seguros): en prod solo existían 2 snapshots (los
# períodos que fueron "el último" en algún sync) pese a que las series SIPEN traen
# historia mensual desde 2003-07 → el selector de períodos del producto ofrecía 2
# opciones, y elegir una servía el pulso ACTUAL.

def _fake_rentabilidad_history():
    """24 meses de rentabilidad live (system + per-AFP) con valor distinto por mes."""
    from datetime import date

    from shared.data.base_client import Record
    from shared.data.lineage import Lineage

    lin = Lineage(source="SIPEN", license="x", fetched_at=date.today())
    sys_recs, ent_recs = [], []
    for i, (y, m) in enumerate([(y, m) for y in (2023, 2024) for m in range(1, 13)]):
        p = f"{y}-{m:02d}"
        sys_recs.append(Record(series="sipen.rentabilidad.cci_nominal_anual", period=p,
                               value=8.0 + i * 0.1, lineage=lin, unit="%"))
        ent_recs.append(Record(series="rentabilidad_nominal_anual", period=p,
                               value=7.5 + i * 0.1, lineage=lin, unit="%",
                               dimension="afp_siembra"))
    return sys_recs, ent_recs


def test_sync_backfills_snapshot_per_period(db, monkeypatch):
    from shared.products.periods import period_sort_key

    monkeypatch.setattr("modules.pension_intel.sipen_sync.fetch_sipen_rentabilidad",
                        lambda period=None: _fake_rentabilidad_history())
    sipen_pension_sync(db)

    periods = sorted((p for (p,) in db.query(PensionSnapshot.period).all()),
                     key=period_sort_key)
    system_periods = sorted({p for (p,) in db.query(PensionSeries.period)
                             .filter(PensionSeries.entity_slug.is_(None),
                                     PensionSeries.value.isnot(None)).all()},
                            key=period_sort_key)
    assert periods == system_periods          # un snapshot por período con dato
    assert len(periods) > 12                  # historia real, no solo "el último"

    # El headline de cada snapshot es la vista AS-OF (no la actual): en 2023-06 la
    # rentabilidad vigente era la de ese mes, no la del último período.
    snap_2306 = db.query(PensionSnapshot).filter_by(period="2023-06").one()
    assert snap_2306.headline["sipen.rentabilidad.cci_nominal_anual"] == 8.5
