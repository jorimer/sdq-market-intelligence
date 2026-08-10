"""Tests for the ONE social connector + sync (Eje 6 Gate A). Offline."""
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from shared.data.one_client import ONEClient, _parse_poverty_csv, region_catalog
from shared.database.base import Base
from modules.social_dev.models.models import SocialIndicator  # noqa: F401 — register table
from modules.social_dev.social_sync import one_social_sync


@pytest.fixture
def db():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=engine)
    s = sessionmaker(bind=engine)()
    try:
        yield s
    finally:
        s.close()
        Base.metadata.drop_all(bind=engine)


def test_parse_poverty_csv_maps_themes_and_regions():
    csv = (
        "Tasa de Pobreza ,Tipo de Regiones,Porcentaje,Año\n"
        "Pobreza General,Enriquillo,31,2024\n"
        "Pobreza Extrema, Cibao Norte ,2,2024\n"          # leading/trailing spaces
        "Pobreza General,Región Inexistente,99,2024\n"     # unknown region → skipped
    ).encode("utf-8")
    out = _parse_poverty_csv(csv)
    assert ("poverty_rate", "enriquillo", "2024", 31.0) in out
    assert ("poverty_extreme", "cibao_norte", "2024", 2.0) in out          # space-tolerant
    assert not any(slug not in dict(region_catalog()) for _, slug, _, _ in out)  # no unknowns


def test_ozama_alias_matches():
    # Ozama (Gran Santo Domingo) under a single-token rename must still map.
    for label in ("Ozama", "Metropolitana", "Gran Santo Domingo"):
        csv = f"a,b,c,d\nPobreza General,{label},20,2024\n".encode("utf-8")
        out = _parse_poverty_csv(csv)
        assert out and out[0][1] == "ozama"


def test_fixture_client_offline():
    recs = ONEClient(mode="fixture").fetch()
    assert recs, "fixture vacío — ¿se generó one.json?"
    assert {r.dimension for r in recs} == {slug for slug, _ in region_catalog()}  # 10 regions
    assert {r.series for r in recs} >= {"poverty_rate"}


def test_sync_persists_and_is_idempotent(db, monkeypatch):
    monkeypatch.setattr(ONEClient, "_fetch_live", ONEClient._fetch_fixture)
    # Las demás fuentes pegan a la red — se sustituyen. Devuelven un conteo POSITIVO
    # porque hacen de sub-syncs que funcionan: un 0 ahora significa "la fuente no trajo
    # nada" y queda declarado en ``errors``, que es justo lo que este test NO mide.
    for _fn in ("_sync_wdi_health", "_sync_bcrd_informality", "_sync_sisdom_income",
                "_sync_minerd_coverage", "_sync_one_schooling", "_sync_wb_findex",
                "_sync_siuben_provincial"):
        monkeypatch.setattr(f"modules.social_dev.social_sync.{_fn}", lambda db, set_phase, *a: 1)

    first = one_social_sync(db)
    assert first["errors"] == []
    assert first["synced"] > 0
    assert first["regions"] == 10
    assert "poverty_rate" in first["themes"]
    n1 = db.query(SocialIndicator).count()
    assert n1 == first["synced"]

    second = one_social_sync(db)          # upsert in place — no duplicates
    assert db.query(SocialIndicator).count() == n1

    row = (
        db.query(SocialIndicator)
        .filter_by(entity_key="enriquillo", theme="poverty_rate", period="2024")
        .first()
    )
    assert row is not None and row.source == "ONE" and row.disaggregation == "region"


def test_una_fuente_caida_queda_declarada_en_errors(db, monkeypatch):
    """Un cero sin explicación es un éxito aparente. Pasó en producción el 2026-08-09:
    cuatro fuentes devolvieron cero y la operación reportó ``errors: []``, así que la
    consola mostró la corrida como buena. Un guard que no reporta no protege: esconde.

    Las dos causas se distinguen porque se actúan distinto — 'no pudimos llegar' es un
    problema a investigar; 'la fuente no publicó' puede ser legítimo."""
    monkeypatch.setattr(ONEClient, "_fetch_live", ONEClient._fetch_fixture)
    monkeypatch.setattr("modules.social_dev.social_sync._sync_wdi_health", lambda db, sp, *a: 7)
    monkeypatch.setattr("modules.social_dev.social_sync._sync_sisdom_income", lambda db, sp, *a: 0)
    monkeypatch.setattr("modules.social_dev.social_sync._sync_bcrd_informality", lambda db, sp, *a: 4)
    monkeypatch.setattr("modules.social_dev.social_sync._sync_one_schooling", lambda db, sp, *a: 0)
    monkeypatch.setattr("modules.social_dev.social_sync._sync_wb_findex", lambda db, sp, *a: 0)

    def _boom(db, sp, *a):
        raise ConnectionError("403 Forbidden")

    monkeypatch.setattr("modules.social_dev.social_sync._sync_minerd_coverage", _boom)
    monkeypatch.setattr("modules.social_dev.social_sync._sync_siuben_provincial", _boom)

    res = one_social_sync(db)
    errores = " · ".join(res["errors"])

    # No se pudo llegar → la causa VIAJA (tipo y mensaje), no se pierde en un log.
    # El rótulo nombra al EMISOR vigente: la cobertura ya no es de la ONE, y mandar a
    # mirar el portal equivocado es parte de no declarar bien la falla.
    assert "cobertura educativa (MINERD · SIIE)" in errores and "403 Forbidden" in errores
    assert "(ONE)" not in errores.split("cobertura educativa")[1][:30]
    assert "indicadores provinciales (SIUBEN)" in errores
    assert "ConnectionError" in errores
    # Cero sin excepción → también consta, con otra redacción.
    assert "la fuente respondió sin observaciones" in errores
    # Lo que sí trajo dato NO ensucia el reporte.
    assert "salud nacional" not in errores
    assert res["coverage_synced"] == 0 and res["provincial_synced"] == 0
    assert res["health_synced"] == 7


def test_sin_fallas_no_se_inventan_errores(db, monkeypatch):
    """El contrapeso del test anterior: si todo trajo dato, ``errors`` queda vacío."""
    monkeypatch.setattr(ONEClient, "_fetch_live", ONEClient._fetch_fixture)
    for fn in ("_sync_wdi_health", "_sync_bcrd_informality", "_sync_sisdom_income",
               "_sync_minerd_coverage", "_sync_one_schooling", "_sync_wb_findex",
               "_sync_siuben_provincial"):
        monkeypatch.setattr(f"modules.social_dev.social_sync.{fn}", lambda db, sp, *a: 3)
    assert one_social_sync(db)["errors"] == []


def test_el_uso_de_una_instantanea_viaja_al_resultado(db, monkeypatch, tmp_path):
    """El respaldo NO puede usarse en silencio.

    Producción no alcanza siuben.gob.do y el tablero del MINERD limita por tasa, así que
    la instantánea comiteada es un camino REAL, no teórico. Y un camino real que no se
    declara es exactamente cómo un fallback se vuelve permanente sin que nadie lo note
    (ya pasó acá con los 'promedios del sistema'). Por eso la procedencia sale en el
    resultado de la operación, al lado de los contadores."""
    from shared.data import snapshots

    monkeypatch.setattr(snapshots, "SNAPSHOT_DIR", tmp_path)
    snapshots.write_snapshot("siuben_provincial",
                             [("siuben_illiteracy_head_share", "elias_pina", "2024-Q4", 36.3)],
                             source="SIUBEN")

    import shared.data.siuben_client as siuben

    def _caido():
        raise ConnectionError("timed out")

    monkeypatch.setattr(siuben, "fetch_siuben_provincial", _caido)
    from modules.social_dev.social_sync import _sync_siuben_provincial

    prov = {}
    assert _sync_siuben_provincial(db, lambda _m: None, prov) == 1
    db.commit()
    # La cifra entró…
    row = (
        db.query(SocialIndicator)
        .filter_by(entity_key="elias_pina", theme="siuben_illiteracy_head_share").first()
    )
    assert row is not None and row.value == 36.3
    # …y el resultado DICE que vino de una foto, con su fecha.
    assert prov["siuben_provincial"].startswith("snapshot:")
    assert len(prov["siuben_provincial"].split(":", 1)[1]) == 10


def test_con_la_fuente_viva_la_procedencia_dice_live(db, monkeypatch):
    """El contrapeso: si la fuente responde, no se rotula como foto."""
    import shared.data.siuben_client as siuben

    monkeypatch.setattr(
        siuben, "fetch_siuben_provincial",
        lambda: [("siuben_illiteracy_head_share", "azua", "2024-Q4", 20.0)])
    from modules.social_dev.social_sync import _sync_siuben_provincial

    prov = {}
    assert _sync_siuben_provincial(db, lambda _m: None, prov) == 1
    assert prov["siuben_provincial"] == "live"


def _ind(db, entity, theme, period, value, source="ONE"):
    db.add(SocialIndicator(entity_key=entity, theme=theme, period=period, value=value, source=source))


def test_assemble_idm_real_plus_rubric_with_sources(db):
    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    _ind(db, "nacional", "life_expectancy", "2024", 73.9, source="WDI")
    _ind(db, "nacional", "child_mortality", "2024", 27.7, source="WDI")
    db.commit()

    from modules.social_dev.service import assemble_idm_dataset
    asm = assemble_idm_dataset(db)
    assert asm["period"] == "2024" and asm["has_live"]
    assert len(asm["dataset"]) == 10                         # the 10 development regions
    enr, src = asm["dataset"]["enriquillo"], asm["sources"]["enriquillo"]
    assert enr["poverty_rate"] == 31.0 and src["poverty_rate"] == "live"   # ONE, by region
    assert enr["life_expectancy"] == 73.9 and src["life_expectancy"] == "live"  # WDI national
    # No national labour yet → income/informality fall back to declared rubric 50.
    assert src["income_per_capita"] == "rubric" and enr["income_per_capita"] == 50
    assert src["informality_rate"] == "rubric" and enr["informality_rate"] == 50


def test_informalidad_nacional_va_live_a_todas_las_regiones(db):
    """La informalidad (ENCFT del BCRD) sí es nacional: la MISMA cifra a las 10 regiones."""
    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    _ind(db, "nacional", "informality_rate", "2024", 55.46, source="BCRD")
    db.commit()

    from modules.social_dev.service import assemble_idm_dataset
    asm = assemble_idm_dataset(db)
    for slug in ("enriquillo", "valdesia"):                  # uniform across regions
        row, src = asm["dataset"][slug], asm["sources"][slug]
        assert row["informality_rate"] == 55.46 and src["informality_rate"] == "live"
    # financial_inclusion has no source → stays declared rubric.
    assert asm["sources"]["enriquillo"]["financial_inclusion"] == "rubric"


def _all_regions_income(db, period="2024", base=13000.0):
    from shared.data.one_client import REGIONS

    for i, (slug, _label) in enumerate(REGIONS):
        _ind(db, slug, "income_per_capita", period, base + i * 500, source="MEPyD")


def test_el_ingreso_ahora_DISTINGUE_regiones(db):
    """El corazón del cambio. El ingreso era el proxy horario de la ONE: una constante
    nacional, la misma cifra para las 10 regiones, que sostenía el NIVEL del índice pero
    no podía mover el orden entre demarcaciones. Con el SISDOM cada región trae la suya."""
    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    _all_regions_income(db)
    db.commit()

    from modules.social_dev.service import assemble_idm_dataset
    asm = assemble_idm_dataset(db)
    valores = {slug: row["income_per_capita"] for slug, row in asm["dataset"].items()}
    assert len(set(valores.values())) == 10, (
        "el ingreso volvió a ser una constante con etiqueta geográfica: " f"{valores}")
    assert all(s["income_per_capita"] == "live" for s in asm["sources"].values())


def test_el_ingreso_incompleto_NO_entra_a_medias(db):
    """Nueve de diez regiones no es 'casi': el motor normaliza min-max ENTRE regiones, así
    que la que falta corre el mínimo o el máximo y cambia el score de las otras nueve.
    O están las 10 o la variable queda en rúbrica uniforme, que no discrimina ni finge."""
    from shared.data.one_client import REGIONS

    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    for slug, _label in REGIONS[:-1]:                       # falta una
        _ind(db, slug, "income_per_capita", "2024", 15000.0, source="MEPyD")
    db.commit()

    from modules.social_dev.service import assemble_idm_dataset
    asm = assemble_idm_dataset(db)
    assert all(s["income_per_capita"] == "rubric" for s in asm["sources"].values())
    assert {row["income_per_capita"] for row in asm["dataset"].values()} == {50.0}


def test_el_ingreso_no_se_cae_si_la_pobreza_avanza_primero(db):
    """El SISDOM es anual y puede quedar un año detrás de la pobreza, que es la serie que
    fija el período objetivo. Se toma el ÚLTIMO valor disponible por región: si se atara
    al período, la variable desaparecería del panel el día que la pobreza publique antes."""
    _ind(db, "enriquillo", "poverty_rate", "2025", 30.0)
    _ind(db, "valdesia", "poverty_rate", "2025", 10.0)
    _all_regions_income(db, period="2024")                  # el ingreso va un año atrás
    db.commit()

    from modules.social_dev.service import assemble_idm_dataset
    asm = assemble_idm_dataset(db)
    assert asm["period"] == "2025"
    assert asm["sources"]["enriquillo"]["income_per_capita"] == "live"
    assert asm["dataset"]["enriquillo"]["income_per_capita"] == 15500.0


def test_parse_one_indicator_xlsx_reads_total_by_year():
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = " Ficha "                              # metadata sheet (ignored)
    wb.active["A1"] = "Ficha técnica"
    ws = wb.create_sheet("Indicador")
    ws.append(["REPÚBLICA DOMINICANA: Tasa de Informalidad…"])  # title row
    ws.append(["Año", "Total", "Hombres", "Mujeres"])           # header
    ws.append([2022, 57.56, 61.21, 52.3])
    ws.append([2023, 56.54, 60.3, 51.27])
    ws.append(["Fuente: ENCFT (BCRD)"])                         # trailing note (ignored)
    buf = io.BytesIO()
    wb.save(buf)

    from shared.data.one_client import parse_one_indicator_xlsx
    assert parse_one_indicator_xlsx(buf.getvalue()) == [(2022, 57.56), (2023, 56.54)]


def test_parse_one_indicator_tolerates_sheet_whitespace():
    """ONE sheet names often carry trailing spaces — must still be found."""
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = " Ficha "
    ws = wb.create_sheet("Indicador ")              # trailing space
    ws.append(["Año", "Total"])
    ws.append([2024, 167.46])
    wb.create_sheet("Notas")                         # later sheet must NOT win
    buf = io.BytesIO()
    wb.save(buf)

    from shared.data.one_client import parse_one_indicator_xlsx
    assert parse_one_indicator_xlsx(buf.getvalue()) == [(2024, 167.46)]


def test_el_descubrimiento_por_media_hash_prefiere_la_revision_mas_nueva():
    """El hash del CDN de la ONE rota y conviven revisiones del mismo archivo. Gana la de
    año final más alto (desempate determinista). Hoy el único consumidor es la
    escolaridad, que es lo último que este módulo saca del portal."""
    from shared.data.one_client import _SCHOOLING_SLUG, _match_media_links

    html = (
        f'<a href="/media/aaa/{_SCHOOLING_SLUG}-2000-2023.xlsx">vieja</a>'
        f'<a href="/media/bbb/{_SCHOOLING_SLUG}-2000-2024.xlsx">nueva</a>'
        '<a href="/media/zzz/poblacion-desocupada-2008-2024.xlsx">distractor</a>'
    )
    links = _match_media_links(html, {"schooling_years": _SCHOOLING_SLUG})
    assert links["schooling_years"].endswith("2000-2024.xlsx")
    assert set(links) == {"schooling_years"}


def test_region_slug_es_el_padron_de_regiones_para_cualquier_emisor():
    """``region_slug`` sobrevivió al parser de cobertura de la ONE porque el padrón de
    regiones —con sus alias— no era de ese cuadro: hoy lo consume el conector del
    MINERD, que es otro emisor. Se fija acá para que un borrado futuro no se lo lleve."""
    from shared.data.one_client import region_slug

    assert region_slug("Región Metropolitana") == "ozama"      # alias + prefijo "Región"
    assert region_slug("Cibao Norte") == "cibao_norte"          # sin prefijo
    assert region_slug("Gran Santo Domingo") == "ozama"
    assert region_slug("Total país") is None                    # no se adivina
    assert region_slug(None) is None


def test_coverage_goes_live_by_region_and_period(db):
    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    _ind(db, "enriquillo", "secondary_coverage", "2024", 66.8)   # valdesia has none
    db.commit()

    from modules.social_dev.service import assemble_idm_dataset
    asm = assemble_idm_dataset(db)
    enr = asm["dataset"]["enriquillo"]
    assert enr["secondary_coverage"] == 66.8                            # by region + period
    assert asm["sources"]["enriquillo"]["secondary_coverage"] == "live"
    # A region without coverage that period → rubric, value excluded (engine skips it).
    assert asm["sources"]["valdesia"]["secondary_coverage"] == "rubric"
    assert "secondary_coverage" not in asm["dataset"]["valdesia"]


def test_indicator_units_fit_postgres_varchar40():
    """sd_indicators.unit is VARCHAR(40): SQLite ignores it, Postgres truncates →
    every declared unit string must fit (dev↔prod parity guard)."""
    from shared.data.siuben_client import DATASETS as SIUBEN_DATASETS
    from shared.data.sisdom_income import UNIT as INCOME_UNIT
    from modules.social_dev.social_sync import COVERAGE_UNIT, FINDEX_UNIT, INFORMALITY_UNIT

    units = ([INFORMALITY_UNIT, INCOME_UNIT, COVERAGE_UNIT, FINDEX_UNIT, "años"]
             + [s.unit for s in SIUBEN_DATASETS])
    too_long = [u for u in units if len(u) > 40]
    assert not too_long, f"unit > 40 chars (rompe en Postgres): {too_long}"


def test_columnas_de_texto_provinciales_caben_en_postgres():
    """Los slugs provinciales y los códigos SIUBEN viajan por columnas acotadas:
    entity_key VARCHAR(60), theme VARCHAR(60), source VARCHAR(40), period VARCHAR(10)."""
    from shared.data.siuben_client import DATASETS as SIUBEN_DATASETS
    from shared.data.siuben_client import SOURCE as SIUBEN_SOURCE
    from shared.reference.provinces import PROVINCES

    assert all(len(slug) <= 60 for slug, _n, _r in PROVINCES)
    assert all(len(s.theme) <= 60 for s in SIUBEN_DATASETS)
    assert len(SIUBEN_SOURCE) <= 40
    assert len("2026-Q4") <= 10          # el período trimestral del SIUBEN


def test_sync_siuben_provincial_upserts_por_provincia(db, monkeypatch):
    """Las series provinciales entran rotuladas y son idempotentes."""
    import shared.data.siuben_client as siuben

    rows = [("siuben_illiteracy_head_share", "elias_pina", "2024-Q4", 36.3),
            ("siuben_illiteracy_head_share", "distrito_nacional", "2024-Q4", 7.5),
            ("siuben_overcrowding_share", "elias_pina", "2024-Q4", 22.1)]
    monkeypatch.setattr(siuben, "fetch_siuben_provincial", lambda: rows)
    from modules.social_dev.social_sync import _sync_siuben_provincial

    assert _sync_siuben_provincial(db, lambda _m: None, {}) == 3
    db.commit()
    row = (
        db.query(SocialIndicator)
        .filter_by(entity_key="elias_pina", theme="siuben_illiteracy_head_share",
                   period="2024-Q4")
        .first()
    )
    assert row is not None and row.value == 36.3
    assert row.source == "SIUBEN" and row.disaggregation == "provincia"
    assert "padrón" in (row.unit or "")        # el universo viaja con el dato

    # Segunda corrida: upsert en el lugar, sin duplicar.
    assert _sync_siuben_provincial(db, lambda _m: None, {}) == 3
    db.commit()
    assert db.query(SocialIndicator).filter_by(source="SIUBEN").count() == 3


def test_siuben_no_contamina_el_idm(db, monkeypatch):
    """Mismo guardia que con la cobertura provincial: el IDM no debe moverse."""
    import shared.data.siuben_client as siuben
    from modules.social_dev.service import assemble_idm_dataset
    from modules.social_dev.social_sync import _sync_siuben_provincial

    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    db.commit()
    before = assemble_idm_dataset(db)

    monkeypatch.setattr(
        siuben, "fetch_siuben_provincial",
        lambda: [("siuben_illiteracy_head_share", "elias_pina", "2024-Q4", 36.3)])
    _sync_siuben_provincial(db, lambda _m: None, {})
    db.commit()

    after = assemble_idm_dataset(db)
    assert after["dataset"] == before["dataset"] and after["sources"] == before["sources"]


def test_findex_financial_inclusion_goes_live_latest_available(db):
    """WB Findex financial access is national + lags; the latest value goes live for
    the current IDM period (so inclusión is real even if the target year has no obs)."""
    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    _ind(db, "nacional", "financial_inclusion", "2023", 40.06, source="WB")  # 2023, not 2024
    db.commit()

    from modules.social_dev.service import assemble_idm_dataset
    asm = assemble_idm_dataset(db)
    assert asm["period"] == "2024"
    for slug in ("enriquillo", "valdesia"):                     # national → every region, live
        assert asm["dataset"][slug]["financial_inclusion"] == 40.06
        assert asm["sources"][slug]["financial_inclusion"] == "live"


def test_sync_wb_findex_upserts_national(db, monkeypatch):
    import shared.data.wdi_client as wdi

    monkeypatch.setattr(
        wdi, "fetch_wb_indicator",
        lambda code, isos, mrv=25: ([{"date": "2023", "value": 40.06},
                                     {"date": "2022", "value": 38.5}], None),
    )
    from modules.social_dev.social_sync import _sync_wb_findex

    n = _sync_wb_findex(db, lambda _m: None)
    db.commit()
    assert n == 2
    row = (
        db.query(SocialIndicator)
        .filter_by(entity_key="nacional", theme="financial_inclusion", period="2023")
        .first()
    )
    assert row is not None and row.value == 40.06 and row.source == "WB"


def test_sync_one_schooling_upserts_national(db, monkeypatch):
    import shared.data.one_client as oc_mod

    monkeypatch.setattr(oc_mod, "fetch_one_education_schooling",
                        lambda: [(2023, 9.61), (2024, 9.61)])
    from modules.social_dev.social_sync import _sync_one_schooling

    n = _sync_one_schooling(db, lambda _m: None)
    db.commit()
    assert n == 2
    row = (
        db.query(SocialIndicator)
        .filter_by(entity_key="nacional", theme="schooling_years", period="2024")
        .first()
    )
    assert row is not None and row.value == 9.61 and row.source == "ONE" and row.unit == "años"


def test_sync_minerd_coverage_upserts_by_region(db, monkeypatch):
    import shared.data.minerd_coverage as mc_mod

    monkeypatch.setattr(
        mc_mod, "fetch_minerd_coverage",
        lambda: [("region", "enriquillo", 2024, 66.8), ("region", "ozama", 2024, 67.7),
                 ("provincia", "distrito_nacional", 2024, 73.1)],
    )
    from modules.social_dev.social_sync import _sync_minerd_coverage

    n = _sync_minerd_coverage(db, lambda _m: None, {})
    db.commit()
    assert n == 3
    row = (
        db.query(SocialIndicator)
        .filter_by(entity_key="enriquillo", theme="secondary_coverage", period="2024")
        .first()
    )
    assert row is not None and row.value == 66.8 and row.source == "MINERD"
    assert row.disaggregation == "region"
    # La provincia convive con la región bajo el mismo tema, distinguida por el nivel.
    prov = (
        db.query(SocialIndicator)
        .filter_by(entity_key="distrito_nacional", theme="secondary_coverage", period="2024")
        .first()
    )
    assert prov is not None and prov.value == 73.1 and prov.disaggregation == "provincia"


def test_cobertura_provincial_no_mueve_el_idm(db, monkeypatch):
    """Guardia estructural: el IDM se ensambla sobre las 10 regiones. Agregar filas
    provinciales bajo el MISMO tema no debe cambiar ni un score ni una procedencia —
    si algún día el ensamblador empezara a barrer entidades, este test lo detecta."""
    import shared.data.minerd_coverage as mc_mod
    from modules.social_dev.service import assemble_idm_dataset
    from modules.social_dev.social_sync import _sync_minerd_coverage

    _ind(db, "enriquillo", "poverty_rate", "2024", 31.0)
    _ind(db, "valdesia", "poverty_rate", "2024", 11.0)
    _ind(db, "enriquillo", "secondary_coverage", "2024", 66.8)
    db.commit()
    before = assemble_idm_dataset(db)

    monkeypatch.setattr(
        mc_mod, "fetch_minerd_coverage",
        lambda: [("provincia", "distrito_nacional", 2024, 73.1),
                 ("provincia", "elias_pina", 2024, 41.2)],
    )
    _sync_minerd_coverage(db, lambda _m: None, {})
    db.commit()

    after = assemble_idm_dataset(db)
    assert after["dataset"] == before["dataset"]
    assert after["sources"] == before["sources"]
    assert len(after["dataset"]) == 10          # sigue siendo el panel de regiones


def test_sync_sisdom_income_upserts_por_region(db, monkeypatch):
    """El ingreso entra rotulado por región y es idempotente."""
    import shared.data.sisdom_income as sisdom

    rows = [("enriquillo", 2024, 13499.35), ("cibao_norte", 2024, 19607.32),
            ("enriquillo", 2023, 11118.8)]
    monkeypatch.setattr(sisdom, "fetch_sisdom_income_per_capita", lambda: rows)
    from modules.social_dev.social_sync import _sync_sisdom_income

    assert _sync_sisdom_income(db, lambda _m: None) == 3
    db.commit()
    row = (
        db.query(SocialIndicator)
        .filter_by(entity_key="enriquillo", theme="income_per_capita", period="2024")
        .first()
    )
    assert row is not None and row.value == 13499.35
    assert row.source == "MEPyD" and row.disaggregation == "region"
    assert "RD$" in (row.unit or ""), "la unidad monetaria viaja con el dato"

    assert _sync_sisdom_income(db, lambda _m: None) == 3     # upsert en el lugar
    db.commit()
    assert db.query(SocialIndicator).filter_by(theme="income_per_capita").count() == 3


def test_el_proxy_nacional_de_la_ONE_se_da_de_baja(db, monkeypatch):
    """Bajo el mismo tema convivirían RD$/hora (~167) y RD$/mes por persona (~18.000):
    dos órdenes de magnitud y dos unidades. El upsert no las alcanza porque cambió la
    entidad, así que quedarían para siempre esperando a que algo las lea."""
    import shared.data.sisdom_income as sisdom

    _ind(db, "nacional", "income_per_capita", "2024", 167.46, source="ONE")
    _ind(db, "nacional", "schooling_years", "2024", 9.61, source="ONE")   # NO se toca
    db.commit()

    monkeypatch.setattr(sisdom, "fetch_sisdom_income_per_capita",
                        lambda: [("enriquillo", 2024, 13499.35)])
    from modules.social_dev.social_sync import _sync_sisdom_income

    _sync_sisdom_income(db, lambda _m: None)
    db.commit()

    assert db.query(SocialIndicator).filter_by(
        entity_key="nacional", theme="income_per_capita").count() == 0
    # La baja es QUIRÚRGICA: la escolaridad nacional de la ONE sigue intacta.
    assert db.query(SocialIndicator).filter_by(
        entity_key="nacional", theme="schooling_years").first() is not None


def test_si_el_SISDOM_falla_NO_se_borra_lo_que_hay(db, monkeypatch):
    """El orden importa: primero traer, después dar de baja. Si se borrara antes, una
    caída del MEPyD dejaría la variable sin nada y con la vieja ya destruida."""
    import shared.data.sisdom_income as sisdom

    _ind(db, "nacional", "income_per_capita", "2024", 167.46, source="ONE")
    db.commit()

    def _boom():
        raise sisdom.SisdomUnavailable("el listado no respondió")

    monkeypatch.setattr(sisdom, "fetch_sisdom_income_per_capita", _boom)
    from modules.social_dev.social_sync import _sync_sisdom_income

    with pytest.raises(sisdom.SisdomUnavailable):     # sube a _best_effort, que la DECLARA
        _sync_sisdom_income(db, lambda _m: None)
    assert db.query(SocialIndicator).filter_by(theme="income_per_capita").count() == 1


def test_la_serie_de_ingreso_se_sirve_como_MONTO_no_como_tasa(db):
    """La Data API tenía la naturaleza clavada en 'rate' porque todas las series del eje
    eran porcentajes. El ingreso viene en RD$: servirlo como tasa haría que un consumidor
    leyera su variación en PUNTOS y publicara 'el ingreso subió 2.400 puntos'."""
    from shared.data.sisdom_income import UNIT
    from modules.social_dev.service import subnational_series

    db.add(SocialIndicator(entity_key="enriquillo", theme="income_per_capita",
                           period="2024", value=13499.35, source="MEPyD",
                           disaggregation="region", unit=UNIT))
    db.add(SocialIndicator(entity_key="enriquillo", theme="poverty_rate",
                           period="2024", value=31.0, source="ONE",
                           disaggregation="region", unit="% de la población"))
    db.commit()

    by = {s["code"]: s for s in subnational_series(db)}
    ingreso = by["income_per_capita.enriquillo"]
    assert ingreso["nature"] == "flow", "un monto en RD$ no es una tasa"
    assert by["poverty_rate.enriquillo"]["nature"] == "rate"   # el resto no se movió
    # La licencia y el emisor viajan: una serie sin licencia no se sabe si se puede citar.
    assert ingreso["license"] and "MEPyD" in ingreso["license"]
    assert "MEPyD" in ingreso["note"]
    assert "ONE" in by["poverty_rate.enriquillo"]["note"]
