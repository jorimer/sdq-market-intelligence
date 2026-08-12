"""La serie histórica de indicadores del proveedor, en formato ancho.

El proveedor entrega, además del mazo de cada ola, una matriz con la HISTORIA completa:
un bloque de columnas por marca, dentro de cada bloque un grupo por ola, y dentro de cada
grupo una columna por corte geográfico. Trece olas, siete marcas y tres cortes en un solo
libro — la diferencia entre un expediente de cuatro olas y uno de trece.

**La rejilla se recorre por etiquetas, nunca por zancada.** El libro real NO tiene un paso
uniforme de tres columnas: el bloque de una marca ocupa 37 columnas para 13 olas, de modo
que asumir «cada ola son tres columnas» hace que la última ola de una marca lea los cortes
de la marca siguiente. El lector localiza los inicios de marca y de ola desde sus propias
filas de encabezado y asigna a cada ola las columnas que van hasta el siguiente inicio.

**No sobrescribe lo que ya está cargado.** Las olas que el libro de la plantilla ya
aportó traen su BASE MUESTRAL; esta matriz no trae bases y viene redondeada a dos
decimales. Reemplazar un 87,0 con base 300 por un 86 sin base degradaría el expediente:
perdería la banda de confianza y con ella la capacidad de sostener un veredicto. Por eso
la carga es **solo-si-falta**, y lo omitido se informa.

**Sin base no hay veredicto, y así queda.** Las observaciones que entran por esta vía no
llevan ``base_n``: sirven para leer trayectoria y posición, no para emitir un veredicto de
decisión. El motor de significancia ya trata la ausencia de base como no puntuable, así
que la disciplina se sostiene sola sin marcas adicionales.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("sdq.brand_intel.tracker_history")

SHEET_KPIS = "KPIs"

#: Etiqueta del proveedor → código del diccionario del tracker. **Deliberadamente corto.**
#: Solo entran los indicadores cuya correspondencia es inequívoca; un código equivocado es
#: peor que ninguno, porque contamina la serie sin que nada lo advierta. Los enunciados de
#: imagen («% Brand Trust», «% Great tasting food») son porcentajes de acuerdo y NO el
#: índice de atributo por doble indexación del tracker: mapearlos ahí mezclaría dos
#: métricas distintas bajo un mismo código.
_INDICATORS: Dict[str, str] = {
    "notoriedad de marca": "awareness_total",
    "restaurante favorito": "favourite_place",
    "opinion general t2b": "brand_opinion_t2b",
}

#: Corte del proveedor → segmento del expediente.
_SEGMENTS: Dict[str, str] = {
    "total": "total",
    "sd": "Santo Domingo",
    "sg": "Santiago",
    "total sd": "Santo Domingo",
}


def _norm(s: Any) -> str:
    txt = unicodedata.normalize("NFKD", str(s or ""))
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = re.sub(r"[^0-9a-zA-Z ]+", " ", txt)
    return re.sub(r"\s+", " ", txt).strip().lower()


def _slugify(name: str) -> str:
    n = _norm(name)
    return re.sub(r"[^a-z0-9]+", "_", n).strip("_")


@dataclass
class HistoryReading:
    """Una celda de la matriz, resuelta a su coordenada del expediente."""

    brand: str
    brand_slug: str
    period: date
    segment: str
    metric_code: str
    value: float


@dataclass
class HistoryReport:
    brands: List[str] = field(default_factory=list)
    periods: List[str] = field(default_factory=list)
    segments: List[str] = field(default_factory=list)
    metrics: List[str] = field(default_factory=list)
    readings: int = 0
    skipped_indicators: List[str] = field(default_factory=list)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "marcas": self.brands, "olas": self.periods, "cortes": self.segments,
            "indicadores_mapeados": self.metrics, "lecturas": self.readings,
            "indicadores_no_mapeados": self.skipped_indicators,
        }


def _blocks(row: Tuple, start: int = 0) -> List[Tuple[int, int, Any]]:
    """Tramos ``(inicio, fin_exclusivo, etiqueta)`` de una fila de encabezado.

    Cada etiqueta no vacía abre un tramo que se cierra en la etiqueta siguiente. Es lo
    que permite recorrer la rejilla sin asumir un ancho fijo.
    """
    marks = [(j, v) for j, v in enumerate(row) if j >= start and v is not None
             and str(v).strip()]
    out = []
    for i, (j, v) in enumerate(marks):
        end = marks[i + 1][0] if i + 1 < len(marks) else len(row)
        out.append((j, end, v))
    return out


def _find_header_rows(ws) -> Tuple[int, Tuple, Tuple, Tuple]:
    """Localiza las filas de marca, ola y corte. Devuelve la fila donde arrancan los datos."""
    rows = list(ws.iter_rows(min_row=1, max_row=14, values_only=True))
    for i in range(len(rows) - 2):
        brand_row, period_row, cut_row = rows[i], rows[i + 1], rows[i + 2]
        tiene_fechas = any(isinstance(v, (datetime, date)) for v in period_row)
        tiene_cortes = any(_norm(v) in _SEGMENTS for v in cut_row if v)
        tiene_marcas = sum(1 for v in brand_row if v and str(v).strip()) >= 2
        if tiene_fechas and tiene_cortes and tiene_marcas:
            return i + 4, brand_row, period_row, cut_row
    raise ValueError(
        "No se reconoció la cabecera de la matriz de indicadores: se esperan tres filas "
        "consecutivas con las marcas, las olas (fechas) y los cortes (TOTAL/SD/SG)."
    )


def read_history(content: bytes) -> Tuple[List[HistoryReading], HistoryReport]:
    """Lee la matriz ancha y devuelve una lectura por celda resuelta."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if SHEET_KPIS not in wb.sheetnames:
            raise ValueError(f"El libro no trae la hoja «{SHEET_KPIS}».")
        ws = wb[SHEET_KPIS]
        data_row, brand_row, period_row, cut_row = _find_header_rows(ws)

        brand_blocks = _blocks(brand_row)
        rep = HistoryReport(brands=[str(b) for _, _, b in brand_blocks])
        # Coordenadas: columna → (marca, ola, segmento). Se construye una sola vez.
        coords: Dict[int, Tuple[str, date, str]] = {}
        for b_start, b_end, brand in brand_blocks:
            for p_start, p_end, per in _blocks(period_row[:b_end], b_start):
                if not isinstance(per, (datetime, date)):
                    continue
                pdate = per.date() if isinstance(per, datetime) else per
                for col in range(p_start, min(p_end, b_end)):
                    seg = _SEGMENTS.get(_norm(cut_row[col])) if col < len(cut_row) else None
                    if seg:
                        coords[col] = (str(brand), pdate, seg)

        readings: List[HistoryReading] = []
        no_map: set = set()
        for row in ws.iter_rows(min_row=data_row, values_only=True):
            etiquetas = [str(v) for v in row[:3] if v is not None and str(v).strip()]
            if not etiquetas:
                continue
            code = None
            for et in etiquetas:
                if _norm(et) in _INDICATORS:
                    code = _INDICATORS[_norm(et)]
                    break
            if code is None:
                # Los rótulos de agrupación («Indicadores», «Statements») no son
                # indicadores; solo se anota lo que parece uno y no se pudo mapear.
                if len(etiquetas) >= 2:
                    no_map.add(etiquetas[-2] if len(etiquetas) > 2 else etiquetas[0])
                continue
            for col, (brand, pdate, seg) in coords.items():
                if col >= len(row):
                    continue
                val = row[col]
                if not isinstance(val, (int, float)) or isinstance(val, bool):
                    continue
                # La matriz viene en fracción (0.9279 = 92,8%); el expediente en puntos.
                pct = float(val) * 100.0 if abs(float(val)) <= 1.5 else float(val)
                readings.append(HistoryReading(
                    brand=brand, brand_slug=_slugify(brand), period=pdate,
                    segment=seg, metric_code=code, value=round(pct, 2)))

        rep.readings = len(readings)
        rep.periods = sorted({r.period.isoformat() for r in readings})
        rep.segments = sorted({r.segment for r in readings})
        rep.metrics = sorted({r.metric_code for r in readings})
        rep.skipped_indicators = sorted(no_map)
        return readings, rep
    finally:
        wb.close()


#: Meses abreviados como los rotula el proveedor en sus mazos («Jun '26»). No es cosmética:
#: el emparejamiento tolerante de la ruta de mazos reduce «Junio '26» a la clave «jun26», y
#: una etiqueta igual al código reduce a «202606» — no casan, y CADA celda de esa ola se
#: rechaza como «ola no reconocida». Pasó en producción: 1.165 celdas del mazo de Ola 5
#: rechazadas por las 11 olas que esta matriz había creado etiquetadas con su propio código.
_MESES_ABBR = ("Ene", "Feb", "Mar", "Abr", "May", "Jun",
               "Jul", "Ago", "Sep", "Oct", "Nov", "Dic")


def human_wave_label(period: date) -> str:
    """El rótulo con el que una persona —y el mazo del proveedor— nombra la ola."""
    return f"{_MESES_ABBR[period.month - 1]} '{period.strftime('%y')}"


def store_history(db: Any, engagement_id: str, readings: List[HistoryReading],
                  source: Optional[str] = None) -> Dict[str, Any]:
    """Carga las lecturas creando las olas y marcas que falten. **Solo-si-falta.**

    Una observación que ya existe NO se toca: la que está vino de la plantilla con su base
    muestral y esta matriz no trae bases. Sustituirla perdería la banda de confianza, que
    es lo único que permite convertir un movimiento en veredicto.
    """
    from modules.brand_intel.ingest.pdf_pipeline import _LabelResolver
    from modules.brand_intel.models.models import (
        BrandEntity, BrandObservation, BrandWave)

    olas = {str(w.code): w for w in db.query(BrandWave).filter(
        BrandWave.engagement_id == engagement_id).all()}
    entidades = db.query(BrandEntity).filter(
        BrandEntity.engagement_id == engagement_id).all()
    marcas = {str(b.slug) for b in entidades}
    # El expediente ya declaró sus marcas con SUS slugs («wendys», «dominos_pizza»);
    # generar uno nuevo desde el rótulo del proveedor («Wendy's» → «wendy_s») crearía
    # una marca paralela y partiría la serie en dos. El resolvedor del módulo calza el
    # rótulo impreso contra las marcas declaradas antes de crear ninguna.
    resolver = _LabelResolver(entidades, list(olas.values()))
    existentes = {
        (str(o.wave_id), str(o.brand_slug), str(o.metric_code), str(o.segment))
        for o in db.query(BrandObservation).filter(
            BrandObservation.engagement_id == engagement_id).all()
    }

    olas_nuevas = marcas_nuevas = nuevas = omitidas = 0
    etiquetas_reparadas = 0
    orden = max((int(w.sort_order or 0) for w in olas.values()), default=0)
    for r in sorted(readings, key=lambda x: x.period):
        code = r.period.strftime("%Y-%m")
        wave = olas.get(code)
        if wave is None:
            orden += 1
            wave = BrandWave(engagement_id=engagement_id, code=code,
                             label=human_wave_label(r.period),
                             sort_order=orden, period_date=r.period)
            db.add(wave)
            db.flush()
            olas[code] = wave
            olas_nuevas += 1
        elif str(wave.label) == code:
            # Reparación acotada: una etiqueta IGUAL al código es el marcador que dejó una
            # carga anterior de esta misma matriz, no una elección de nadie. Se reemplaza
            # por el rótulo humano; una etiqueta que una persona escribió no se toca jamás.
            wave.label = human_wave_label(r.period)
            etiquetas_reparadas += 1
        slug = resolver.brand(r.brand) or r.brand_slug
        if slug not in marcas:
            db.add(BrandEntity(engagement_id=engagement_id, slug=slug,
                               name=r.brand, is_focal=False, in_category_set=True))
            db.flush()
            marcas.add(slug)
            marcas_nuevas += 1
        key = (str(wave.id), slug, r.metric_code, r.segment)
        if key in existentes:
            omitidas += 1
            continue
        db.add(BrandObservation(
            engagement_id=engagement_id, wave_id=wave.id, brand_slug=slug,
            metric_code=r.metric_code, segment=r.segment, value=r.value,
            base_n=None,          # la matriz no trae bases: no puntuable, y así se declara
            unit="pct", source=source))
        existentes.add(key)
        nuevas += 1

    # La CRONOLOGÍA gobierna el orden de la serie, no el orden de carga. Al incorporar
    # historia anterior a lo ya cargado, numerar desde el máximo existente dejaría 2018
    # después de 2026 y toda lectura de trayectoria —tendencia, pronóstico, ola previa—
    # se construiría sobre una secuencia falsa sin que nada fallara.
    db.flush()
    renumeradas = _renumber_waves(db, engagement_id)

    return {
        "olas_creadas": olas_nuevas, "marcas_creadas": marcas_nuevas,
        "etiquetas_reparadas": etiquetas_reparadas,
        "observaciones_nuevas": nuevas,
        "observaciones_omitidas_por_existir": omitidas,
        "olas_reordenadas": renumeradas,
    }


def _renumber_waves(db: Any, engagement_id: str) -> int:
    """Reasigna ``sort_order`` por fecha de período. Devuelve cuántas cambiaron.

    Las olas sin fecha —la de proyección, que existe solo para sostener un pronóstico
    congelado— quedan al final conservando su orden relativo.
    """
    from modules.brand_intel.models.models import BrandWave

    todas = db.query(BrandWave).filter(
        BrandWave.engagement_id == engagement_id).all()
    ordenadas = sorted(
        todas,
        key=lambda w: (w.period_date is None, w.period_date or date.max,
                       int(w.sort_order or 0)),
    )
    cambios = 0
    for i, w in enumerate(ordenadas, start=1):
        if int(w.sort_order or 0) != i:
            w.sort_order = i          # type: ignore[assignment]
            cambios += 1
    return cambios
