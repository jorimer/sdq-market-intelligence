"""La ingesta del tablero de ventas del operador.

El tablero llega como un libro de trabajo interno —hojas de presentación, tablas
dinámicas, una hoja «Base de Datos» con el detalle— y se lee SOLO de esa hoja de
detalle: las demás son vistas derivadas de ella, y leer un agregado ya calculado por
otro impide recomputarlo.

**Cero visión, cero modelo.** A diferencia de los mazos y los planes, aquí no hay nada
que interpretar: son columnas rotuladas. La ingesta es determinista de punta a punta, y
eso es deliberado — una cifra de venta mal leída contamina toda descomposición aguas
abajo y ningún gráfico posterior puede distinguirla de una buena.

**No se guarda el cheque promedio** aunque el tablero lo traiga: se computa al leer.
Almacenar un derivado junto a sus operandos garantiza que algún día dejen de cuadrar.

**La ciudad se normaliza.** El tablero real trae «Santo domingo» y «Santo Domingo» como
etiquetas distintas para la misma plaza —dieciocho locales repartidos entre las dos—, y
agregarlas por separado partiría en dos el mercado que concentra el volumen.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("sdq.brand_intel.sales")

SHEET = "Base de Datos"

#: Encabezado → campo del modelo. Las claves se comparan normalizadas (sin acentos, sin
#: puntuación, minúsculas), porque el tablero varía la puntuación entre entregas
#: («A/C  Comparativo %» con doble espacio, «Gcs D.T 2025» con y sin punto).
_COLUMNS_RAW: Dict[str, str] = {
    "id tienda": "store_id",
    "restaurantes": "store_name",
    "ciudad": "city",
    "zona": "zone",
    "fecha": "business_date",
    "ventas 2026": "sales",
    "venta 2025": "sales_prior",
    "gcs 2026": "transactions",
    "gcs 2025": "transactions_prior",
    "ventas dt 2026": "sales_drive_thru",
    "ventas dt 2025": "sales_drive_thru_prior",
    "gcs dt 2026": "transactions_drive_thru",
    "gcs dt 2025": "transactions_drive_thru_prior",
    "ventas delivery 2026": "sales_delivery",
    "ventas delivery 2025": "sales_delivery_prior",
    "gcs delivery 2026": "transactions_delivery",
    "gcs delivery 2025": "transactions_delivery_prior",
    "ventas mostrador 2026": "sales_counter",
    "ventas mostrador 2025": "sales_counter_prior",
    "gcs mostrador 2026": "transactions_counter",
    "gcs mostrador 2025": "transactions_counter_prior",
}

#: Indexado por ``_key`` para absorber la puntuación y los espacios del tablero.
_COLUMNS: Dict[str, str] = {k.replace(" ", ""): v for k, v in _COLUMNS_RAW.items()}

_INT_FIELDS = {f for f in _COLUMNS.values() if f.startswith("transactions")}
_REQUIRED = {"store_id", "business_date", "sales"}


def _norm(s: Any) -> str:
    """Etiqueta comparable: sin acentos, sin puntuación, un solo espacio, minúsculas."""
    txt = unicodedata.normalize("NFKD", str(s or ""))
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = re.sub(r"[^0-9a-zA-Z ]+", " ", txt)
    return re.sub(r"\s+", " ", txt).strip().lower()


def _key(s: Any) -> str:
    """Clave de encabezado: ``_norm`` SIN espacios.

    El tablero escribe el canal de autoservicio como «D.T», que al quitar la puntuación
    queda «d t» y no calza con «dt»; y alterna espacios dobles entre entregas («A/C
    Comparativo  %»). Comparar sin espacios absorbe ambas variantes — el defecto real:
    con la comparación por espacios, las cuatro columnas de Automac entraban como «sin
    mapear» y el canal quedaba en NULL sin que nada fallara.
    """
    return _norm(s).replace(" ", "")


def normalize_city(raw: Any) -> Optional[str]:
    """La plaza en forma canónica. Une las variantes de capitalización del tablero."""
    n = _norm(raw)
    if not n:
        return None
    canon = {
        "santo domingo": "Santo Domingo",
        "santiago": "Santiago",
        "puerto plata": "Puerto Plata",
        "san francisco": "San Francisco",
        "la vega": "La Vega",
    }
    if n in canon:
        return canon[n]
    # Plaza no vista antes: se conserva con capitalización de título en vez de
    # descartarse — un local nuevo no puede desaparecer de la serie en silencio.
    return " ".join(w.capitalize() for w in n.split())


def _parse_date(raw: Any) -> Optional[date]:
    """La fecha de la jornada. El tablero la entrega como «2026-06-01, lunes»."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    txt = str(raw or "").strip()
    if not txt:
        return None
    head = txt.split(",")[0].strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(head, fmt).date()
        except ValueError:
            continue
    return None


def _num(raw: Any) -> Optional[float]:
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    txt = str(raw or "").strip().replace(",", "")
    if not txt or txt in {"-", "n/d", "N/D"}:
        return None
    try:
        return float(txt)
    except ValueError:
        return None


@dataclass
class SalesIngestReport:
    """Qué se leyó y qué se descartó. Sin esto, una ingesta silenciosa parece completa."""

    rows_read: int = 0
    rows_stored: int = 0
    rows_skipped: int = 0
    skip_reasons: Optional[Dict[str, int]] = None
    stores: int = 0
    cities: Optional[List[str]] = None
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    unmapped_columns: Optional[List[str]] = None
    ignored_columns: Optional[Dict[str, str]] = None

    def as_dict(self) -> Dict[str, Any]:
        return {
            "filas_leidas": self.rows_read,
            "filas_guardadas": self.rows_stored,
            "filas_descartadas": self.rows_skipped,
            "motivos_descarte": self.skip_reasons or {},
            "locales": self.stores,
            "plazas": self.cities or [],
            "desde": self.date_from.isoformat() if self.date_from else None,
            "hasta": self.date_to.isoformat() if self.date_to else None,
            "columnas_sin_mapear": self.unmapped_columns or [],
            "columnas_ignoradas": self.ignored_columns or {},
        }


#: Columnas que el tablero trae y esta ingesta NO toma A PROPÓSITO, con el motivo. La
#: distinción importa: «columna que ignoramos por diseño» y «columna que no reconocimos»
#: son cosas distintas, y meterlas en la misma lista mata la señal — el tablero real trae
#: 27 columnas derivadas, así que un aviso indiferenciado sepulta el único caso que
#: interesa (fue así como cuatro columnas de Automac entraron vacías sin que nada avisara).
_IGNORED_EXACT: Dict[str, str] = {
    "pais": "metadato, no una medida",
    "consultor de operaciones": "metadato, no una medida",
    "mes": "se deriva de la fecha",
    "ano": "se deriva de la fecha",
    "a c 2025": "cheque promedio: se recomputa como ventas÷transacciones",
    "a c 2026": "cheque promedio: se recomputa como ventas÷transacciones",
}

#: Fragmentos que delatan una columna YA CALCULADA por el operador. No entran porque la
#: plataforma las recomputa desde los absolutos: un porcentaje ajeno no se puede auditar.
_IGNORED_PATTERNS: Tuple[Tuple[str, str], ...] = (
    ("proyeccion", "proyección del operador, no un dato observado"),
    ("proy", "proyección del operador, no un dato observado"),
    ("comparativ", "variación ya calculada: se recomputa desde los absolutos"),
    ("diferencia", "diferencia ya calculada: se recomputa desde los absolutos"),
    ("dif", "diferencia ya calculada: se recomputa desde los absolutos"),
)


def classify_unmapped(labels: List[str]) -> Tuple[List[str], Dict[str, str]]:
    """Parte los rótulos no tomados en (NO RECONOCIDOS, ignorados con su motivo).

    Solo la primera lista es una alarma: es la que dice «esta columna parecía un dato y se
    quedó afuera».
    """
    desconocidas: List[str] = []
    ignoradas: Dict[str, str] = {}
    for raw in labels:
        k = _norm(raw)
        motivo = _IGNORED_EXACT.get(k)
        if motivo is None:
            motivo = next((m for frag, m in _IGNORED_PATTERNS if frag in k), None)
        if motivo:
            ignoradas[str(raw).strip()] = motivo
        else:
            desconocidas.append(str(raw).strip())
    return desconocidas, ignoradas


def _find_header(ws) -> Tuple[int, Dict[int, str], List[str]]:
    """Fila de encabezados, mapa columna→campo y los rótulos que no se reconocieron.

    El encabezado no está en la primera fila (el tablero abre con filas de título), así
    que se busca la primera fila que contenga los rótulos obligatorios en vez de
    asumir una posición — una entrega que agregue una fila de título arriba no debe
    romper la ingesta.
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        mapped = {i: _COLUMNS[_key(c)] for i, c in enumerate(row)
                  if c is not None and _key(c) in _COLUMNS}
        if _REQUIRED.issubset(set(mapped.values())):
            unmapped = [str(c) for i, c in enumerate(row)
                        if c is not None and i not in mapped and str(c).strip()]
            return idx, mapped, unmapped
    raise ValueError(
        "No se encontró la fila de encabezados del tablero: se esperan al menos las "
        "columnas 'ID Tienda', 'Fecha' y 'Ventas <año>' en la hoja «Base de Datos»."
    )


def read_sales_workbook(content: bytes) -> Tuple[List[Dict[str, Any]], SalesIngestReport]:
    """Lee el tablero y devuelve las filas normalizadas con su informe de ingesta."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            raise ValueError(
                f"El libro no trae la hoja «{SHEET}». Las hojas presentes son: "
                f"{', '.join(wb.sheetnames)}. Se requiere la hoja de detalle, no un "
                "agregado ya calculado."
            )
        ws = wb[SHEET]
        header_row, colmap, unmapped = _find_header(ws)

        rows: List[Dict[str, Any]] = []
        skips: Dict[str, int] = {}
        read = 0
        for raw in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(c is not None for c in raw):
                continue
            read += 1
            rec: Dict[str, Any] = {}
            for i, field in colmap.items():
                if i >= len(raw):
                    continue
                val = raw[i]
                if field == "business_date":
                    rec[field] = _parse_date(val)
                elif field in ("store_id", "store_name", "zone"):
                    rec[field] = str(val).strip() if val is not None else None
                elif field == "city":
                    rec[field] = normalize_city(val)
                else:
                    n = _num(val)
                    rec[field] = (int(round(n)) if n is not None and field in _INT_FIELDS
                                  else n)
            # Una fila sin local, sin fecha o sin venta no es una jornada: se descarta
            # con su motivo en vez de entrar con huecos que el motor tomaría por ceros.
            missing = [f for f in ("store_id", "business_date") if not rec.get(f)]
            if missing:
                skips["sin " + " ni ".join(missing)] = skips.get(
                    "sin " + " ni ".join(missing), 0) + 1
                continue
            if rec.get("sales") is None and rec.get("transactions") is None:
                skips["sin venta ni transacciones"] = skips.get(
                    "sin venta ni transacciones", 0) + 1
                continue
            rows.append(rec)

        desconocidas, ignoradas = classify_unmapped(unmapped)
        fechas = [r["business_date"] for r in rows if r.get("business_date")]
        rep = SalesIngestReport(
            rows_read=read, rows_stored=len(rows), rows_skipped=read - len(rows),
            skip_reasons=skips or None,
            stores=len({r["store_id"] for r in rows}),
            cities=sorted({r["city"] for r in rows if r.get("city")}),
            date_from=min(fechas) if fechas else None,
            date_to=max(fechas) if fechas else None,
            unmapped_columns=desconocidas or None,
            ignored_columns=ignoradas or None,
        )
        if desconocidas:
            # WARNING, no info: una columna que parecía un dato quedó afuera y nadie la
            # pidió — es exactamente la forma del defecto de «Ventas D.T».
            logger.warning("Tablero con %d columna(s) NO RECONOCIDA(s): %s",
                           len(desconocidas), desconocidas[:8])
        if ignoradas:
            logger.info("Tablero con %d columna(s) ignoradas por diseño", len(ignoradas))
        return rows, rep
    finally:
        wb.close()


def store_sales(db: Any, engagement_id: str, rows: List[Dict[str, Any]],
                source_file: Optional[str] = None) -> Dict[str, Any]:
    """Guarda las jornadas, actualizando la que ya exista para ese local y fecha.

    Idempotente por (encargo, fecha, local): el operador entrega el tablero varias veces
    y las jornadas se repiten entre entregas, a veces corregidas. La última entrega
    manda sobre la jornada; las jornadas que una entrega no cubre no se tocan.
    """
    from modules.brand_intel.models.models import BrandSalesDaily

    existentes = {
        (r.business_date, r.store_id): r
        for r in db.query(BrandSalesDaily).filter(
            BrandSalesDaily.engagement_id == engagement_id).all()
    }
    nuevas = actualizadas = 0
    for rec in rows:
        key = (rec["business_date"], rec["store_id"])
        row = existentes.get(key)
        if row is None:
            row = BrandSalesDaily(engagement_id=engagement_id,
                                  source_file=source_file, **rec)
            db.add(row)
            existentes[key] = row
            nuevas += 1
        else:
            for field, val in rec.items():
                setattr(row, field, val)
            setattr(row, "source_file", source_file)
            actualizadas += 1
    return {"jornadas_nuevas": nuevas, "jornadas_actualizadas": actualizadas}
