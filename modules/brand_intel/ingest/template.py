"""The Excel template handed to the research provider.

Deliberately a *long* (tidy) layout: one row per measurement, rather than a grid of
metrics as columns. A grid would need a new column — and a schema change — for every
indicator a client adds. The long shape absorbs any tracker without touching the model,
which is what lets a second client onboard by filling a spreadsheet.

Five sheets:
  Encargo         — the mandate's identity.
  Olas            — one row per wave, with fieldwork window and nominal base.
  Marcas          — the competitive set, and which brands count toward the denominator.
  Observaciones   — the data. One row per wave x brand x metric x segment.
  Diccionario     — the metric vocabulary, so whoever fills it in has the codes at hand.

``base_n`` is a first-class column, not an afterthought: it is what makes every confidence
band computable. A row without it is loaded and displayed but can never sustain a verdict.
"""
from __future__ import annotations

import io
from typing import Any, Dict, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.brand_intel.engines.metrics import METRICS, core_metrics

_NAVY = "0B1F3A"
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_NOTE_FONT = Font(italic=True, size=9, color="6B7A8F")


def _write_header(ws, headers, widths) -> None:
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_template(
    engagement_slug: str = "",
    focal_brand: str = "",
    client_name: str = "",
    provider: str = "",
    waves: Optional[Sequence[Dict[str, Any]]] = None,
    brands: Optional[Sequence[Dict[str, Any]]] = None,
    request_wave: Optional[Dict[str, Any]] = None,
    segments: Sequence[str] = ("total",),
    request_rows: Optional[Sequence[Dict[str, Any]]] = None,
) -> bytes:
    """El libro .xlsx, y con ``request_wave`` una PLANILLA DE PEDIDO ya armada.

    La diferencia no es cosmética. En blanco, quien la recibe —un analista del proveedor,
    que no conoce nuestros slugs ni nuestros códigos— tiene que construir cada fila: eso no
    se llena, o se llena mal. Con las filas ya generadas para la ola pedida, solo faltan
    ``valor`` y ``base_n``, y llenarla es pegar dos columnas.

    Es además la ruta que históricamente FUNCIONÓ: la única entrega que llegó a
    observaciones confirmadas vino de esta plantilla, no de leer el mazo con visión.
    """
    wb = Workbook()

    # ── Encargo ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Encargo"
    _write_header(ws, ["campo", "valor", "nota"], [26, 40, 62])
    rows = [
        ("slug", engagement_slug, "Identificador estable del encargo. Ej: mcdonalds-rd"),
        ("cliente", client_name, "Quién contrata el estudio."),
        ("marca_focal", focal_brand, "La marca sobre la que trata el informe."),
        ("mercado", "República Dominicana", "País o mercado del estudio."),
        ("categoria", "", "Categoría competitiva. Ej: QSR, banca minorista."),
        ("proveedor", provider, "Proveedor de investigación. Ej: Ipsos Dominicana."),
    ]
    for r, (k, v, note) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=3, value=note).font = _NOTE_FONT

    # ── Olas ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Olas")
    _write_header(
        ws,
        ["codigo", "etiqueta", "orden", "fecha_referencia",
         "campo_inicio", "campo_fin", "base_nominal"],
        [16, 16, 8, 20, 16, 16, 14],
    )
    if waves:
        # Las olas REALES del encargo: quien recibe la planilla no tiene que adivinar cómo
        # las nombramos, y una ola escrita distinto es lo que rompe el emparejamiento.
        for r, w in enumerate(waves, start=2):
            ws.cell(row=r, column=1, value=w.get("code"))
            ws.cell(row=r, column=2, value=w.get("label"))
            ws.cell(row=r, column=3, value=r - 1)
            ws.cell(row=r, column=7, value=w.get("base"))
        ws.cell(row=len(waves) + 3, column=1,
                value="(ya cargadas — no hace falta tocarlas)").font = _NOTE_FONT
    else:
        ws.cell(row=2, column=1, value="2025-05")
        ws.cell(row=2, column=2, value="May '25")
        ws.cell(row=2, column=3, value=1)
        ws.cell(row=2, column=7, value=300)
        ws.cell(row=3, column=1,
                value="(una fila por ola, en orden cronológico)").font = _NOTE_FONT

    # ── Marcas ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Marcas")
    _write_header(ws, ["slug", "nombre", "es_focal", "en_set_categoria", "orden"],
                  [22, 28, 12, 20, 8])
    if brands:
        for r, b in enumerate(brands, start=2):
            ws.cell(row=r, column=1, value=b.get("slug"))
            ws.cell(row=r, column=2, value=b.get("name"))
            ws.cell(row=r, column=3, value="SI" if b.get("is_focal") else "NO")
            ws.cell(row=r, column=4, value="SI")
            ws.cell(row=r, column=5, value=r - 1)
    else:
        ws.cell(row=2, column=1, value="mcdonalds")
        ws.cell(row=2, column=2, value="McDonald's")
        ws.cell(row=2, column=3, value="SI")
        ws.cell(row=2, column=4, value="SI")
    if not brands:
        ws.cell(row=2, column=5, value=1)
    # La nota va DEBAJO de las marcas reales: en la fila 3 pisaría la segunda, y el
    # archivo saldría a un tercero con una marca menos sin que nada avisara.
    ws.cell(row=(len(brands) + 3) if brands else 3, column=1,
            value="(en_set_categoria = NO para marcas medidas "
                  "pero fuera de la categoría)").font = _NOTE_FONT

    # ── Observaciones ──────────────────────────────────────────────
    ws = wb.create_sheet("Observaciones")
    _write_header(
        ws,
        ["entrega", "ola", "marca", "metrica", "segmento", "atributo", "valor",
         "base_n", "unidad", "fuente"],
        [22, 14, 20, 26, 20, 28, 12, 10, 10, 34],
    )
    if request_wave:
        # PLANILLA DE PEDIDO: una fila por celda, con todo resuelto salvo las dos columnas
        # que solo el proveedor puede llenar.
        #
        # `request_rows` es la lista EXACTA de coordenadas a pedir, derivada de lo que ese
        # proveedor ya publicó en entregas anteriores. Importa: el cruce genérico —todas
        # las marcas × todas las métricas× todos los cortes— pide celdas que el estudio no
        # mide (la fidelidad se publica para UNA marca y le pediríamos dieciocho), y una
        # fila vacía frente a alguien con prisa es una invitación a completarla con algo.
        # Sin la lista, se cae al cruce genérico del núcleo, que sirve para un encargo
        # nuevo del que todavía no sabemos qué publica.
        fila = 2
        etiqueta = str(request_wave.get("label") or request_wave.get("code") or "")
        if request_rows:
            pedido_filas = [
                (r.get("brand_slug"), r.get("metric_code"), r.get("segment") or "total",
                 r.get("attribute") or "")
                for r in request_rows
            ]
        else:
            pedido_filas = [
                (b.get("slug"), m.code, corte, "")
                for b in (brands or []) for m in core_metrics()
                for corte in (segments or ("total",))
            ]
        for slug, metrica, corte, atributo in pedido_filas:
            ws.cell(row=fila, column=1, value=etiqueta)
            ws.cell(row=fila, column=2, value=request_wave.get("code"))
            ws.cell(row=fila, column=3, value=slug)
            ws.cell(row=fila, column=4, value=metrica)
            ws.cell(row=fila, column=5, value=corte)
            ws.cell(row=fila, column=6, value=atributo)
            ws.cell(row=fila, column=9, value="pct")
            fila += 1
        nota_fila = fila + 1
    else:
        nota_fila = 3
        ws.cell(row=2, column=1, value="Ola 4 · mar\'26")
        ws.cell(row=2, column=2, value="2025-05")
        ws.cell(row=2, column=3, value="mcdonalds")
        ws.cell(row=2, column=4, value="reach_7d")
        ws.cell(row=2, column=5, value="total")
        ws.cell(row=2, column=6, value="")      # atributo: vacío salvo métrica por atributo
        ws.cell(row=2, column=7, value=26)
        ws.cell(row=2, column=8, value=300)
        ws.cell(row=2, column=9, value="pct")
        ws.cell(row=2, column=10, value="Hot Tracker · lámina 18")
    # Las notas van DEBAJO de las filas, no en la 3: con la planilla pre-llenada
    # escribirían encima de dos celdas de datos y nadie lo notaría hasta la carga.
    if request_wave:
        ws.cell(row=nota_fila, column=1,
                value="↑ Solo faltan «valor» y «base_n». Todo lo demás ya está resuelto: "
                      "no hace falta conocer nuestros códigos.").font = _NOTE_FONT
        nota_fila += 1
    ws.cell(row=nota_fila, column=1,
            value="entrega = de qué informe salió la cifra. Un tracker reexpone sus olas "
                  "anteriores en cada entrega y a veces las corrige: sin esta columna, "
                  "cargar dos informes en un mismo libro funde lo que dijo cada uno y la "
                  "cifra vigente pasa a depender del orden de las filas. "
                  "Vacío = todo el libro es una sola entrega.").font = _NOTE_FONT
    ws.cell(row=nota_fila + 1, column=1,
            value="marca vacía = métrica de categoría · base_n vacío = el dato se muestra "
                  "pero no puede sostener un veredicto").font = _NOTE_FONT
    ws.cell(row=nota_fila + 2, column=1,
            value="atributo = para las métricas que se miden POR atributo, enunciado o "
                  "escala (índice de atributo, evaluación de delivery, rango de gasto): "
                  "UNA FILA POR CADA UNO. Sin esa columna, todos caen en la misma celda y "
                  "se pisan entre sí.").font = _NOTE_FONT

    # ── Diccionario ────────────────────────────────────────────────
    ws = wb.create_sheet("Diccionario")
    _write_header(ws, ["metrica", "etiqueta", "tipo", "admite_banda", "nota"],
                  [26, 34, 14, 15, 60])
    for r, m in enumerate(METRICS, start=2):
        ws.cell(row=r, column=1, value=m.code)
        ws.cell(row=r, column=2, value=m.label)
        ws.cell(row=r, column=3, value=m.kind)
        ws.cell(row=r, column=4, value="SI" if m.supports_bands else "NO")
        ws.cell(row=r, column=5, value=m.description).font = _NOTE_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def template_filename(engagement_slug: Optional[str] = None) -> str:
    base = engagement_slug or "encargo"
    return f"SDQ-MIP_plantilla_tracker_{base}.xlsx"
