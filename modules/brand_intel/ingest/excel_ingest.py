"""Reader for the provider's workbook.

Two principles, both load-bearing:

* **Reject, never coerce.** A row that cannot be understood is reported with its sheet, row
  number and reason — not silently dropped and not filled with a default. The ingest report
  is part of the deliverable: the client sees exactly what landed and what did not.
* **Idempotent.** Re-uploading a corrected workbook updates in place instead of duplicating.
  Corrections are normal in tracker work, so the natural motion must be safe.

Unknown metric codes are rejected rather than stored. A typo that silently created a new
metric would quietly split a series in two and no chart would show the break.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from modules.brand_intel.engines.metrics import TRACKER_VOCABULARY, Vocabulary
from modules.brand_intel.models.models import (
    BrandEngagement,
    BrandEntity,
    BrandObservation,
    BrandWave,
)

_TRUE = {"si", "sí", "s", "yes", "y", "true", "1", "verdadero"}


@dataclass
class IngestReport:
    """What landed, what did not, and why."""

    waves_created: int = 0
    waves_updated: int = 0
    brands_created: int = 0
    brands_updated: int = 0
    observations_created: int = 0
    observations_updated: int = 0
    rejected: List[Dict[str, Any]] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    def reject(self, sheet: str, row: int, reason: str) -> None:
        self.rejected.append({"sheet": sheet, "row": row, "reason": reason})

    def as_dict(self) -> Dict[str, Any]:
        return {
            "olas": {"creadas": self.waves_created, "actualizadas": self.waves_updated},
            "marcas": {"creadas": self.brands_created, "actualizadas": self.brands_updated},
            "observaciones": {
                "creadas": self.observations_created,
                "actualizadas": self.observations_updated,
            },
            "rechazadas": self.rejected,
            "advertencias": self.warnings,
            "total_rechazadas": len(self.rejected),
        }


def _norm(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _as_bool(v: Any, default: bool = False) -> bool:
    s = _norm(v)
    return default if s is None else s.lower() in _TRUE


def _as_int(v: Any) -> Optional[int]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _as_float(v: Any) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _as_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _rows(ws, min_row: int = 2):
    """Yield (row_number, values) skipping blank and note rows."""
    for idx, row in enumerate(ws.iter_rows(min_row=min_row, values_only=True), start=min_row):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        first = _norm(row[0])
        if first and first.startswith("("):     # the template's inline notes
            continue
        yield idx, row


def ingest_workbook(
    db: Session, engagement: BrandEngagement, content: bytes,
    document_name: str = "libro.xlsx",
) -> IngestReport:
    """Load a provider workbook into an existing engagement."""
    report = IngestReport()
    wb = load_workbook(io.BytesIO(content), data_only=True)

    for sheet in ("Olas", "Marcas", "Observaciones"):
        if sheet not in wb.sheetnames:
            report.reject(sheet, 0, "Hoja ausente en el archivo.")
    if report.rejected:
        return report

    _ingest_waves(db, engagement, wb["Olas"], report)
    _ingest_brands(db, engagement, wb["Marcas"], report)
    db.flush()
    from modules.brand_intel import service as svc

    _ingest_observations(db, engagement, wb["Observaciones"], report,
                         svc.vocabulary_for(db, str(engagement.id)),
                         document_name=document_name)
    return report


def _ingest_waves(db, engagement, ws, report: IngestReport) -> None:
    for rownum, row in _rows(ws):
        code = _norm(row[0])
        if not code:
            report.reject("Olas", rownum, "Falta el código de la ola.")
            continue
        label = _norm(row[1]) or code
        order = _as_int(row[2])
        if order is None:
            report.reject("Olas", rownum, "Falta el orden cronológico de la ola.")
            continue

        existing = (
            db.query(BrandWave)
            .filter(BrandWave.engagement_id == engagement.id, BrandWave.code == code)
            .first()
        )
        target = existing or BrandWave(engagement_id=engagement.id, code=code)
        target.label = label
        target.sort_order = order
        target.period_date = _as_date(row[3])
        target.field_start = _as_date(row[4]) if len(row) > 4 else None
        target.field_end = _as_date(row[5]) if len(row) > 5 else None
        target.nominal_base = _as_int(row[6]) if len(row) > 6 else None

        if target.period_date is None:
            report.warnings.append(
                f"Ola '{code}' sin fecha de referencia: no podrá deflactarse."
            )
        if existing:
            report.waves_updated += 1
        else:
            db.add(target)
            report.waves_created += 1


def _ingest_brands(db, engagement, ws, report: IngestReport) -> None:
    for rownum, row in _rows(ws):
        slug = _norm(row[0])
        if not slug:
            report.reject("Marcas", rownum, "Falta el slug de la marca.")
            continue
        existing = (
            db.query(BrandEntity)
            .filter(BrandEntity.engagement_id == engagement.id, BrandEntity.slug == slug)
            .first()
        )
        target = existing or BrandEntity(engagement_id=engagement.id, slug=slug)
        target.name = _norm(row[1]) or slug
        target.is_focal = _as_bool(row[2])
        target.in_category_set = _as_bool(row[3], default=True)
        target.sort_order = _as_int(row[4]) or 0
        if existing:
            report.brands_updated += 1
        else:
            db.add(target)
            report.brands_created += 1


def _ingest_observations(db, engagement, ws, report: IngestReport,
                         vocab: Vocabulary = TRACKER_VOCABULARY,
                         document_name: str = "libro.xlsx") -> None:
    """Carga las observaciones agrupándolas por ENTREGA.

    Un tracker reexpone sus olas anteriores en cada entrega y a veces las corrige. Antes
    esta ruta escribía la observación directamente, sin registrar de qué entrega venía: dos
    informes en un mismo libro se fundían, y cuál cifra quedaba vigente dependía del orden
    de las filas. La columna `entrega` separa lo que dijo cada una, y la promoción pasa por
    el mismo camino que la ruta de PDF —``promote_readings``— para que la regla de
    precedencia sea una sola y no dos que acaban divergiendo.
    """
    from modules.brand_intel.ingest.pdf_pipeline import Reading, promote_readings
    from modules.brand_intel.models.models import BrandExtraction

    waves = {
        w.code: w.id
        for w in db.query(BrandWave).filter(BrandWave.engagement_id == engagement.id).all()
    }
    brands = {
        b.slug
        for b in db.query(BrandEntity).filter(BrandEntity.engagement_id == engagement.id).all()
    }

    por_entrega: Dict[str, List[Any]] = {}
    for rownum, row in _rows(ws):
        entrega = (_norm(row[0]) if len(row) > 0 else "") or document_name
        wave_code = _norm(row[1]) if len(row) > 1 else ""
        wave_id = waves.get(wave_code or "")
        if not wave_id:
            report.reject("Observaciones", rownum,
                          f"Ola desconocida: '{wave_code}'. Debe existir en la hoja Olas.")
            continue

        brand = (_norm(row[2]) if len(row) > 2 else "") or ""
        if brand and brand not in brands:
            report.reject("Observaciones", rownum,
                          f"Marca desconocida: '{brand}'. Debe existir en la hoja Marcas.")
            continue

        metric = (_norm(row[3]) if len(row) > 3 else "") or ""
        if not metric or vocab.get(metric) is None:
            report.reject("Observaciones", rownum,
                          f"Métrica desconocida: '{metric}'. Ver la hoja Diccionario.")
            continue

        value = _as_float(row[5]) if len(row) > 5 else None
        if value is None:
            report.reject("Observaciones", rownum, "Valor ausente o no numérico.")
            continue

        por_entrega.setdefault(entrega, []).append(Reading(
            wave_id=str(wave_id),
            brand_slug=brand or None,
            metric_code=metric,
            segment=(_norm(row[4]) if len(row) > 4 else "") or "total",
            value=value,
            base_n=_as_int(row[6]) if len(row) > 6 else None,
            unit=(_norm(row[7]) if len(row) > 7 else "") or "pct",
            source=(_norm(row[8]) if len(row) > 8 else "") or entrega,
        ))

    # Una fila de extracción por entrega: es lo que ata cada lectura a su informe y lo que
    # permite después responder "¿esta cifra cambió, y qué entrega la cambió?".
    for entrega, lecturas in por_entrega.items():
        extraction = (
            db.query(BrandExtraction)
            .filter(BrandExtraction.engagement_id == engagement.id,
                    BrandExtraction.document_name == entrega,
                    BrandExtraction.method == "excel")
            .first()
        )
        if extraction is None:
            extraction = BrandExtraction(
                engagement_id=engagement.id, document_name=entrega, method="excel",
                status="confirmed",
            )
            db.add(extraction)
        db.flush()
        out = promote_readings(db, str(engagement.id), extraction, lecturas)
        report.observations_created += int(out["creadas"])
        report.observations_updated += int(out["actualizadas"])
        for d in out["discrepancias"]:
            report.warnings.append(
                f"«{d['marca']} · {d['metrica']} · {d['ola']}» viene con dos cifras "
                f"distintas en la entrega '{entrega}': no se promueve ninguna.")
        for c in out["cifras_que_cambian"]:
            if "vigente" in c:
                report.warnings.append(
                    f"«{c['marca']} · {c['metrica']} · {c['ola']}»: la entrega '{entrega}' "
                    f"dice {c['esta_entrega']}, pero manda una entrega posterior "
                    f"({c['vigente']}). Queda en el registro.")
