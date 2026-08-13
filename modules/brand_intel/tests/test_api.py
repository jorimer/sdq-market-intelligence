"""API tests, focused on the isolation boundary.

This is the platform's only module serving private per-client data, so the access rules
are correctness properties. In particular: a caller from another organization must get a
**404, not a 403** — the existence of another client's engagement is itself private, and a
403 would confirm it.
"""
import io

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from shared.auth.dependencies import get_current_user
from shared.auth.models import UserRole
from shared.database.session import get_db
from modules.brand_intel.api.router import router
from modules.brand_intel.models.models import BrandEngagement


def _client(db, role=UserRole.admin, org=None):
    app = FastAPI()
    app.include_router(router, prefix="/api/v1/brand-intel")

    class _U:
        def __init__(self, r, o):
            self.role = r
            self.organization_id = o
            self.email = "analista@sdq.test"      # destructive routes log the actor

    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: _U(role, org)
    return TestClient(app)


# ── isolation ─────────────────────────────────────────────────────────

def test_other_organization_gets_404_not_403(db, engagement):
    """A 403 would confirm the engagement exists. It must be indistinguishable from absent."""
    engagement.organization_id = "org-1"
    db.commit()
    r = _client(db, role=UserRole.viewer, org="org-2").get(
        "/api/v1/brand-intel/engagements/demo")
    assert r.status_code == 404


def test_owning_organization_can_read(db, engagement):
    engagement.organization_id = "org-1"
    db.commit()
    r = _client(db, role=UserRole.viewer, org="org-1").get(
        "/api/v1/brand-intel/engagements/demo")
    assert r.status_code == 200
    assert r.json()["focal_brand"] == "Focal"


def test_staff_can_read_any_engagement(db, engagement):
    engagement.organization_id = "org-1"
    db.commit()
    r = _client(db, role=UserRole.admin, org=None).get(
        "/api/v1/brand-intel/engagements/demo")
    assert r.status_code == 200


def test_unknown_engagement_is_404(db, engagement):
    assert _client(db).get("/api/v1/brand-intel/engagements/fantasma").status_code == 404


def test_list_only_returns_visible_engagements(db, engagement):
    engagement.organization_id = "org-1"
    db.add(BrandEngagement(slug="otro", client_name="Otro", focal_brand="X",
                           market="RD", organization_id="org-2"))
    db.commit()
    body = _client(db, role=UserRole.viewer, org="org-1").get(
        "/api/v1/brand-intel/engagements").json()
    assert [e["slug"] for e in body] == ["demo"]


# ── engagements ───────────────────────────────────────────────────────

def test_create_engagement_and_reject_duplicate_slug(db):
    c = _client(db)
    payload = {"slug": "nuevo", "client_name": "Cliente", "focal_brand": "Marca"}
    assert c.post("/api/v1/brand-intel/engagements", json=payload).status_code == 201
    assert c.post("/api/v1/brand-intel/engagements", json=payload).status_code == 409


# ── template and ingest ───────────────────────────────────────────────

def test_template_downloads_as_xlsx(db, engagement):
    r = _client(db).get("/api/v1/brand-intel/template.xlsx?engagement=demo")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"          # xlsx is a zip container


def test_ingest_rejects_a_non_xlsx_upload(db, engagement):
    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/ingest",
        files={"file": ("datos.csv", b"a,b,c", "text/csv")})
    assert r.status_code == 400


def test_ingest_accepts_a_valid_workbook(db, engagement):
    wb = Workbook()
    ws = wb.active
    ws.title = "Olas"
    ws.append(["codigo", "etiqueta", "orden", "fecha_referencia",
               "campo_inicio", "campo_fin", "base_nominal"])
    ws.append(["w4", "Ola 4", 4, "2026-07-01", None, None, 300])
    ws = wb.create_sheet("Marcas")
    ws.append(["slug", "nombre", "es_focal", "en_set_categoria", "orden"])
    ws.append(["focal", "Focal", "SI", "SI", 1])
    ws = wb.create_sheet("Observaciones")
    ws.append(["entrega", "ola", "marca", "metrica", "segmento", "valor", "base_n",
               "unidad", "fuente"])
    ws.append(["Ola 4", "w4", "focal", "reach_7d", "total", 50, 300, "pct", "test"])
    buf = io.BytesIO()
    wb.save(buf)

    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/ingest",
        files={"file": ("datos.xlsx", buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    body = r.json()
    assert body["olas"]["creadas"] == 1
    assert body["observaciones"]["creadas"] == 1
    assert body["total_rechazadas"] == 0


# ── analysis endpoints ────────────────────────────────────────────────

@pytest.mark.parametrize("path", [
    "category", "attribution", "funnel", "ticket", "signal-filter",
    "forecast/backtest", "forecast/track-record", "decisions",
    "scenarios", "vigilance",
])
def test_analysis_endpoints_respond(db, engagement, path):
    r = _client(db).get(f"/api/v1/brand-intel/engagements/demo/{path}")
    assert r.status_code == 200
    assert isinstance(r.json(), dict)


def test_category_endpoint_returns_the_share_series(db, engagement):
    body = _client(db).get("/api/v1/brand-intel/engagements/demo/category").json()
    assert body["available"] is True
    assert {b["brand"] for b in body["share"]} == {"focal", "rival"}


# ── decisions ─────────────────────────────────────────────────────────

def test_decision_check_reports_infeasibility(db, engagement):
    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/decisions/check",
        json={"metric_code": "reach_7d", "baseline_wave_code": "w2",
              "brand_slug": "focal", "success_threshold": 1.0})
    assert r.status_code == 200
    assert r.json()["feasible"] is False


def test_decision_check_rejects_an_unknown_wave(db, engagement):
    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/decisions/check",
        json={"metric_code": "reach_7d", "baseline_wave_code": "w9",
              "success_threshold": 5.0})
    assert r.status_code == 400


def test_unevaluable_decision_is_recorded_with_its_reason(db, engagement):
    """Recorded, not silently dropped: it has to be redesignable rather than re-proposed."""
    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/decisions",
        json={"title": "Movimiento minúsculo", "metric_code": "reach_7d",
              "baseline_wave_code": "w2", "brand_slug": "focal",
              "target_wave_code": "w3", "success_threshold": 0.5})
    assert r.status_code == 201
    body = r.json()
    assert body["status"] == "unevaluable"
    assert "inevaluable" in body["feasibility"]["reason"]


def test_feasible_decision_is_created_open(db, engagement):
    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/decisions",
        json={"title": "Subir alcance 10 puntos", "metric_code": "reach_7d",
              "baseline_wave_code": "w2", "brand_slug": "focal",
              "target_wave_code": "w3", "success_threshold": 10.0,
              "owner": "Marketing"})
    assert r.status_code == 201
    assert r.json()["status"] == "open"


# ── forecast and report ───────────────────────────────────────────────

def test_issue_forecast_rejects_unknown_wave(db, engagement):
    r = _client(db).post("/api/v1/brand-intel/engagements/demo/forecast/issue?wave=w9")
    assert r.status_code == 400


def test_report_json_and_html(db, engagement):
    c = _client(db)
    body = c.get("/api/v1/brand-intel/engagements/demo/report").json()
    assert "executive" in body and "limits" in body

    r = c.get("/api/v1/brand-intel/engagements/demo/report.html")
    assert r.status_code == 200
    assert r.text.startswith("<!doctype html>")


# ── borrado ───────────────────────────────────────────────────────────

def test_delete_removes_the_engagement_and_every_row_that_belongs_to_it(db, engagement):
    """A surviving child row is private client data with no owner and no access path."""
    from modules.brand_intel.models.models import (
        BrandEntity, BrandObservation, BrandWave,
    )

    assert db.query(BrandObservation).filter(
        BrandObservation.engagement_id == engagement.id).count() > 0

    r = _client(db).delete(
        f"/api/v1/brand-intel/engagements/{engagement.slug}",
        params={"confirm": engagement.slug},
    )
    assert r.status_code == 200, r.text
    assert r.json()["deleted"] == engagement.slug

    for model in (BrandObservation, BrandEntity, BrandWave):
        assert db.query(model).filter(model.engagement_id == engagement.id).count() == 0
    assert db.query(BrandEngagement).filter(
        BrandEngagement.slug == engagement.slug).count() == 0


def test_delete_without_the_echoed_slug_changes_nothing(db, engagement):
    """A DELETE that fires on the URL alone is one stale tab away from a lost dataset."""
    r = _client(db).delete(
        f"/api/v1/brand-intel/engagements/{engagement.slug}", params={"confirm": "otra"})
    assert r.status_code == 400
    assert db.query(BrandEngagement).filter(
        BrandEngagement.slug == engagement.slug).count() == 1


def test_an_analyst_cannot_delete_an_engagement(db, engagement):
    r = _client(db, role=UserRole.analyst).delete(
        f"/api/v1/brand-intel/engagements/{engagement.slug}",
        params={"confirm": engagement.slug},
    )
    assert r.status_code == 403
    assert db.query(BrandEngagement).filter(
        BrandEngagement.slug == engagement.slug).count() == 1


def test_delete_is_gated_by_role_before_the_engagement_is_even_looked_up(db, engagement):
    """The refusal must not depend on the slug, or it becomes an existence oracle.

    Deletion is admin-only, and admins are staff who can already see every engagement, so
    here the isolation boundary is the role — not the organization check that guards
    reads. A non-admin gets the same 403 whether the engagement exists or not.
    """
    c = _client(db, role=UserRole.analyst, org="org-2")
    existing = c.delete(f"/api/v1/brand-intel/engagements/{engagement.slug}",
                        params={"confirm": engagement.slug})
    absent = c.delete("/api/v1/brand-intel/engagements/fantasma",
                      params={"confirm": "fantasma"})
    assert existing.status_code == absent.status_code == 403
    assert db.query(BrandEngagement).filter(
        BrandEngagement.slug == engagement.slug).count() == 1


def test_delete_of_an_unknown_engagement_is_404_for_an_admin(db, engagement):
    r = _client(db).delete("/api/v1/brand-intel/engagements/fantasma",
                           params={"confirm": "fantasma"})
    assert r.status_code == 404


# ── clientes ──────────────────────────────────────────────────────────

def test_create_client_and_reject_a_duplicate_code(db):
    c = _client(db)
    payload = {"code": "MCD-RD", "name": "Operador RD"}
    assert c.post("/api/v1/brand-intel/clients", json=payload).status_code == 201
    # Un código duplicado uniría dos clientes que no son el mismo.
    assert c.post("/api/v1/brand-intel/clients", json=payload).status_code == 409


def test_an_engagement_created_under_a_client_inherits_its_organization(db):
    """Olvidar la organización creaba un estudio invisible para el propio cliente."""
    c = _client(db)
    c.post("/api/v1/brand-intel/clients",
           json={"code": "MCD-RD", "name": "Operador RD", "organization_id": "org-9"})
    r = c.post("/api/v1/brand-intel/engagements",
               json={"slug": "mcd-tracker", "client": "MCD-RD",
                     "client_name": "ignorado", "focal_brand": "McDonald's"})
    assert r.status_code == 201
    assert r.json()["client"] == "MCD-RD"

    eng = db.query(BrandEngagement).filter(BrandEngagement.slug == "mcd-tracker").one()
    assert eng.organization_id == "org-9"
    assert eng.client_name == "Operador RD"      # el nombre lo manda el cliente


def test_several_studies_hang_off_one_client(db):
    c = _client(db)
    c.post("/api/v1/brand-intel/clients", json={"code": "MCD-RD", "name": "Operador RD"})
    for slug, brand in [("mcd-tracker", "McDonald's"), ("mcd-clima", "Clima interno")]:
        assert c.post("/api/v1/brand-intel/engagements",
                      json={"slug": slug, "client": "MCD-RD",
                            "client_name": "se ignora", "focal_brand": brand},
                      ).status_code == 201

    listado = c.get("/api/v1/brand-intel/clients").json()
    assert [(x["code"], x["engagements"]) for x in listado] == [("MCD-RD", 2)]
    encargos = c.get("/api/v1/brand-intel/engagements").json()
    assert {e["client_code"] for e in encargos if e["slug"].startswith("mcd")} == {"MCD-RD"}


def test_an_unknown_client_is_refused_rather_than_silently_ignored(db):
    r = _client(db).post("/api/v1/brand-intel/engagements",
                         json={"slug": "huerfano", "client": "NO-EXISTE",
                               "client_name": "Cliente", "focal_brand": "Marca"})
    assert r.status_code == 404
    assert db.query(BrandEngagement).filter(
        BrandEngagement.slug == "huerfano").count() == 0


def test_the_client_is_not_the_isolation_boundary(db, engagement):
    """El acceso se sigue resolviendo sobre el encargo, no sobre el cliente.

    Si el cliente concediera acceso, bastaría pertenecer a su organización para leer un
    estudio marcado para otra — y esta es la propiedad que no puede romperse al añadir
    una capa por encima.
    """
    c = _client(db)
    c.post("/api/v1/brand-intel/clients",
           json={"code": "ACME", "name": "Acme", "organization_id": "org-1"})
    c.post("/api/v1/brand-intel/engagements",
           json={"slug": "acme-otro", "client": "ACME", "client_name": "Acme",
                 "focal_brand": "Marca", "organization_id": "org-2"})

    # Un usuario de la organización del CLIENTE no puede abrir un estudio de otra.
    r = _client(db, role=UserRole.viewer, org="org-1").get(
        "/api/v1/brand-intel/engagements/acme-otro")
    assert r.status_code == 404


# ── el informe como documento ─────────────────────────────────────────

@pytest.mark.parametrize("fmt,firma,mime", [
    ("pdf", b"%PDF", "application/pdf"),
    ("docx", b"PK", "wordprocessingml"),
])
def test_the_report_downloads_as_a_document(db, engagement, tmp_path, fmt, firma, mime):
    """El estándar de la casa pide online · PDF · Word; el módulo solo daba HTML."""
    r = _client(db).get(f"/api/v1/brand-intel/engagements/demo/report.{fmt}")
    assert r.status_code == 200, r.text
    assert r.content[:4].startswith(firma)
    assert mime in r.headers["content-type"]
    assert f"informe_contexto_demo.{fmt}" in r.headers["content-disposition"]


def test_the_html_route_still_wins_over_the_document_route(db, engagement):
    """`.html` tiene su propio renderizador: una ruta comodín declarada antes se lo traga."""
    r = _client(db).get("/api/v1/brand-intel/engagements/demo/report.html")
    assert r.status_code == 200
    assert "text/html" in r.headers["content-type"]
    assert r.text.lstrip().startswith("<!doctype html")


def test_an_unknown_report_format_is_404(db, engagement):
    assert _client(db).get(
        "/api/v1/brand-intel/engagements/demo/report.xlsx").status_code == 404


# ── conclusiones del proveedor ────────────────────────────────────────

def test_conclusions_endpoint_lists_what_the_provider_claims(db, engagement):
    from modules.brand_intel.ingest import conclusions as conc

    conc.store_conclusions(db, str(engagement.id), [conc.Conclusion(
        claim="Focal mantiene el liderazgo como marca favorita.", page_number=19,
        kind="hallazgo", subjects=("Focal",), topic="Lugar favorito",
        metric_code="", direction="estable", wave_label="Ola 2", confident=True,
    )])
    db.commit()

    r = _client(db).get("/api/v1/brand-intel/engagements/demo/conclusions")
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 1
    row = body["conclusions"][0]
    # La lámina SÍ viaja en esta vista: es el recibo que permite comprobar que lo que se
    # atribuye al proveedor está escrito en su mazo. En el informe del cliente no aparece.
    assert row["page_number"] == 19
    assert row["subject_slugs"] == ["focal"]
    assert row["metric_code"] == "favourite_place"


def test_conclusions_of_another_organization_are_404(db, engagement):
    """Lo que concluyó el estudio de otro cliente es tan privado como sus cifras."""
    engagement.organization_id = "org-1"
    db.commit()
    r = _client(db, role=UserRole.viewer, org="org-2").get(
        "/api/v1/brand-intel/engagements/demo/conclusions")
    assert r.status_code == 404


# ── contraste y discrepancias ─────────────────────────────────────────

def _open_discrepancy(db, engagement):
    """Fabrica un movimiento significativo y deja una discrepancia abierta."""
    from modules.brand_intel.ingest import conclusions as conc
    from modules.brand_intel.models.models import BrandObservation, BrandWave

    w2 = (db.query(BrandWave).filter(BrandWave.engagement_id == engagement.id,
                                     BrandWave.code == "w2").one())
    obs = (db.query(BrandObservation)
           .filter(BrandObservation.engagement_id == engagement.id,
                   BrandObservation.wave_id == w2.id,
                   BrandObservation.brand_slug == "focal",
                   BrandObservation.metric_code == "favourite_place").one())
    obs.value = 60.0
    conc.store_conclusions(db, str(engagement.id), [conc.Conclusion(
        claim="Focal pierde preferencia de forma sostenida.", page_number=19,
        kind="hallazgo", subjects=("Focal",), topic="",
        metric_code="favourite_place", direction="baja",
        wave_label="Ola 2", confident=True,
    )])
    db.commit()
    r = _client(db).post("/api/v1/brand-intel/engagements/demo/contrast")
    assert r.status_code == 200
    return r.json()


def test_contrast_opens_a_discrepancy_and_lists_it(db, engagement):
    out = _open_discrepancy(db, engagement)
    assert out["discrepan"] == 1

    r = _client(db).get("/api/v1/brand-intel/engagements/demo/discrepancies")
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 1 and body["bloqueantes"] == 1
    assert body["discrepancies"][0]["status"] == "abierta"
    assert body["discrepancies"][0]["page_number"] == 19


def test_discrepancy_can_be_advanced_with_a_note(db, engagement):
    _open_discrepancy(db, engagement)
    did = _client(db).get(
        "/api/v1/brand-intel/engagements/demo/discrepancies").json()[
        "discrepancies"][0]["id"]

    r = _client(db).patch(
        f"/api/v1/brand-intel/engagements/demo/discrepancies/{did}",
        json={"status": "retirada",
              "resolution_note": "Ipsos usa media móvil; el corte difiere."})
    assert r.status_code == 200
    assert r.json()["status"] == "retirada"
    assert r.json()["updated_by"] == "analista@sdq.test"


def test_closing_without_a_note_is_422(db, engagement):
    _open_discrepancy(db, engagement)
    did = _client(db).get(
        "/api/v1/brand-intel/engagements/demo/discrepancies").json()[
        "discrepancies"][0]["id"]
    r = _client(db).patch(
        f"/api/v1/brand-intel/engagements/demo/discrepancies/{did}",
        json={"status": "acordada"})
    assert r.status_code == 422


def test_discrepancies_of_another_organization_are_404(db, engagement):
    """La mesa con el proveedor es tan privada como las cifras del cliente."""
    engagement.organization_id = "org-1"
    db.commit()
    r = _client(db, role=UserRole.viewer, org="org-2").get(
        "/api/v1/brand-intel/engagements/demo/discrepancies")
    assert r.status_code == 404


def test_mesa_document_downloads_with_its_own_watermarked_identity(db, engagement):
    _open_discrepancy(db, engagement)
    r = _client(db).get("/api/v1/brand-intel/engagements/demo/mesa.pdf")
    assert r.status_code == 200
    assert r.content[:5] == b"%PDF-"
    assert "nota_mesa_demo.pdf" in r.headers["content-disposition"]


def test_mesa_document_is_404_when_the_table_is_empty(db, engagement):
    r = _client(db).get("/api/v1/brand-intel/engagements/demo/mesa.pdf")
    assert r.status_code == 404
    assert "vacía" in r.json()["detail"]


def test_mesa_document_of_another_organization_is_404(db, engagement):
    engagement.organization_id = "org-1"
    db.commit()
    r = _client(db, role=UserRole.viewer, org="org-2").get(
        "/api/v1/brand-intel/engagements/demo/mesa.pdf")
    assert r.status_code == 404


# ── S6 · las dos series que entraban por script ───────────────────────
#
# Mientras solo se cargaran con un script, actualizar el expediente dependía de una
# persona: no era una capacidad del producto. Estos tests cubren además la PRIMERA
# cobertura del lector del tablero, que hasta ahora solo se había ejercido a mano.


def _sales_workbook(*, cutoff_day: int = 3, sheet: str = "Base de Datos") -> bytes:
    """Un tablero mínimo con la forma real: encabezado, dos locales, tres jornadas."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Reporte de ventas"])          # el encabezado no está en la fila 1
    ws.append(["ID Tienda", "Restaurantes", "Ciudad", "Zona", "Fecha",
               "Ventas 2026", "Venta 2025", "Gcs 2026", "Gcs 2025",
               "Ventas D.T 2026", "Ventas D.T 2025", "Gcs D.T 2026", "Gcs D.T 2025"])
    for day in range(1, cutoff_day + 1):
        for sid, name, city in (("001", "Ágora", "Santo Domingo"),
                                ("002", "Bella Vista", "Santiago")):
            ws.append([sid, name, city, "Metro", f"2026-06-{day:02d}, lunes",
                       10000 + day, 9500 + day, 400 + day, 390 + day,
                       3000, 2900, 120, 118])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_upload_sales_stores_and_reports_what_it_read(db, engagement):
    c = _client(db)
    r = c.post("/api/v1/brand-intel/engagements/demo/sales",
               files={"file": ("tablero.xlsx", _sales_workbook(),
                               "application/vnd.openxmlformats-officedocument"
                               ".spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    body = r.json()
    lectura = body["lectura"]
    assert lectura["filas_leidas"] == 6 and lectura["locales"] == 2
    assert lectura["desde"] == "2026-06-01" and lectura["hasta"] == "2026-06-03"
    # El informe de ingesta viaja SIEMPRE: una ingesta silenciosa parece completa, y así
    # entraron cuatro columnas de Automac como NULL sin que nada fallara.
    assert "columnas_sin_mapear" in lectura and "motivos_descarte" in lectura
    assert body["guardado"]

    # Idempotente: la misma entrega otra vez actualiza, no duplica.
    r2 = c.post("/api/v1/brand-intel/engagements/demo/sales",
                files={"file": ("tablero.xlsx", _sales_workbook(), "application/xlsx")})
    assert r2.status_code == 201
    from modules.brand_intel.models.models import BrandSalesDaily
    assert db.query(BrandSalesDaily).filter(
        BrandSalesDaily.engagement_id == engagement.id).count() == 6


def test_upload_sales_rejects_the_wrong_sheet_with_its_reason(db, engagement):
    """El fallo es del ARCHIVO: 400 nombrando la hoja que falta y las que hay. Un
    agregado ya calculado no sirve, y el mensaje tiene que decirlo."""
    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/sales",
        files={"file": ("resumen.xlsx", _sales_workbook(sheet="Resumen"),
                        "application/xlsx")})
    assert r.status_code == 400
    assert "Base de Datos" in r.json()["detail"]


def test_upload_sales_rejects_a_non_workbook(db, engagement):
    r = _client(db).post("/api/v1/brand-intel/engagements/demo/sales",
                         files={"file": ("plan.pdf", b"%PDF-1.4", "application/pdf")})
    assert r.status_code == 400
    assert ".xlsx" in r.json()["detail"]


def test_upload_sales_needs_analyst(db, engagement):
    r = _client(db, role=UserRole.viewer).post(
        "/api/v1/brand-intel/engagements/demo/sales",
        files={"file": ("t.xlsx", _sales_workbook(), "application/xlsx")})
    assert r.status_code == 403


def test_sales_status_declares_its_own_window_not_the_wave(db, engagement):
    """La venta NO se recorta por ola: la delimita el tablero del operador. El endpoint no
    acepta `?wave=` para no insinuar lo contrario, y el estado sale igual sin serie."""
    c = _client(db)
    r = c.get("/api/v1/brand-intel/engagements/demo/sales")
    assert r.status_code == 200
    body = r.json()
    assert body["available"] is False and "jornadas de venta" in body["reason"]
    # Un `?wave=` se ignora en vez de fingir un recorte que no ocurre.
    assert c.get("/api/v1/brand-intel/engagements/demo/sales?wave=X").status_code == 200


def test_upload_tracker_history_rejects_a_matrix_without_its_sheet(db, engagement):
    r = _client(db).post(
        "/api/v1/brand-intel/engagements/demo/tracker-history",
        files={"file": ("matriz.xlsx", _sales_workbook(sheet="Otra"),
                        "application/xlsx")})
    assert r.status_code == 400
    assert "KPIs" in r.json()["detail"]


def test_upload_tracker_history_needs_analyst(db, engagement):
    r = _client(db, role=UserRole.viewer).post(
        "/api/v1/brand-intel/engagements/demo/tracker-history",
        files={"file": ("m.xlsx", _sales_workbook(sheet="KPIs"), "application/xlsx")})
    assert r.status_code == 403


def test_the_review_detail_exposes_the_attribute_dimension(db, engagement):
    """Sin el atributo en la respuesta, revisar 248 celdas de índice de atributo es cruzar
    el portón a ciegas: dos celdas de la misma marca y ola se ven idénticas en pantalla y
    solo difieren en la dimensión que no se muestra."""
    from modules.brand_intel.models.models import BrandExtraction, BrandExtractionCell

    ext = BrandExtraction(engagement_id=engagement.id, document_name="mazo.pdf",
                          status="validated", n_pages=1, pages_done=1)
    db.add(ext)
    db.flush()
    for attr, valor in (("hamburguesas deliciosas", 87.0), ("café", 181.0)):
        db.add(BrandExtractionCell(
            extraction_id=ext.id, engagement_id=engagement.id, page_number=32,
            wave_code="w1", brand_slug="focal", metric_code="attribute_index",
            segment="total", attribute=attr, value=valor, unit="index"))
    db.commit()

    r = _client(db).get(
        f"/api/v1/brand-intel/engagements/demo/extractions/{ext.id}")
    assert r.status_code == 200
    celdas = r.json()["cells"]
    assert {c["attribute"] for c in celdas} == {"hamburguesas deliciosas", "café"}
    # Y las dos son distinguibles: mismo todo salvo el atributo.
    assert len({(c["metric"], c["brand"], c["wave"], c["segment"]) for c in celdas}) == 1
