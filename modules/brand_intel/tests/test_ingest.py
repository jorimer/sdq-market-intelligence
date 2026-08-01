"""Ingest tests: the template round-trips, bad rows are rejected with a reason, and a
re-upload updates in place instead of duplicating.

Corrections are routine in tracker work, so the second upload of a fixed workbook is the
normal motion — it has to be safe.
"""
import io

import pytest
from openpyxl import Workbook, load_workbook

from modules.brand_intel import service as svc
from modules.brand_intel.ingest.excel_ingest import ingest_workbook
from modules.brand_intel.ingest.template import build_template
from modules.brand_intel.models.models import (
    BrandEngagement,
    BrandEntity,
    BrandObservation,
    BrandWave,
)


def _workbook(waves, brands, observations) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Olas"
    ws.append(["codigo", "etiqueta", "orden", "fecha_referencia",
               "campo_inicio", "campo_fin", "base_nominal"])
    for row in waves:
        ws.append(row)

    ws = wb.create_sheet("Marcas")
    ws.append(["slug", "nombre", "es_focal", "en_set_categoria", "orden"])
    for row in brands:
        ws.append(row)

    ws = wb.create_sheet("Observaciones")
    ws.append(["entrega", "ola", "marca", "metrica", "segmento", "valor", "base_n",
               "unidad", "fuente"])
    for row in observations:
        # Las filas de los tests se escriben sin `entrega`: el libro entero es una sola.
        ws.append([""] + list(row) if len(row) < 9 else list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def blank_engagement(db):
    eng = BrandEngagement(slug="nuevo", client_name="C", focal_brand="F", market="RD")
    db.add(eng)
    db.commit()
    return eng


def test_template_has_the_five_sheets():
    wb = load_workbook(io.BytesIO(build_template("demo", "Focal", "Cliente", "Proveedor")))
    assert set(wb.sheetnames) == {"Encargo", "Olas", "Marcas", "Observaciones", "Diccionario"}


def test_template_dictionary_lists_the_metric_vocabulary():
    wb = load_workbook(io.BytesIO(build_template()))
    codes = {r[0] for r in wb["Diccionario"].iter_rows(min_row=2, values_only=True)}
    assert "reach_7d" in codes
    assert "favourite_place" in codes


def test_ingest_loads_waves_brands_and_observations(db, blank_engagement):
    content = _workbook(
        waves=[["w1", "Ola 1", 1, "2025-01-01", None, None, 300]],
        brands=[["focal", "Focal", "SI", "SI", 1]],
        observations=[["w1", "focal", "reach_7d", "total", 26, 300, "pct", "lámina 18"]],
    )
    report = ingest_workbook(db, blank_engagement, content)
    db.commit()

    assert report.waves_created == 1
    assert report.brands_created == 1
    assert report.observations_created == 1
    assert report.rejected == []
    obs = db.query(BrandObservation).one()
    assert obs.value == 26 and obs.base_n == 300 and obs.source == "lámina 18"


def test_unknown_metric_is_rejected_with_a_reason(db, blank_engagement):
    content = _workbook(
        waves=[["w1", "Ola 1", 1, "2025-01-01", None, None, 300]],
        brands=[["focal", "Focal", "SI", "SI", 1]],
        observations=[["w1", "focal", "reach_siete_dias", "total", 26, 300, "pct", ""]],
    )
    report = ingest_workbook(db, blank_engagement, content)
    assert report.observations_created == 0
    assert len(report.rejected) == 1
    assert "Métrica desconocida" in report.rejected[0]["reason"]
    assert report.rejected[0]["row"] == 2


def test_unknown_wave_or_brand_is_rejected(db, blank_engagement):
    content = _workbook(
        waves=[["w1", "Ola 1", 1, "2025-01-01", None, None, 300]],
        brands=[["focal", "Focal", "SI", "SI", 1]],
        observations=[
            ["w9", "focal", "reach_7d", "total", 26, 300, "pct", ""],
            ["w1", "fantasma", "reach_7d", "total", 26, 300, "pct", ""],
        ],
    )
    report = ingest_workbook(db, blank_engagement, content)
    assert report.observations_created == 0
    reasons = " ".join(r["reason"] for r in report.rejected)
    assert "Ola desconocida" in reasons and "Marca desconocida" in reasons


def test_non_numeric_value_is_rejected(db, blank_engagement):
    content = _workbook(
        waves=[["w1", "Ola 1", 1, "2025-01-01", None, None, 300]],
        brands=[["focal", "Focal", "SI", "SI", 1]],
        observations=[["w1", "focal", "reach_7d", "total", "n/d", 300, "pct", ""]],
    )
    report = ingest_workbook(db, blank_engagement, content)
    assert report.observations_created == 0
    assert "no numérico" in report.rejected[0]["reason"]


def test_missing_base_is_loaded_but_flagged_downstream(db, blank_engagement):
    """A row without its base is data, not an error — it just cannot carry a verdict."""
    content = _workbook(
        waves=[["w1", "Ola 1", 1, "2025-01-01", None, None, 300]],
        brands=[["focal", "Focal", "SI", "SI", 1]],
        observations=[["w1", "focal", "reach_7d", "total", 26, None, "pct", ""]],
    )
    report = ingest_workbook(db, blank_engagement, content)
    db.commit()
    assert report.observations_created == 1
    assert db.query(BrandObservation).one().base_n is None


def test_reupload_updates_in_place_and_does_not_duplicate(db, blank_engagement):
    base = dict(
        waves=[["w1", "Ola 1", 1, "2025-01-01", None, None, 300]],
        brands=[["focal", "Focal", "SI", "SI", 1]],
    )
    ingest_workbook(db, blank_engagement, _workbook(
        observations=[["w1", "focal", "reach_7d", "total", 26, 300, "pct", ""]], **base))
    db.commit()

    report = ingest_workbook(db, blank_engagement, _workbook(
        observations=[["w1", "focal", "reach_7d", "total", 31, 300, "pct", ""]], **base))
    db.commit()

    assert report.observations_updated == 1
    assert report.observations_created == 0
    assert db.query(BrandObservation).count() == 1
    assert db.query(BrandObservation).one().value == 31
    assert db.query(BrandWave).count() == 1
    assert db.query(BrandEntity).count() == 1


def test_wave_without_reference_date_warns_about_deflation(db, blank_engagement):
    content = _workbook(
        waves=[["w1", "Ola 1", 1, None, None, None, 300]],
        brands=[["focal", "Focal", "SI", "SI", 1]],
        observations=[],
    )
    report = ingest_workbook(db, blank_engagement, content)
    assert any("deflactarse" in w for w in report.warnings)


def test_missing_sheet_is_reported_not_crashed(db, blank_engagement):
    wb = Workbook()
    wb.active.title = "Otra"
    buf = io.BytesIO()
    wb.save(buf)
    report = ingest_workbook(db, blank_engagement, buf.getvalue())
    assert any(r["reason"] == "Hoja ausente en el archivo." for r in report.rejected)


# ── dos entregas en un mismo libro ────────────────────────────────────

def test_two_deliveries_in_one_workbook_are_kept_apart(db, engagement):
    """El caso real: un libro con los resultados de dos informes del tracker.

    Sin la columna `entrega` las dos se fundían y la cifra vigente dependía del orden de
    las filas. Aquí la Ola 4 reexpone Nov '25 con una corrección (25 → 26): debe ganar la
    entrega cuya ola más reciente es posterior, esté donde esté en el libro.
    """
    from modules.brand_intel.models.models import (
        BrandExtraction, BrandObservation, BrandObservationReading,
    )

    # La entrega vieja va DESPUÉS en el libro, que es lo que antes decidía el resultado.
    content = _workbook(
        waves=[("e4", "Mar '26", 4, None, None, None, 300),
               ("e3", "Nov '25", 3, None, None, None, 300)],
        brands=[("focal", "Focal", "SI", "SI", 1)],
        observations=[
            ("Ola 4 · mar'26", "e3", "focal", "reach_7d", "total", 26, 300, "pct", ""),
            ("Ola 4 · mar'26", "e4", "focal", "reach_7d", "total", 27, 300, "pct", ""),
            ("Ola 3 · nov'25", "e3", "focal", "reach_7d", "total", 25, 300, "pct", ""),
        ],
    )
    ingest_workbook(db, engagement, content, document_name="dos-olas.xlsx")
    db.commit()

    # Dos entregas registradas, no una.
    entregas = {e.document_name for e in db.query(BrandExtraction)
                .filter(BrandExtraction.method == "excel").all()}
    assert entregas == {"Ola 4 · mar'26", "Ola 3 · nov'25"}

    # Lo que dijo cada una queda: dos lecturas para Nov '25, sin pisarse.
    e3 = next(w.id for w in svc.waves(db, engagement.id) if w.code == "e3")
    lecturas = sorted(
        r.value for r in db.query(BrandObservationReading)
        .filter(BrandObservationReading.wave_id == e3,
                BrandObservationReading.metric_code == "reach_7d").all())
    assert lecturas == [25.0, 26.0]

    # Y la vigente es la de la entrega más reciente, pese al orden de las filas.
    vigente = (db.query(BrandObservation)
               .filter(BrandObservation.wave_id == e3,
                       BrandObservation.metric_code == "reach_7d").one())
    assert vigente.value == 26.0


def test_a_workbook_without_the_delivery_column_is_one_delivery(db, engagement):
    """Compatibilidad: un libro que no declara entrega es una sola, con el nombre del fichero."""
    from modules.brand_intel.models.models import BrandExtraction

    content = _workbook(
        waves=[("e4", "Mar '26", 4, None, None, None, 300)],
        brands=[("focal", "Focal", "SI", "SI", 1)],
        observations=[("e4", "focal", "reach_7d", "total", 27, 300, "pct", "")],
    )
    ingest_workbook(db, engagement, content, document_name="mi-libro.xlsx")
    db.commit()
    nombres = {e.document_name for e in db.query(BrandExtraction)
               .filter(BrandExtraction.method == "excel").all()}
    assert nombres == {"mi-libro.xlsx"}


# ── compatibilidad con libros anteriores a la columna `entrega` ────────

def test_a_workbook_from_before_the_delivery_column_still_loads(db, engagement):
    """La regresión: añadir `entrega` al principio corrió TODAS las columnas.

    Un libro descargado antes del cambio —y ya rellenado por una persona— veía rechazadas
    sus 503 filas con «Ola desconocida: 'mcdonalds'», porque la marca caía donde el lector
    esperaba la ola. Las columnas se leen por NOMBRE de cabecera, no por posición.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Olas"
    ws.append(["codigo", "etiqueta", "orden", "fecha_referencia",
               "campo_inicio", "campo_fin", "base_nominal"])
    ws.append(["v1", "Ola vieja", 1, None, None, None, 300])
    ws = wb.create_sheet("Marcas")
    ws.append(["slug", "nombre", "es_focal", "en_set_categoria", "orden"])
    ws.append(["focal", "Focal", "SI", "SI", 1])
    ws = wb.create_sheet("Observaciones")
    # Cabecera VIEJA: sin `entrega`.
    ws.append(["ola", "marca", "metrica", "segmento", "valor", "base_n", "unidad", "fuente"])
    ws.append(["v1", "focal", "reach_7d", "total", 41, 300, "pct", "Hot Tracker · lámina 18"])
    buf = io.BytesIO()
    wb.save(buf)

    rep = ingest_workbook(db, engagement, buf.getvalue(), document_name="viejo.xlsx")
    db.commit()
    d = rep.as_dict()
    assert d["total_rechazadas"] == 0, d["rechazadas"]
    assert d["observaciones"]["creadas"] == 1
    # Y se avisa de que el libro no separa entregas, en vez de callarlo.
    assert any("no declara la columna 'entrega'" in w for w in d["advertencias"])


def test_the_templates_own_note_rows_are_not_rejected_as_data(db, engagement):
    """Un libro correcto no puede reportar filas rechazadas por una frase que escribió
    el propio sistema al generar la plantilla."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Olas"
    ws.append(["codigo", "etiqueta", "orden", "fecha_referencia",
               "campo_inicio", "campo_fin", "base_nominal"])
    ws.append(["v1", "Ola", 1, None, None, None, 300])
    ws = wb.create_sheet("Marcas")
    ws.append(["slug", "nombre", "es_focal", "en_set_categoria", "orden"])
    ws.append(["focal", "Focal", "SI", "SI", 1])
    ws = wb.create_sheet("Observaciones")
    ws.append(["entrega", "ola", "marca", "metrica", "segmento", "valor", "base_n",
               "unidad", "fuente"])
    ws.append(["Ola 1", "v1", "focal", "reach_7d", "total", 41, 300, "pct", ""])
    ws.append(["marca vacía = métrica de categoría · base_n vacío = …"])   # la nota
    buf = io.BytesIO()
    wb.save(buf)

    d = ingest_workbook(db, engagement, buf.getvalue()).as_dict()
    assert d["total_rechazadas"] == 0, d["rechazadas"]
    assert d["observaciones"]["creadas"] == 1


def test_a_workbook_missing_the_required_columns_says_so_once(db, engagement):
    """Rechazar 500 filas una a una esconde que el problema es la cabecera."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Olas"
    ws.append(["codigo", "etiqueta", "orden", "fecha_referencia",
               "campo_inicio", "campo_fin", "base_nominal"])
    ws.append(["v1", "Ola", 1, None, None, None, 300])
    ws = wb.create_sheet("Marcas")
    ws.append(["slug", "nombre", "es_focal", "en_set_categoria", "orden"])
    ws.append(["focal", "Focal", "SI", "SI", 1])
    ws = wb.create_sheet("Observaciones")
    ws.append(["periodo", "brand", "kpi", "valor"])          # cabecera equivocada
    ws.append(["v1", "focal", "reach_7d", 41])
    buf = io.BytesIO()
    wb.save(buf)

    d = ingest_workbook(db, engagement, buf.getvalue()).as_dict()
    assert d["total_rechazadas"] == 1
    assert "columnas obligatorias" in d["rechazadas"][0]["reason"]
