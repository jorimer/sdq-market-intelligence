"""Oficina Nacional de Estadística (ONE) — live connector for social data.

Feeds Eje 6 (`social_dev`) with real ONE statistics. First dataset wired: the
*Tasa de Pobreza Monetaria General y Extrema por Regiones de Desarrollo* (the one
IDM variable ONE publishes structured and by the 10 development regions), CSV from
``descargas.one.gob.do``, 2000-…. Other ONE datasets (services, education) are
added to ``DATASETS`` over time; the studies in PDF (Censo/ENHOGAR) go through the
publications-digest path, not here.

``Record.dimension`` carries the region slug (no separate field), mirroring how
:mod:`shared.data.bcrd_sectors` carries the sector slug. Missing values stay
``None`` — never interpolated. ``live`` mode downloads the CSV; ``fixture`` mode
reads ``one.json`` for offline/tests.

A second dataset is wired below as plain module functions (not the ``Record``
client): the national ONE/BCRD labour series (informality + income proxy) that
fill two IDM variables, scraped from the Trabajo landing by media-hash (the DGA
pattern) and parsed from their *Indicador* sheet.
"""
import csv
import io
import logging
import re
import unicodedata
from datetime import date
from html import unescape
from typing import Dict, List, Optional

from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage

logger = logging.getLogger("sdq.data.one")

# Poverty dataset (datos.gob.do org ONE) — public CSV on the ONE download CDN.
POVERTY_URL = (
    "https://descargas.one.gob.do/download/OGTIC/"
    "Tasa_de_Pobreza_Monetaria_General_y_Extrema_por_Regiones_de_Desarrollo,_2000-2024.csv"
)
# CSV columns: "Tasa de Pobreza" {Pobreza Extrema|General} · region · "Porcentaje" · "Año".
POVERTY_GENERAL = "pobreza general"
POVERTY_EXTREME = "pobreza extrema"

# The 10 development regions → stable slugs. Source labels vary in leading spaces
# / "Ozama o Metropolitana", so we match on a normalized key.
REGIONS: List[tuple] = [
    ("cibao_norte", "Cibao Norte"),
    ("cibao_sur", "Cibao Sur"),
    ("cibao_nordeste", "Cibao Nordeste"),
    ("cibao_noroeste", "Cibao Noroeste"),
    ("valdesia", "Valdesia"),
    ("enriquillo", "Enriquillo"),
    ("el_valle", "El Valle"),
    ("higuamo", "Higuamo"),
    ("ozama", "Ozama o Metropolitana"),
    ("yuma", "Yuma"),
]


def _norm(s: object) -> str:
    """Accent/case/space-insensitive key for matching provider labels."""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return " ".join(t.casefold().split())


# Known label variants the provider has used for the same region → slug. Ozama
# (Gran Santo Domingo, the most populous region) circulates under several names;
# a single-token rename must not silently drop it (matching is exact-normalized,
# not token-subset). Extend here if the ONE renames a region.
_REGION_ALIASES: Dict[str, str] = {
    "ozama": "ozama",
    "metropolitana": "ozama",
    "gran santo domingo": "ozama",
}
_REGION_BY_NORM: Dict[str, str] = {_norm(label): slug for slug, label in REGIONS}
_REGION_BY_NORM.update({_norm(k): v for k, v in _REGION_ALIASES.items()})


def region_catalog() -> List[tuple]:
    """``[(slug, display_name)]`` for seeding the region peer set."""
    return [(slug, label) for slug, label in REGIONS]


def _parse_poverty_csv(raw: bytes) -> List[tuple]:
    """Parse the poverty CSV → ``[(theme, region_slug, year, value)]``.

    ``theme`` is ``poverty_rate`` (general) or ``poverty_extreme``. Rows whose
    region isn't a known development region are skipped (logged), never guessed.
    """
    text = None
    for enc in ("utf-8-sig", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV de pobreza ONE: no se pudo decodificar")

    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return []
    out: List[tuple] = []
    unknown = set()
    for r in rows[1:]:
        if len(r) < 4:
            continue
        kind = _norm(r[0])
        slug = _REGION_BY_NORM.get(_norm(r[1]))
        year = r[3].strip()
        if slug is None:
            unknown.add(r[1].strip())
            continue
        if not year.isdigit():
            continue
        try:
            value = float(r[2].replace(",", "").strip())
        except ValueError:
            continue
        if kind == POVERTY_GENERAL:
            out.append(("poverty_rate", slug, year, value))
        elif kind == POVERTY_EXTREME:
            out.append(("poverty_extreme", slug, year, value))
    if unknown:
        logger.warning("[ONE] regiones no reconocidas en el CSV de pobreza: %s", sorted(unknown))
    return out


class ONEClient(FixtureBackedClient):
    """ONE social statistics (Eje 6). Live = download the configured CSVs."""

    source = "ONE"
    license = "datos oficiales ONE — uso público con cita"
    license_ok = True
    fixture_file = "one.json"
    live_phase = "Fase 4 (Eje 6 · social)"

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        self.check_license()
        if self.mode == "live":
            return self._fetch_live(series, period)
        return self._fetch_fixture(series, period)

    # ── Live ──────────────────────────────────────────────────────
    def _fetch_live(self, series: Optional[str], period: Optional[str]) -> List[Record]:  # pragma: no cover - network I/O
        import httpx

        resp = httpx.get(POVERTY_URL, timeout=40, follow_redirects=True)
        resp.raise_for_status()
        published = self._published_at(resp)
        lineage = Lineage(
            source="ONE", license=self.license, fetched_at=date.today(),
            published_at=published, url=POVERTY_URL,
            note="Tasa de pobreza monetaria por regiones de desarrollo",
        )
        out: List[Record] = []
        for theme, slug, year, value in _parse_poverty_csv(resp.content):
            out.append(Record(series=theme, period=year, value=value,
                              lineage=lineage, unit="% de la población", dimension=slug))
        return _filter(out, series, period)

    @staticmethod
    def _published_at(resp) -> Optional[date]:  # pragma: no cover - network I/O
        from datetime import datetime

        lm = resp.headers.get("last-modified")
        if not lm:
            return None
        try:
            return datetime.strptime(lm, "%a, %d %b %Y %H:%M:%S %Z").date()
        except ValueError:
            return None

    # ── Fixture (offline / tests) ─────────────────────────────────
    def _fetch_fixture(self, series: Optional[str], period: Optional[str]) -> List[Record]:
        """Fixture shape: ``{"<slug>": {"<theme>": {"<year>": value}}}``."""
        fixture = self._load_fixture(self.fixture_file)
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today())
        out: List[Record] = []
        for slug, themes in fixture.items():
            for theme, obs in themes.items():
                for yr, val in obs.items():
                    out.append(Record(
                        series=theme, period=str(yr),
                        value=None if val is None else float(val),
                        lineage=lineage, unit="% de la población", dimension=slug,
                    ))
        return _filter(out, series, period)


def _filter(records: List[Record], series: Optional[str], period: Optional[str]) -> List[Record]:
    if series:
        records = [r for r in records if r.series == series]
    if period:
        records = [r for r in records if r.period == period]
    return records


one_client = ONEClient()


# ── ONE labour statistics (national annual series) ─────────────────────────
# Real ONE/BCRD (ENFT/ENCFT) labour indicators that fill two IDM variables held
# as declared rubric: informality_rate (exact match) and income (a declared
# PROXY: hourly labour income, not household per-capita income). National annual
# series, applied to every region like the WDI health vars. Files live on the
# ONE Umbraco CDN under ``/media/<hash>/<slug>.xlsx`` (the DGA media-hash pattern);
# the hash rotates, so we scrape the landing page for the current links.
LABOR_LANDING = (
    "https://www.one.gob.do/datos-y-estadisticas/temas/estadisticas-sociales/trabajo/"
)
# IDM theme → filename slug fragment (accent-insensitive) to match on the landing.
_LABOR_SLUGS: Dict[str, str] = {
    "informality_rate": "tasa-de-informalidad-en-el-empleo-por-sexo",
    "income_per_capita": "ingreso-laboral-promedio-por-hora-trabajada-en-ocupacion-principal",
}
_MEDIA_XLSX_RE = re.compile(r"/media/[a-z0-9]+/[^\"'> ]+?\.xlsx", re.IGNORECASE)
_HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP ONE labour connector)"}


def _end_year(path: str) -> int:
    """Highest 4-digit year in a filename (its coverage end), or 0."""
    return max((int(y) for y in re.findall(r"(?:19|20)\d{2}", path)), default=0)


def _match_media_links(html_text: str, slugs: Dict[str, str]) -> Dict[str, str]:
    """``{key: /media/<hash>/<slug>.xlsx}`` matching each *slugs* fragment in the HTML.

    Accent/case-insensitive (so a rotated media hash is followed automatically);
    when several revisions match the same slug (e.g. ``…-2004-2023`` and
    ``…-2004-2024``) the latest end-year wins (deterministic, like the DGA ``-v2``
    tie-break). A file that isn't found is simply absent (never fabricated)."""
    best: Dict[str, tuple] = {}  # key -> (end_year, url)
    for raw in sorted({unescape(m) for m in _MEDIA_XLSX_RE.findall(html_text)}):
        norm = _norm(raw)
        for key, slug in slugs.items():
            if slug in norm:
                ey = _end_year(raw)
                if key not in best or ey > best[key][0]:
                    best[key] = (ey, raw)
    return {key: url for key, (_ey, url) in best.items()}


def discover_labor_links(html_text: str) -> Dict[str, str]:
    """``{idm_theme: /media/<hash>/<slug>.xlsx}`` for the Trabajo labour files."""
    return _match_media_links(html_text, _LABOR_SLUGS)


def parse_one_indicator_xlsx(content: bytes) -> List[tuple]:
    """Parse an ONE *Indicador* sheet → ``[(year, value)]`` from the Total column.

    The file's first sheet is a *Ficha técnica* (metadata); the series lives in
    the ``Indicador`` sheet as ``Año | Total | Hombres | Mujeres``. Year rows are
    found by a 4-digit year in column A; missing/non-numeric stays out."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    # Sheet names often carry trailing spaces/case; fall back to the last sheet.
    sheet = next((n for n in wb.sheetnames if n.strip().casefold() == "indicador"), None)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[-1]]
    out: List[tuple] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        a, total = row[0], (row[1] if len(row) > 1 else None)
        if isinstance(a, (int, float)) and 1990 <= int(a) <= 2100 and isinstance(total, (int, float)):
            out.append((int(a), round(float(total), 4)))
    return out


def fetch_one_labor() -> List[tuple]:  # pragma: no cover - network I/O
    """Live: scrape the Trabajo landing, download the matched files, parse them →
    ``[(idm_theme, year, value)]`` (national). Best-effort per file."""
    import urllib.parse

    import httpx

    resp = httpx.get(LABOR_LANDING, timeout=40, follow_redirects=True, headers=_HEADERS)
    resp.raise_for_status()
    links = discover_labor_links(resp.text)
    out: List[tuple] = []
    for theme, path in links.items():
        url = "https://www.one.gob.do" + urllib.parse.quote(path)
        try:
            f = httpx.get(url, timeout=60, follow_redirects=True, headers=_HEADERS)
            f.raise_for_status()
            for year, value in parse_one_indicator_xlsx(f.content):
                out.append((theme, year, value))
        except (httpx.HTTPError, ValueError, KeyError) as e:
            logger.warning("[ONE] descarga/parseo de %s (%s) falló: %s", theme, path, e)
    return out


# ── ONE education — regional net-coverage (secondary) by development region ─
# Real ONE/MINERD net secondary-coverage by the 10 development regions, 2010-2024
# (the IDM's education dimension otherwise has only the static ENHOGAR-2022 study).
# Adds real REGIONAL + TEMPORAL differentiation (access, complementing attainment).
EDUCATION_LANDING = (
    "https://www.one.gob.do/datos-y-estadisticas/temas/estadisticas-sociales/educacion/"
)
_EDUCATION_SLUGS: Dict[str, str] = {
    "secondary_coverage": "tasa-neta-de-cobertura-por-nivel-region-provincia",
}
# National average years of schooling (15+) — the IDM's schooling_years (ENHOGAR
# only reports literacy by region, not years of schooling). Simple Año|Total sheet.
_SCHOOLING_SLUG = "anos-promedio-de-educacion-de-la-poblacion-de-15-anos-y-mas"


def _coverage_region_slug(label: str) -> Optional[str]:
    """Map a 'Región X' row label to a development-region slug (None if not one)."""
    n = _norm(label).replace("region ", "", 1).strip()
    return _REGION_BY_NORM.get(n)


def region_slug(label: object) -> Optional[str]:
    """Slug de la región de desarrollo para una etiqueta del proveedor, o ``None``.

    Público porque el padrón de regiones —con sus alias, ver ``_REGION_ALIASES``— es la
    misma verdad para cualquier fuente que las nombre, no solo para el cuadro de
    cobertura de la ONE. Tolera el prefijo "Región " y las variantes de Ozama."""
    return _coverage_region_slug(str(label or ""))


_SCHOOL_YEAR_RE = re.compile(r"^(20\d{2})-(20\d{2})")
_SECONDARY_LEVELS = {"secundario", "medio"}  # "Medio" is the pre-2014 secondary label


def parse_one_coverage_xlsx(content: bytes) -> List[tuple]:
    """Parse 'tasa neta de cobertura por nivel, según región y provincia' →
    ``[(geo_level, slug, year, secondary_coverage)]``.

    ``geo_level`` is ``"region"`` (the 10 development regions) or ``"provincia"`` (the
    32 provinces). **Both levels are kept.** The file has always carried the province
    breakdown — its own title says *según región y provincia* — and this parser used to
    throw those rows away, which left the axis unable to distinguish demarcations inside
    a region. Ozama alone holds Distrito Nacional and Santo Domingo; collapsing them to
    one value erases most of the country's population into a single number.

    Columns are grouped by school year (``A-B`` → calendar year ``B``), each split
    into levels (Inicial/Primario/Secundario — older years: Inicial/Básico/Medio).
    The secondary column is located by matching the level sub-header within each
    group's span (NOT a fixed offset), so a layout/level-order change can't silently
    read the wrong level. The column header and 'Total país' are skipped; a label that
    resolves to neither a region nor a province is dropped and logged, never guessed."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    # 1) year-group starts: a row with ``20YY-20YY+1`` tokens (B == A+1).
    groups: Dict[int, int] = {}   # start_col -> ending calendar year
    year_row = -1
    for idx, r in enumerate(rows[:8]):
        cols: Dict[int, int] = {}
        for ci, c in enumerate(r):
            m = _SCHOOL_YEAR_RE.match(c.strip()) if isinstance(c, str) else None
            if m and int(m.group(2)) == int(m.group(1)) + 1:
                cols[ci] = int(m.group(2))
        if cols:
            groups, year_row = cols, idx
            break
    if not groups:
        logger.warning("[ONE] cobertura: no se encontró la fila de años-lectivos (layout cambió?)")
        return []

    # 2) secondary column per group: match the level sub-header within the group span.
    starts = sorted(groups)

    def _span_end(sc: int) -> int:
        later = [s for s in starts if s > sc]
        return later[0] if later else sc + 3

    sec_col: Dict[int, int] = {}
    for r in rows[year_row: year_row + 4]:
        for sc in starts:
            for col in range(sc, min(_span_end(sc), len(r))):
                if isinstance(r[col], str) and _norm(r[col]) in _SECONDARY_LEVELS:
                    sec_col[sc] = col
        if len(sec_col) == len(starts):
            break
    if not sec_col:
        logger.warning("[ONE] cobertura: no se ubicó la columna 'Secundario/Medio' (layout cambió?)")
        return []

    # 3) filas de agregado: regiones ('Región …') Y provincias. El header y 'Total país'
    #    quedan fuera; una etiqueta que no resuelve se descarta y se registra.
    from shared.reference.provinces import province_slug

    out: List[tuple] = []
    regions_mapped = provinces_mapped = 0
    unknown: set = set()
    for r in rows:
        if not r or not isinstance(r[0], str) or not r[0].strip():
            continue
        label = r[0]
        key = _norm(label)
        if key in ("region y provincia", "total pais"):   # encabezado / total nacional
            continue
        if key.startswith("region "):
            slug, level = _coverage_region_slug(label), "region"
            if slug:
                regions_mapped += 1
        else:
            slug, level = province_slug(label), "provincia"
            if slug:
                provinces_mapped += 1
        if not slug:
            unknown.add(label.strip())
            continue
        for sc, yr in groups.items():
            col = sec_col.get(sc)
            v = r[col] if col is not None and col < len(r) else None
            if isinstance(v, (int, float)):
                out.append((level, slug, yr, round(float(v), 2)))
    if unknown:
        logger.warning("[ONE] cobertura: etiquetas no reconocidas (descartadas): %s",
                       sorted(unknown))
    if not regions_mapped:
        logger.warning("[ONE] cobertura: ninguna fila 'Región …' mapeó (layout cambió?)")
    if not provinces_mapped:
        logger.warning("[ONE] cobertura: ninguna fila de provincia mapeó (layout cambió?)")
    return out


def fetch_one_education_coverage() -> List[tuple]:  # pragma: no cover - network I/O
    """Live: scrape the Educación landing, download the net-coverage file, parse it →
    ``[(geo_level, slug, year, secondary_coverage)]`` — regiones Y provincias."""
    import urllib.parse

    import httpx

    resp = httpx.get(EDUCATION_LANDING, timeout=40, follow_redirects=True, headers=_HEADERS)
    resp.raise_for_status()
    path = _match_media_links(resp.text, _EDUCATION_SLUGS).get("secondary_coverage")
    if not path:
        return []
    url = "https://www.one.gob.do" + urllib.parse.quote(path)
    f = httpx.get(url, timeout=90, follow_redirects=True, headers=_HEADERS)
    f.raise_for_status()
    return parse_one_coverage_xlsx(f.content)


def fetch_one_education_schooling() -> List[tuple]:  # pragma: no cover - network I/O
    """Live: national average years of schooling (15+) from the Educación landing →
    ``[(year, value)]`` (the IDM's ``schooling_years``, national)."""
    import urllib.parse

    import httpx

    resp = httpx.get(EDUCATION_LANDING, timeout=40, follow_redirects=True, headers=_HEADERS)
    resp.raise_for_status()
    path = _match_media_links(resp.text, {"schooling_years": _SCHOOLING_SLUG}).get("schooling_years")
    if not path:
        return []
    f = httpx.get("https://www.one.gob.do" + urllib.parse.quote(path), timeout=60,
                  follow_redirects=True, headers=_HEADERS)
    f.raise_for_status()
    return parse_one_indicator_xlsx(f.content)
