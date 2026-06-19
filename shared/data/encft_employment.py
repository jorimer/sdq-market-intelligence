"""ENCFT employment by activity — outcome backbone for the Gate-E sectorial backtest.

Pulls the ONE annual series *Población ocupada por actividad económica* (ENFT
2008-2016 + ENCFT 2017-2024, published by the ONE on the Trabajo landing as a
media-hash ``.xlsx``) — the real, open, historical employment panel the Gate-E
backtest validates against (the de-risked alternative to the TSS PDF bulletins).

The workbook has **10 activity branches**, not the 17 national-accounts sectors;
:mod:`shared.data.sector_crosswalk` maps branch ↔ BCRD-17 slugs. This connector
emits employment **per branch** (the real resolution) — splitting a bundle across
its member slugs would be fabrication. ``Record.dimension`` carries the branch
key, mirroring how :mod:`shared.data.bcrd_sectors` carries the sector slug.

Fail-closed guards (no silent partial panel):
  * an activity-looking row that doesn't map → the ONE renamed a branch → raise.
  * any of the 10 branches missing → raise.
  * Σ branches must equal the workbook's own ``Total`` row each year (the
    anti-stacked-blocks lesson: verify against a real magnitude, not just an
    internal invariant) → raise on deviation.
Missing values stay ``None`` — never interpolated.
"""
import io
import logging
import re
import urllib.parse
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from shared.data._text import find_media_xlsx, norm
from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage
from shared.data.sector_crosswalk import ENCFT_BRANCHES, map_label

logger = logging.getLogger("sdq.data.encft_employment")

# ONE Trabajo landing (same page the social informality/income files come from);
# the workbook link is a rotating /media/<hash>/… path → scraped, not hard-coded.
LABOR_LANDING = (
    "https://www.one.gob.do/datos-y-estadisticas/temas/estadisticas-sociales/trabajo/"
)
# Filename fragment uniquely identifying the target workbook (the ocupada-by-activity
# table; "desocupada" and the PEA tables are excluded — they carry "según" / "pea").
OCUPADOS_FRAGMENT = "poblacion-ocupada-sexo-ano-actividad-economica"
_HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP ENCFT employment connector)"}

VAR_EMPLOYMENT = "employment"
UNIT_EMPLOYMENT = "personas ocupadas (15+)"
LICENSE = "datos oficiales ONE — uso público con cita"
SUM_TOLERANCE_PCT = 0.5  # Σ branches vs the Total row that fails closed


class EmploymentError(RuntimeError):
    """The ENCFT employment workbook was missing, unreadable or structurally changed."""


def _year_columns(rows: List[tuple]) -> Tuple[Dict[int, int], int]:
    """``({year: column}, header_row_index)`` from the year-header row.

    Each year spans three columns (Total/Hombres/Mujeres); the year token sits at
    the group's first (Total) column, so its position *is* the Total column. Picks
    the row with the most year tokens (so a stray title row can't be mistaken for
    it) and returns its index so the caller skips it (its own cells are the years).
    """
    best: Dict[int, int] = {}
    best_idx = -1
    for idx, r in enumerate(rows[:15]):
        cur: Dict[int, int] = {}
        for ci, c in enumerate(r):
            if ci == 0:
                continue  # activity label column
            y: Optional[int] = None
            if isinstance(c, (int, float)) and not isinstance(c, bool) and float(c).is_integer():
                y = int(c)
            elif isinstance(c, str):
                m = re.match(r"\s*(\d{4})", c)
                y = int(m.group(1)) if m else None
            if y is not None and 2000 <= y <= 2035:
                cur.setdefault(y, ci)  # first column of the year = Total
        if len(cur) > len(best):
            best, best_idx = cur, idx
    if not best:
        raise EmploymentError("no se encontró la fila de años en el cuadro ENCFT de empleo")
    return best, best_idx


def parse_ocupados_xlsx(content: bytes) -> Tuple[Dict[str, Dict[int, float]], Dict[int, float]]:
    """Parse the workbook → ``(branches, national_total)``.

    ``branches`` is ``{branch_key: {year: value}}`` for the 10 activity branches;
    ``national_total`` is ``{year: value}`` from the ``Total`` row (used only to
    verify the partition). Pure of network — operates on bytes, unit-testable.
    """
    import openpyxl  # lazy: app boots without a workbook in hand

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    year_col, header_idx = _year_columns(rows)
    branches: Dict[str, Dict[int, float]] = {}
    national: Dict[int, float] = {}
    unmapped: List[str] = []

    for i, r in enumerate(rows):
        if i == header_idx:
            continue  # the year-header row itself (its cells are the years)
        if not r or not isinstance(r[0], str) or not r[0].strip():
            continue
        label = r[0]
        values: Dict[int, float] = {}
        for year, col in year_col.items():
            v = r[col] if col < len(r) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                values[year] = round(float(v), 3)
        if norm(label) == "total":          # the national total row
            if values:
                national = values
            continue
        key = map_label(label)
        if key is None:
            if values:  # a row with employment numbers we couldn't map → nomenclature changed
                unmapped.append(label.strip())
            continue     # otherwise a note/footnote row → skip
        branches[key] = values

    if unmapped:
        raise EmploymentError(
            f"ramas ENCFT sin mapear (¿cambió la nomenclatura de la ONE?): {sorted(unmapped)}"
        )
    return branches, national


def _check_partition(branches: Dict[str, Dict[int, float]], national: Dict[int, float]) -> None:
    """Fail-closed: all 10 branches present and Σ branches == Total each year."""
    missing = [b.key for b in ENCFT_BRANCHES if b.key not in branches]
    if missing:
        raise EmploymentError(f"ramas ENCFT ausentes en el cuadro: {sorted(missing)}")
    if not national:
        raise EmploymentError("fila 'Total' nacional ausente: no se puede verificar la partición")
    # No emitir un año de rama que no tenga fila 'Total' contra la cual verificarse
    # (sería un panel parcial sin magnitud real de control → fail-closed).
    branch_years = {y for obs in branches.values() for y in obs}
    uncovered = sorted(branch_years - set(national))
    if uncovered:
        raise EmploymentError(
            f"años con datos de rama sin fila 'Total' para verificar: {uncovered} "
            "— no se puede validar la partición"
        )
    for year, total in sorted(national.items()):
        leaf_sum = sum(branches[k].get(year, 0.0) for k in branches)
        if not total:
            if leaf_sum:  # Total=0 con ramas no nulas → estructura inconsistente
                raise EmploymentError(
                    f"{year}: Total nacional=0 con Σ ramas={leaf_sum:.0f} — estructura inconsistente"
                )
            continue
        dev = 100.0 * (leaf_sum / total - 1.0)
        if abs(dev) > SUM_TOLERANCE_PCT:
            raise EmploymentError(
                f"{year}: Σ ramas={leaf_sum:.0f} vs Total={total:.0f} (dif {dev:+.2f}%) "
                "— posible rama faltante o cambio de estructura de la ONE"
            )


def build_employment_records(
    branches: Dict[str, Dict[int, float]],
    national: Dict[int, float],
    *,
    url: Optional[str] = None,
    license_: str = LICENSE,
    published_at: Optional[date] = None,
) -> List[Record]:
    """Verify the partition, then emit one ``employment`` Record per (branch, year).

    Pure function (no I/O). ``national`` is used only by the partition guard — the
    national total is not emitted as a branch record.
    """
    _check_partition(branches, national)
    lineage = Lineage(
        source="ONE", license=license_, fetched_at=date.today(), url=url,
        published_at=published_at,
        note="Población ocupada por actividad económica (ENFT 2008-2016 / ENCFT 2017-2024)",
    )
    out: List[Record] = []
    for key in (b.key for b in ENCFT_BRANCHES):
        for year, value in sorted(branches[key].items()):
            out.append(Record(
                series=VAR_EMPLOYMENT, period=str(year), value=value,
                lineage=lineage, unit=UNIT_EMPLOYMENT, dimension=key,
            ))
    return out


def _filter(records: List[Record], series: Optional[str], period: Optional[str]) -> List[Record]:
    if series:
        records = [r for r in records if r.series == series]
    if period:
        records = [r for r in records if r.period == period]
    return records


class EmploymentClient(FixtureBackedClient):
    """ONE ENCFT employment by activity branch (Gate-E sectorial outcome)."""

    source = "ONE"
    license = LICENSE
    license_ok = True
    fixture_file = "encft_employment.json"
    live_phase = "Fase 4 (Eje 3 · empleo ENCFT)"

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        self.check_license()
        if self.mode == "live":
            return self._fetch_live(series, period)
        return self._fetch_fixture(series, period)

    # ── Live ──────────────────────────────────────────────────────
    def _fetch_live(self, series: Optional[str], period: Optional[str]) -> List[Record]:  # pragma: no cover - network I/O
        import httpx

        resp = httpx.get(LABOR_LANDING, timeout=40, follow_redirects=True, headers=_HEADERS)
        resp.raise_for_status()
        path = find_media_xlsx(resp.text, {"ocupados": OCUPADOS_FRAGMENT}).get("ocupados")
        if not path:
            raise EmploymentError(
                "no se encontró el cuadro 'Población ocupada por actividad económica' "
                "en la landing de Trabajo de la ONE (¿cambió el enlace?)"
            )
        url = "https://www.one.gob.do" + urllib.parse.quote(path)
        published = self._published_at(url)
        f = httpx.get(url, timeout=90, follow_redirects=True, headers=_HEADERS)
        f.raise_for_status()
        branches, national = parse_ocupados_xlsx(f.content)
        records = build_employment_records(branches, national, url=url, published_at=published)
        return _filter(records, series, period)

    @staticmethod
    def _published_at(url: str) -> Optional[date]:  # pragma: no cover - network I/O
        """The CDN's ``Last-Modified`` for the workbook (best-effort provenance)."""
        import httpx
        try:
            resp = httpx.head(url, timeout=20, follow_redirects=True, headers=_HEADERS)
            lm = resp.headers.get("last-modified")
            if lm:
                return datetime.strptime(lm, "%a, %d %b %Y %H:%M:%S %Z").date()
        except Exception as e:  # noqa: BLE001 — provenance is best-effort
            logger.warning("[encft_employment] no se pudo leer Last-Modified: %s", e)
        return None

    # ── Fixture (offline / tests) ─────────────────────────────────
    def _fetch_fixture(self, series: Optional[str], period: Optional[str]) -> List[Record]:
        """Fixture shape: ``{"<branch_key>": {"employment": {"<year>": value}}}``."""
        fixture = self._load_fixture(self.fixture_file)
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today())
        out: List[Record] = []
        for key, vars_ in fixture.items():
            for var, obs in vars_.items():
                for yr, val in obs.items():
                    out.append(Record(
                        series=var, period=str(yr),
                        value=None if val is None else float(val),
                        lineage=lineage, unit=UNIT_EMPLOYMENT, dimension=key,
                    ))
        return _filter(out, series, period)


employment_client = EmploymentClient()
