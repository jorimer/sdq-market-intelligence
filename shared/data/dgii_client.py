"""Dirección General de Impuestos Internos (DGII) — fiscal pulse connector.

Effective tax collection (*recaudación efectiva*) by major tax group, monthly —
the national fiscal pulse that feeds Eje 2 (macro). The DGII publishes the
*Informe de Recaudación Mensual* as a ZIP containing an Excel; the ``Recaudo
Histórico Mensual`` sheet carries the current year's collection by tax concept ×
month. We emit the **7 top-level tax groups** (ISR, propiedad, ITBIS/mercancías,
comercio exterior, ecológicos, contraprestación, otros) — a clean, explainable
composition of total collection — not every sub-tax.

Source license: public DGII statistical publication (use with citation); the same
figures are mirrored on datos.gob.do under ODbL. Missing/future months stay absent
(never fabricated). ``Record.dimension`` carries the tax-group slug.

By **tax type**, not by economic sector: a per-sector collection table exists only
in the annual *Boletín Estadístico* PDF, frozen at 2019 (Eje 3 use deferred).
"""
import io
import logging
import re
import zipfile
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from shared.data._text import norm
from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage

logger = logging.getLogger("sdq.data.dgii")

INFORME_URL = "https://dgii.gov.do/estadisticas/informeRecaudacionMensual/Paginas/default.aspx"
SHEET = "Recaudo Historico Mensual"
LICENSE = "datos públicos DGII (Informe de Recaudación) — uso con cita; ODbL en datos.gob.do"
VAR = "recaudacion"
UNIT = "RD$"
_HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP DGII fiscal connector)"}

_MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
          "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

# The 7 top-level tax groups (UPPERCASE rows). (slug, normalized-label fragment).
GROUPS: List[Tuple[str, str]] = [
    ("ingresos", "impuestos sobre los ingresos"),
    ("propiedad", "impuestos sobre la propiedad"),
    ("mercancias_servicios", "impuestos internos sobre mercancias y servicios"),
    ("comercio_exterior", "impuestos s/el comercio y las transacciones"),
    ("ecologicos", "impuestos ecologicos"),
    ("contraprestacion", "ingresos por contraprestacion"),
    ("otros", "otros ingresos"),
]


class DGIIError(RuntimeError):
    """The DGII recaudación report was missing, unreadable or structurally changed."""


def latest_zip_per_year(html: str) -> Dict[str, str]:
    """``{year: zip_path}`` — the most complete monthly ZIP for each year on the page.

    Filenames look like ``Recaudacion enero-mayo 2026.zip``; per year, keep the one
    whose last named month is latest (the most complete cumulative report). Stitching
    one ZIP per year yields a multi-year monthly series.
    """
    from html import unescape

    best: Dict[str, tuple] = {}  # year -> (last_month, path)
    for path in sorted({unescape(m) for m in re.findall(r'/estadisticas/[^"\']+?\.zip', html, re.I)}):
        low = norm(path)
        ym = re.search(r"20\d{2}", low)
        if not ym:
            continue
        year = ym.group(0)
        last_month = max((i for i, mn in enumerate(_MESES, 1) if mn in low), default=0)
        if year not in best or last_month > best[year][0]:
            best[year] = (last_month, path)
    return {y: p for y, (_lm, p) in best.items()}


def parse_recaudacion_xlsx(content: bytes) -> Tuple[str, Dict[str, Dict[int, float]]]:
    """Parse the ``Recaudo Histórico Mensual`` sheet → ``(year, {slug: {month: value}})``.

    Pure of network. Reads the year ("Año YYYY"), the month columns from the
    ``Conceptos | Enero…Diciembre | Total`` header, and the 7 top-level tax-group
    rows. Future months (value 0) are skipped. Fail-closed if a group is missing.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if SHEET not in wb.sheetnames:
            raise DGIIError(f"hoja '{SHEET}' ausente en el Excel de recaudación DGII")
        rows = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]
    finally:
        wb.close()

    # year
    year = None
    for r in rows[:5]:
        for c in r:
            m = re.search(r"a[ñn]o\s*(20\d{2})", str(c), re.I) if c else None
            if m:
                year = m.group(1)
                break
        if year:
            break
    if not year:
        raise DGIIError("no se encontró el año ('Año YYYY') en el Excel de recaudación")

    # month columns from the header row (the one with 'Conceptos' + month names)
    month_col: Dict[int, int] = {}
    for r in rows[:6]:
        labels = {norm(c): j for j, c in enumerate(r) if isinstance(c, str)}
        if "conceptos" in labels:
            for mi, mname in enumerate(_MESES, 1):
                if mname in labels:
                    month_col[mi] = labels[mname]
            break
    if len(month_col) < 12:
        raise DGIIError("encabezado de meses incompleto en el Excel de recaudación")

    out: Dict[str, Dict[int, float]] = {}
    for r in rows:
        label = r[0]
        # the 7 top-level groups are the UPPERCASE rows; sub-items are mixed-case,
        # so this avoids matching a sub-concept by substring.
        if not isinstance(label, str) or not label.strip() or label.strip() != label.strip().upper():
            continue
        key = norm(label)
        slug = next((s for s, frag in GROUPS if frag in key), None)
        if slug is None:
            continue
        obs: Dict[int, float] = {}
        for mi, col in month_col.items():
            v = r[col] if col < len(r) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool) and v != 0:
                obs[mi] = round(float(v), 2)
        out[slug] = obs

    missing = [s for s, _f in GROUPS if s not in out]
    if missing:
        raise DGIIError(f"grupos de impuesto ausentes (¿cambió el Excel DGII?): {missing}")
    return year, out


def build_records(year: str, groups: Dict[str, Dict[int, float]], *,
                  url: Optional[str] = None, published_at: Optional[date] = None) -> List[Record]:
    """Emit one Record per (tax group, month) plus a computed ``total`` group."""
    lineage = Lineage(source="DGII", license=LICENSE, fetched_at=date.today(),
                      url=url, published_at=published_at,
                      note="Recaudación efectiva por grupo de impuesto (Informe Mensual DGII)")
    out: List[Record] = []
    months = sorted({mi for obs in groups.values() for mi in obs})
    for slug, _f in GROUPS:
        for mi in sorted(groups[slug]):
            out.append(Record(series=VAR, period=f"{year}-{mi:02d}", value=groups[slug][mi],
                              lineage=lineage, unit=UNIT, dimension=slug))
    # total = Σ groups per month (verify-against-real-magnitude is intrinsic here)
    for mi in months:
        tot = sum(groups[s].get(mi, 0.0) for s, _f in GROUPS)
        if tot:
            out.append(Record(series=VAR, period=f"{year}-{mi:02d}", value=round(tot, 2),
                              lineage=lineage, unit=UNIT, dimension="total"))
    return out


def _filter(records: List[Record], series: Optional[str], period: Optional[str]) -> List[Record]:
    if series:
        records = [r for r in records if r.series == series]
    if period:
        records = [r for r in records if r.period == period]
    return records


class DGIIClient(FixtureBackedClient):
    """DGII effective tax collection by tax group (Eje 2 fiscal pulse)."""

    source = "DGII"
    license = LICENSE
    license_ok = True
    fixture_file = "dgii_recaudacion.json"
    live_phase = "Fase 4 (Eje 2 · pulso fiscal)"

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        self.check_license()
        if self.mode == "live":
            return self._fetch_live(series, period)
        return self._fetch_fixture(series, period)

    # ── Live ──────────────────────────────────────────────────────
    def _fetch_live(self, series: Optional[str], period: Optional[str]) -> List[Record]:  # pragma: no cover - network I/O
        import httpx

        import urllib.parse

        page = httpx.get(INFORME_URL, timeout=40, follow_redirects=True, headers=_HEADERS)
        page.raise_for_status()
        by_year = latest_zip_per_year(page.text)
        if not by_year:
            raise DGIIError("no se encontró un ZIP de recaudación en la página de la DGII")
        records: List[Record] = []
        errors: List[str] = []
        for yr in sorted(by_year):
            url = "https://dgii.gov.do" + urllib.parse.quote(by_year[yr])
            try:
                z = httpx.get(url, timeout=90, follow_redirects=True, headers=_HEADERS)
                z.raise_for_status()
                zf = zipfile.ZipFile(io.BytesIO(z.content))
                xlsx = next((n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls"))), None)
                if not xlsx:
                    raise DGIIError("el ZIP no contiene un Excel")
                pyear, groups = parse_recaudacion_xlsx(zf.read(xlsx))
                records.extend(build_records(pyear, groups, url=url, published_at=self._published_at(url)))
            except Exception as e:  # noqa: BLE001 — per-year best-effort (older years use a different sheet)
                errors.append(f"{yr}: {e}")
        if not records:
            raise DGIIError(f"ningún año DGII parseó: {errors}")
        if errors:
            logger.warning("[dgii] años omitidos: %s", errors)
        return _filter(records, series, period)

    @staticmethod
    def _published_at(url: str) -> Optional[date]:  # pragma: no cover - network I/O
        import httpx
        try:
            resp = httpx.head(url, timeout=20, follow_redirects=True, headers=_HEADERS)
            lm = resp.headers.get("last-modified")
            if lm:
                return datetime.strptime(lm, "%a, %d %b %Y %H:%M:%S %Z").date()
        except Exception as e:  # noqa: BLE001 — provenance best-effort
            logger.warning("[dgii] no se pudo leer Last-Modified: %s", e)
        return None

    # ── Fixture (offline / tests) ─────────────────────────────────
    def _fetch_fixture(self, series: Optional[str], period: Optional[str]) -> List[Record]:
        """Fixture shape: ``{"<slug>": {"recaudacion": {"<YYYY-MM>": value}}}``."""
        fixture = self._load_fixture(self.fixture_file)
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today())
        out: List[Record] = []
        for slug, vars_ in fixture.items():
            for var, obs in vars_.items():
                for period_, val in obs.items():
                    out.append(Record(series=var, period=str(period_),
                                      value=None if val is None else float(val),
                                      lineage=lineage, unit=UNIT, dimension=slug))
        return _filter(out, series, period)


dgii_client = DGIIClient()
