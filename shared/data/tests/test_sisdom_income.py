"""SISDOM (MEPyD) — ingreso per cápita por región. Offline (sin red)."""
import io

import openpyxl
import pytest

from shared.data.one_client import REGIONS
from shared.data.sisdom_income import SHEET, SisdomUnavailable, parse_income_per_capita

_REGION_LABELS = [label for _slug, label in REGIONS]


def _workbook(rows, sheet_name=SHEET + " ", labels=None):
    """Libro con la forma real del cuadro: bloques Nacional | Zona | Regiones."""
    wb = openpyxl.Workbook()
    wb.active.title = "Indice"
    wb.active["A1"] = "Índice del área temática"
    ws = wb.create_sheet(sheet_name)
    ws.append(["SISTEMA DE INDICADORES SOCIALES DE LA REPÚBLICA DOMINICANA"])
    ws.append([])
    ws.append(["Área Temática: Pobreza y Distribución del Ingreso"])
    ws.append(["2000 - 2024"])
    ws.append(["Índice", None, "03 3 021  Ingreso familiar mensual oficial por persona"])
    ws.append([])
    ws.append(["Desagregaciones/ Periodo", None, "Nacional", "Zona de residencia", None,
               "Regiones de Desarrollo (Ley 345-22)"])
    ws.append([None, None, None, "Urbana", "Rural"] + list(labels or _REGION_LABELS))
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _year_block(year_label, nacional, base):
    """Marzo / Septiembre / Año — la etiqueta del año va SOLO en la primera fila."""
    reg = [base + i for i in range(len(_REGION_LABELS))]
    return [
        [year_label, "Marzo", nacional, nacional + 1, nacional - 1] + reg,
        [None, "Septiembre", nacional, nacional + 1, nacional - 1] + reg,
        [None, "Año", nacional, nacional + 1, nacional - 1] + reg,
    ]


def test_toma_la_fila_anual_y_las_10_regiones():
    """Solo la fila 'Año' entra: Marzo y Septiembre son cortes dentro del mismo año y
    estamparlos con el año daría tres valores para un período."""
    content = _workbook(_year_block("2024", 18000, 13000))
    rows = parse_income_per_capita(content)

    assert len(rows) == 10, "una observación por región, solo de la fila anual"
    assert {slug for slug, _y, _v in rows} == {slug for slug, _l in REGIONS}
    assert all(y == 2024 for _s, y, _v in rows)
    by = {s: v for s, _y, v in rows}
    assert by["cibao_norte"] == 13000.0          # primera columna regional
    assert by["ozama"] == 13008.0                # 'Ozama o Metropolitana' resuelve


def test_el_nacional_y_la_zona_no_entran_como_region():
    """El cuadro pone Nacional/Urbana/Rural ANTES de las regiones. Si el parser leyera por
    posición se comería esas tres columnas; lee por rótulo, y ninguna resuelve a región."""
    rows = parse_income_per_capita(_workbook(_year_block("2024", 99999, 13000)))
    assert 99999.0 not in {v for _s, _y, v in rows}
    assert 100000.0 not in {v for _s, _y, v in rows}   # 'Urbana' = nacional + 1


def test_el_ano_se_arrastra_entre_filas():
    """La etiqueta de año aparece una sola vez por bloque; sin arrastrarla, la fila 'Año'
    quedaría huérfana y el cuadro entero saldría vacío."""
    content = _workbook(_year_block("2023", 15000, 11000) + _year_block("2024", 18000, 13000))
    rows = parse_income_per_capita(content)
    assert {y for _s, y, _v in rows} == {2023, 2024}
    assert len(rows) == 20


def test_2016_duplicado_gana_la_asteriscada():
    """El cuadro trae 2016 dos veces: ENFT (sin asterisco) y ENCFT (con). 2017-2024 solo
    existen asteriscados, así que quedarse con la vieja fabricaría un salto de nivel
    contra 2017 que después nadie podría explicar. El orden no debe importar."""
    viejo = [[None, "Año", 10983, 11906, 7404] + [7000] * 10]
    nuevo = [["2016*", "Año", 10096, 10688, 7798] + [9000] * 10]
    directo = parse_income_per_capita(_workbook(_year_block("2016", 10983, 7000) + nuevo))
    assert {v for _s, y, v in directo if y == 2016} == {9000.0}

    # Y al revés (la asteriscada primero): sigue ganando la asteriscada, no la última.
    invertido = parse_income_per_capita(
        _workbook([["2016*", "Año", 10096, 10688, 7798] + [9000] * 10]
                  + [["2016", "Año", 10983, 11906, 7404] + [7000] * 10] + viejo))
    assert {v for _s, y, v in invertido if y == 2016} == {9000.0}


def test_un_duplicado_que_no_se_puede_resolver_falla():
    """Dos filas del mismo año con la MISMA marca no tienen criterio: elegir una sería
    adivinar, y adivinar en silencio es lo que este repo no hace."""
    rows = ([["2020", "Año", 1, 2, 3] + [500] * 10]
            + [["2020", "Año", 1, 2, 3] + [900] * 10])
    with pytest.raises(SisdomUnavailable, match="dos veces"):
        parse_income_per_capita(_workbook(rows))


def test_si_falta_una_region_falla_en_vez_de_entregar_nueve():
    """El IDM normaliza min-max ENTRE regiones: con nueve, el mínimo o el máximo se corre
    y cambia el score de las demás. Un panel incompleto no es 'menos dato', es otro dato."""
    incompleto = _REGION_LABELS[:-1] + ["Región Que No Existe"]
    with pytest.raises(SisdomUnavailable, match="9 de 10"):
        parse_income_per_capita(
            _workbook(_year_block("2024", 18000, 13000), labels=incompleto))


def test_hoja_ausente_lo_dice_con_su_nombre():
    with pytest.raises(SisdomUnavailable, match="no trae la hoja"):
        parse_income_per_capita(_workbook(_year_block("2024", 1, 1), sheet_name="Otra"))


# El descubrimiento del libro YA NO vive acá: se mudó a `shared.data.sisdom_common` junto
# con el resto de la plumbing, y sus tests a `test_sisdom_common.py`. Este módulo tenía una
# COPIA del descubrimiento y por eso se quedó apuntando al host muerto cuando el emisor se
# mudó — lo vigila `test_una_sola_plumbing_del_sisdom.py`.


def _multi_block_workbook():
    """Dos regionalizaciones apiladas, como en los cuadros hermanos del SISDOM: la vieja
    (Dec 685-00) primero y la vigente (Ley 345-22) después, con cifras distintas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["SISDOM"])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Desagregaciones/ Periodo", None, "Nacional",
               "Regiones de desarrollo (Dec 685-00)"] + [None] * 9
              + ["Regiones de Desarrollo (Ley 345-22)"])
    ws.append([None, None, None] + _REGION_LABELS + _REGION_LABELS)
    ws.append(["2024", "Año", 18000] + [1000] * 10 + [9000] * 10)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_con_varias_regionalizaciones_gana_la_VIGENTE_no_la_primera():
    """El modo de falla es silencioso y por eso es el peligroso: la Dec 685-00 usa otro
    conjunto de regiones, así que quedarse con el primer bloque que resuelva entregaría
    cifras de OTRA regionalización sin que nada lo advierta. Se elige por lo declarado."""
    rows = parse_income_per_capita(_multi_block_workbook())
    assert len(rows) == 10
    assert {v for _s, _y, v in rows} == {9000.0}, "se tomó el bloque viejo (Dec 685-00)"


def test_regionalizaciones_todas_desconocidas_fallan():
    """Si ninguna de las declaradas es una que sepamos leer, elegir sería adivinar."""
    wb = openpyxl.load_workbook(io.BytesIO(_multi_block_workbook()))
    ws = wb[SHEET]
    ws.cell(row=7, column=14).value = "Regiones de desarrollo (Dec 999-99)"
    ws.cell(row=7, column=4).value = "Regiones de desarrollo (Dec 685-00)"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(SisdomUnavailable, match="regionalizaciones"):
        parse_income_per_capita(buf.getvalue())
