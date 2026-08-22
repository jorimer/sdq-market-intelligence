"""El cuadro por zona: nada por posición, solo filas anuales, y el quiebre no se disimula.

Los indicadores 2.3 y 2.6 de la Ley 1-12 piden pobreza RURAL. El panel abierto que ya se
ingiere está desagregado por región de desarrollo, así que ninguna combinación de sus filas
produce la cifra rural: hacía falta otro cuadro.
"""
import io

import pytest

from shared.data.sisdom_pobreza_zona import LINEAS, SISDOMZonaError, parse_zona_rural


def _libro(cabecera_zona, cabecera_linea, filas, hoja="03 3 003a "):
    """Un cuadro mínimo con la forma real: zonas en una fila, líneas en la de abajo."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja
    for c, v in cabecera_zona.items():
        ws.cell(8, c).value = v
    for c, v in cabecera_linea.items():
        ws.cell(9, c).value = v
    f = 10
    for anio, corte, valores in filas:
        ws.cell(f, 1).value = anio
        ws.cell(f, 2).value = corte
        for c, v in valores.items():
            ws.cell(f, c).value = v
        f += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


ZONA = {3: "Nacional", 5: "Urbana", 7: "Rural", 9: "Regiones de Desarrollo"}
LINEA = {3: " Indigencia monetaria", 4: " Pobreza monetaria",
         5: " Indigencia monetaria", 6: " Pobreza monetaria",
         7: " Indigencia monetaria", 8: " Pobreza monetaria",
         9: " Indigencia monetaria"}


class TestSoloLasFilasANUALES:
    def test_marzo_y_septiembre_no_son_el_ano(self):
        """Son cortes de la encuesta. Publicar uno como el año es el mismo error que tomar
        un anexo mensual por el ejercicio completo."""
        b = _libro(ZONA, LINEA, [
            ("2010", "Marzo", {7: 99.0, 8: 99.0}),
            ("", "Septiembre", {7: 88.0, 8: 88.0}),
            ("", "Año", {7: 16.76, 8: 50.20})])
        s, _ = parse_zona_rural(b)
        assert s["indigencia"] == [("2010", 16.76)]
        assert 99.0 not in dict(s["indigencia"]).values()

    def test_el_ano_se_arrastra_a_sus_cortes(self):
        """La celda del año solo aparece en la primera fila del bloque."""
        b = _libro(ZONA, LINEA, [("2010", "Marzo", {7: 1.0, 8: 2.0}),
                                 ("", "Año", {7: 16.76, 8: 50.2})])
        s, _ = parse_zona_rural(b)
        assert dict(s["indigencia"]) == {"2010": 16.76}


class TestNadaPorPosicion:
    def test_la_columna_RURAL_se_busca_por_etiqueta(self):
        """Con el bloque rural corrido, leer por posición devolvería la columna urbana. Las
        dos series se parecen lo suficiente como para que nadie lo note."""
        zona = {3: "Nacional", 6: "Urbana", 10: "Rural"}
        linea = {3: "Indigencia monetaria", 6: "Indigencia monetaria",
                 7: "Pobreza monetaria", 10: "Indigencia monetaria",
                 11: "Pobreza monetaria"}
        b = _libro(zona, linea, [("2010", "Año", {6: 4.5, 7: 24.0, 10: 16.76, 11: 50.2})])
        s, _ = parse_zona_rural(b)
        assert dict(s["indigencia"])["2010"] == 16.76, "tomó la urbana"

    def test_sin_las_cabeceras_LEVANTA(self):
        b = _libro({3: "Nacional"}, {3: "Total"}, [("2010", "Año", {3: 1.0})])
        with pytest.raises(SISDOMZonaError, match="cabeceras"):
            parse_zona_rural(b)

    def test_sin_la_hoja_no_se_inventa_otra(self):
        b = _libro(ZONA, LINEA, [("2010", "Año", {7: 1.0, 8: 2.0})], hoja="otra cosa")
        with pytest.raises(SISDOMZonaError, match="hoja"):
            parse_zona_rural(b)


class TestElQuiebreNoSeDisimula:
    """El emisor publica el año de solape dos veces, una por metodología. Encadenar los
    tramos sin declararlo es el defecto que ya se pagó con la informalidad laboral."""

    FILAS = [("2015", "Año", {7: 10.51, 8: 40.24}),
             ("2016", "Año", {7: 8.82, 8: 37.91}),
             ("2016*", "Año", {7: 8.64, 8: 40.11}),
             ("2017*", "Año", {7: 6.65, 8: 36.77})]

    def test_se_expone_el_ano_de_solape_y_el_salto(self):
        _, q = parse_zona_rural(_libro(ZONA, LINEA, self.FILAS))
        assert q["anio_de_solape"] == "2016"
        assert q["salto_pct"]["indigencia"] == pytest.approx(2.0, abs=0.1)
        assert q["salto_pct"]["pobreza"] == pytest.approx(5.8, abs=0.1)

    def test_manda_la_metodologia_VIGENTE_en_el_ano_repetido(self):
        """Igual que con las ediciones del anexo eléctrico: la reciente es la que el emisor
        sostiene hoy."""
        s, _ = parse_zona_rural(_libro(ZONA, LINEA, self.FILAS))
        assert dict(s["indigencia"])["2016"] == 8.64

    def test_el_ano_repetido_no_duplica_la_serie(self):
        s, _ = parse_zona_rural(_libro(ZONA, LINEA, self.FILAS))
        años = [a for a, _ in s["indigencia"]]
        assert len(años) == len(set(años)) == 3

    def test_sin_solape_el_quiebre_queda_en_None(self):
        """El contrapeso: no se inventa un quiebre donde el emisor no lo declara."""
        _, q = parse_zona_rural(_libro(ZONA, LINEA, [("2015", "Año", {7: 10.5, 8: 40.2})]))
        assert q["anio_de_solape"] is None and q["salto_pct"] == {}


def test_las_dos_lineas_del_emisor_estan_declaradas():
    assert set(LINEAS) == {"indigencia", "pobreza"}


# ── La zona es un PARÁMETRO, y elegir mal sirve la columna de al lado ──────────────────────

_FILAS_DOS_ZONAS = [("2010", "Año", {3: 10.79, 4: 40.53, 7: 16.76, 8: 50.20})]


def test_la_zona_nacional_y_la_rural_salen_del_mismo_cuadro_y_no_se_mezclan():
    """Los indicadores 2.1 y 2.4 son del país y el 2.3 y el 2.6 de la zona rural. Las cuatro
    columnas viven en el mismo cuadro y se parecen: leer la zona equivocada publica una cifra
    plausible contra la meta de otra población. Es el defecto que demotó al 2.1 y al 2.4."""
    from shared.data.sisdom_pobreza_zona import parse_zona

    libro = _libro(ZONA, LINEA, _FILAS_DOS_ZONAS)
    nac, _ = parse_zona(libro, "nacional")
    rur, _ = parse_zona(libro, "rural")
    assert dict(nac["indigencia"])["2010"] == 10.79
    assert dict(rur["indigencia"])["2010"] == 16.76
    assert dict(nac["pobreza"])["2010"] == 40.53
    assert dict(rur["pobreza"])["2010"] == 50.20


def test_una_zona_que_el_cuadro_no_publica_LEVANTA():
    """Pedir una zona inexistente tiene que fallar y no caer en la de al lado."""
    from shared.data.sisdom_pobreza_zona import parse_zona

    with pytest.raises(SISDOMZonaError, match="desconocida"):
        parse_zona(_libro(ZONA, LINEA, _FILAS_DOS_ZONAS), "regional")


def test_la_zona_por_defecto_sigue_siendo_la_rural():
    """`parse_zona_rural` existía antes de que la zona fuera parámetro; quien lo llamaba no
    tiene que enterarse del cambio."""
    from shared.data.sisdom_pobreza_zona import parse_zona

    libro = _libro(ZONA, LINEA, _FILAS_DOS_ZONAS)
    assert parse_zona_rural(libro)[0] == parse_zona(libro, "rural")[0]
