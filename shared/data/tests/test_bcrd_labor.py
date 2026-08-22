"""ENCFT del BCRD — parser de informalidad (lógica pura, sin red).

La planilla se arma acá con la forma REAL del libro del BCRD: ventanas móviles de
cuatro trimestres en el encabezado, de las cuales solo algunas son de año calendario, y
dos filas de informalidad que miden cosas distintas y se parecen peligrosamente.
"""
import io

import openpyxl
import pytest

from shared.data.bcrd_labor import (
    BcrdLaborUnavailable,
    INFORMALITY_LABEL,
    SHEET,
    SHEET_FEMALE as FEMENINO,
    SHEET_MALE as MASCULINO,
    parse_informality,
)


def _libro(filas, hoja=SHEET):
    wb = openpyxl.Workbook()
    wb.active.title = "Indicadores"          # otra hoja: NO debe ganar
    wb.active.append(["Ocupación Informal", 99.9])
    ws = wb.create_sheet(hoja)
    for f in filas:
        ws.append(f)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_ENCABEZADO = ["Condición", " III 2014 - II 2015", " I 2015 - IV 2015",
               " II 2015 - I 2016", " I 2016 - IV 2016"]


def test_solo_toma_las_ventanas_de_ano_calendario():
    """Las móviles no son anuales. Estamparlas con un año sería inventar una
    correspondencia que el emisor no declara."""
    contenido = _libro([
        ["Banco Central de la República Dominicana"],
        _ENCABEZADO,
        ["Sector Informal", 51.31, 51.24, 51.00, 50.00],
        ["Ocupación Informal", 57.57, 57.46, 57.53, 57.68],
    ])
    assert parse_informality(contenido) == [(2015, 57.46), (2016, 57.68)]


def test_no_confunde_ocupacion_informal_con_sector_informal():
    """Miden cosas distintas y corren ~6 puntos separadas: tomar la fila equivocada
    cambiaría el nivel de la serie sin que nada lo advirtiera."""
    contenido = _libro([
        _ENCABEZADO,
        ["Sector Informal", 51.31, 51.24, 51.00, 50.00],
        ["Ocupación Informal", 57.57, 57.46, 57.53, 57.68],
    ])
    assert [v for _y, v in parse_informality(contenido)] == [57.46, 57.68]
    assert [v for _y, v in parse_informality(contenido, label="Sector Informal")] == [51.24, 50.00]


def test_la_hoja_correcta_es_la_de_promedios_no_la_primera():
    """El libro trae la serie trimestral en 'Indicadores' y la anual en la de
    promedios. Leer la primera hoja daría una serie con el período equivocado."""
    contenido = _libro([_ENCABEZADO, ["Ocupación Informal", 1.0, 57.46, 2.0, 57.68]])
    assert parse_informality(contenido) == [(2015, 57.46), (2016, 57.68)]


def test_layout_cambiado_levanta_con_el_motivo():
    sin_ventanas = _libro([["Condición", "2015", "2016"],
                           ["Ocupación Informal", 57.46, 57.68]])
    with pytest.raises(BcrdLaborUnavailable, match="año calendario"):
        parse_informality(sin_ventanas)

    sin_fila = _libro([_ENCABEZADO, ["Otra cosa", 1, 2, 3, 4]])
    with pytest.raises(BcrdLaborUnavailable, match="no se encontró la fila"):
        parse_informality(sin_fila)

    with pytest.raises(BcrdLaborUnavailable, match="no trae la hoja"):
        parse_informality(_libro([_ENCABEZADO], hoja="Otra Hoja"))


def test_valores_fuera_de_rango_se_descartan_no_se_recortan():
    """Una tasa no puede ser 0 ni >100. Un valor así es un artefacto de lectura, y
    recortarlo lo disfrazaría de dato bueno."""
    contenido = _libro([
        _ENCABEZADO,
        ["Ocupación Informal", 57.57, 0, 57.53, 145.0],
    ])
    with pytest.raises(BcrdLaborUnavailable, match="rango"):
        parse_informality(contenido)


def test_la_etiqueta_por_defecto_es_la_del_empleo():
    assert INFORMALITY_LABEL == "ocupacion informal"


# ── Desocupación y brechas de género ──────────────────────────────────────────────────────

def _libro_por_sexo(fem_filas, mas_filas):
    """Libro con la forma REAL de las hojas por sexo: el año aparece UNA vez en el
    encabezado y se propaga a sus trimestres (celdas combinadas), 2014 arranca con dos
    trimestres y el año en curso viene incompleto."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "Indicadores"
    for hoja, filas in ((FEMENINO, fem_filas), (MASCULINO, mas_filas)):
        ws = wb.create_sheet(hoja)
        # III-2014, IV-2014 | 2015 completo | 2016 completo | 2017 con un solo trimestre
        ws.append(["Indicador", 2014, None, 2015, None, None, None,
                   2016, None, None, None, 2017])
        for f in filas:
            ws.append(f)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_la_razon_ignora_los_anos_incompletos():
    """2014 trae dos trimestres y el año en curso uno: promediarlos daría un punto que no
    es un año y que nadie podría reproducir."""
    from shared.data.bcrd_labor import parse_gender_ratio
    contenido = _libro_por_sexo(
        [["Tasa de Ocupación", 40, 40, 42, 42, 42, 42, 45, 45, 45, 45, 99]],
        [["Tasa de Ocupación", 80, 80, 84, 84, 84, 84, 90, 90, 90, 90, 99]])
    assert parse_gender_ratio(contenido, "Tasa de Ocupación") == [(2015, 0.5), (2016, 0.5)]


def test_la_razon_divide_los_promedios_anuales_y_no_promedia_razones():
    """Las dos cuentas dan distinto. Se elige dividir los promedios porque es la que
    reproduce cualquiera que tome las tasas anuales publicadas de cada sexo."""
    from shared.data.bcrd_labor import parse_gender_ratio
    # Trimestres F: 10,20,30,40 (media 25) · M: 20,20,20,100 (media 40) → 25/40 = 0,625.
    # El promedio de las razones trimestrales daría 0,825 — bien distinto.
    contenido = _libro_por_sexo(
        [["Tasa de Ocupación", 1, 1, 10, 20, 30, 40, 1, 1, 1, 1, 1]],
        [["Tasa de Ocupación", 1, 1, 20, 20, 20, 100, 1, 1, 1, 1, 1]])
    assert dict(parse_gender_ratio(contenido, "Tasa de Ocupación"))[2015] == 0.625


def test_una_hoja_por_sexo_ausente_falla_en_vez_de_devolver_media_serie():
    from shared.data.bcrd_labor import BcrdLaborUnavailable, parse_gender_ratio
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = FEMENINO
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(BcrdLaborUnavailable, match="Masculino"):
        parse_gender_ratio(buf.getvalue(), "Tasa de Ocupación")


def test_la_desocupacion_se_lee_de_SU1_y_no_de_la_fila_parecida():
    """«SU1: Tasa de Desocupación» y «Tasa de Desocupación (abiertos con iniciadores)»
    miden poblaciones distintas y corren a décimas de distancia: confundirlas cambiaría el
    nivel sin que nada lo advierta."""
    from shared.data.bcrd_labor import UNEMPLOYMENT_LABEL
    contenido = _libro([
        ["Banco Central de la República Dominicana"],
        _ENCABEZADO,
        ["Tasa de Desocupación (abiertos con iniciadores)", 7.99, 7.86, 7.80, 7.70],
        ["SU1: Tasa de Desocupación", 7.88, 7.33, 7.20, 7.08],
    ])
    assert parse_informality(contenido, label=UNEMPLOYMENT_LABEL) == [(2015, 7.33), (2016, 7.08)]


# ── Brecha regional (indicador 2.38 de la END) ─────────────────────────────────────────────

def _libro_regiones(bloques, hoja="Regiones"):
    """`bloques` = [(año, [Ozama, Cibao, Sur, Este], extras)] con la forma real del cuadro:
    el año en una fila sola y las cuatro medidas SU debajo, con `Total País` primero."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Indicadores"
    ws = wb.create_sheet(hoja)
    for anio, tasas, extras in bloques:
        ws.append([anio])
        ws.append(["Indicador", "Total País", "Región Ozama", "Región Norte",
                   "Región Sur", "Región Este"])
        ws.append(["SU1: Tasa de Desocupación", 7.3] + list(extras))
        ws.append(["SU2: Desocupación ampliada", 13.3] + list(tasas))
        ws.append(["Tasa de Inactividad", 38.1, 1, 2, 3, 4])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_la_brecha_es_maximo_menos_minimo_entre_las_cuatro_regiones():
    from shared.data.bcrd_labor import parse_regional_gap

    d = dict(parse_regional_gap(_libro_regiones(
        [(2020, [8.44, 10.42, 11.57, 11.64], [1, 2, 3, 4])])))
    assert d == {2020: round(11.64 - 8.44, 2)}


def test_NO_mete_el_total_pais_en_la_brecha():
    """`Total País` es un agregado, no una región. Contarlo produciría una brecha contra un
    promedio, que es OTRA definición — y la que el oráculo descartó por un factor."""
    from shared.data.bcrd_labor import parse_regional_gap

    # El total (13.3) queda por encima del máximo regional: si entrara, la brecha crecería.
    d = dict(parse_regional_gap(_libro_regiones([(2020, [8.0, 9.0, 10.0, 11.0], [1, 2, 3, 4])])))
    assert d == {2020: 3.0}


def test_lee_SU2_y_no_la_fila_de_al_lado():
    """SU1 es la desocupación ABIERTA y corre unos siete puntos por debajo. La ley nombra la
    ampliada: atar la fila vecina mediría otra población."""
    from shared.data.bcrd_labor import parse_regional_gap

    d = dict(parse_regional_gap(_libro_regiones(
        [(2020, [8.0, 9.0, 10.0, 11.0], [1.0, 1.5, 2.0, 9.0])])))
    assert d == {2020: 3.0}, "tomó la fila SU1, cuya brecha sería 8.0"


def test_un_ano_con_una_region_ausente_se_descarta():
    """Con tres regiones el máximo o el mínimo puede faltar, y la brecha saldría más chica
    sin que nada avise. Es la misma regla que el cociente sin denominador."""
    from shared.data.bcrd_labor import parse_regional_gap

    with pytest.raises(BcrdLaborUnavailable):
        parse_regional_gap(_libro_regiones([(2020, [8.0, 9.0, None, 11.0], [1, 2, 3, 4])]))


def test_sin_la_hoja_de_regiones_falla_con_el_motivo():
    from shared.data.bcrd_labor import parse_regional_gap

    with pytest.raises(BcrdLaborUnavailable, match="Regiones"):
        parse_regional_gap(_libro_regiones([(2020, [1, 2, 3, 4], [1, 2, 3, 4])], hoja="Otra"))


def test_el_ano_puede_venir_como_texto_o_como_numero():
    """El cuadro real usa las dos formas en el mismo libro."""
    from shared.data.bcrd_labor import parse_regional_gap

    d = dict(parse_regional_gap(_libro_regiones(
        [("2019", [1.0, 2.0, 3.0, 4.0], [1, 2, 3, 4]),
         (2020, [1.0, 2.0, 3.0, 5.0], [1, 2, 3, 4])])))
    assert d == {2019: 3.0, 2020: 4.0}
