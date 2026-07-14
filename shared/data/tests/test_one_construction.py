"""Tests del conector ONE de construcción por tipología (valor tasado / m² / licencias /
construcciones). Workbooks sintéticos que replican las dos orientaciones reales de la ONE
(filas=Tipo jerárquica 2020+ y filas=Mes plana 2013-2019); nunca tocan la red.
"""
import openpyxl
from openpyxl.styles import Alignment, Font

from shared.data.one_construction import (
    merge_metrics,
    parse_metric_workbook,
)


def _tipo_sheet(ws, groups):
    """Hoja orientación filas=Tipo: título + header + fila Total (en blanco, como 2020) +
    filas de grupo (indent 0, negrita) intercaladas con sub-líneas (indent 1)."""
    ws["A1"] = "Cuadro 4.10. Valor tasado ... según tipo de construcción, 2025*"
    ws["A2"] = "(En RD$)"
    hdr = ["Tipo de construcción", "Total", "Enero", "Feb"]
    for j, v in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=v)
    ws.cell(row=5, column=1, value="Total")  # celda Total en blanco a propósito
    r = 6
    for name, total, subs in groups:
        c = ws.cell(row=r, column=1, value=name)
        c.font = Font(bold=True)
        c.alignment = Alignment(indent=0)
        ws.cell(row=r, column=2, value=total)
        r += 1
        for sname, sval in subs:
            sc = ws.cell(row=r, column=1, value=sname)
            sc.font = Font(bold=False)
            sc.alignment = Alignment(indent=1)
            ws.cell(row=r, column=2, value=sval)
            r += 1


def _mes_sheet(ws, tipos_total, grand):
    """Hoja orientación filas=Mes: tipos en columnas, fila 'Total' con el gran total."""
    ws["A1"] = "Cuadro 4.10 ... 2013*"
    hdr = ["Mes", "Total"] + [t for t, _ in tipos_total]
    for j, v in enumerate(hdr, start=1):
        ws.cell(row=4, column=j, value=v)
    ws.cell(row=5, column=1, value="Enero")
    ws.cell(row=6, column=1, value="Total")
    ws.cell(row=6, column=2, value=grand)
    for j, (_, v) in enumerate(tipos_total, start=3):
        ws.cell(row=6, column=j, value=v)


def _wb_tipo_year(year, groups):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _tipo_sheet(wb.create_sheet(str(year)), groups)
    return wb


def test_parse_tipo_orientation_groups_only_total_from_sum():
    # Grupo con sub-líneas: solo el grupo cuenta; el total se DERIVA de la suma de grupos
    # (la celda 'Total' del proveedor está en blanco, como en 2020).
    wb = _wb_tipo_year(2025, [
        ("Edificios de Apartamentos", 45789879594.0,
         [("Apartamentos", 44310860330.0), ("Apart. Turístico", 1479019264.0)]),
        ("Comerciales y Oficinas", 5127690356.0, []),
    ])
    out = parse_metric_workbook(wb)
    assert set(out) == {2025}
    bt = out[2025]["by_typology"]
    assert bt == {"Edificios de Apartamentos": 45789879594.0,
                  "Comerciales y Oficinas": 5127690356.0}  # sub-líneas excluidas
    assert out[2025]["total"] == 45789879594.0 + 5127690356.0  # suma de grupos


def test_parse_mes_orientation_total_only():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _mes_sheet(wb.create_sheet("2013"), [("Almacén", 10.0), ("Apartamento", 90.0)], grand=100.0)
    out = parse_metric_workbook(wb)
    assert out[2013]["by_typology"] is None       # taxonomía plana vieja → no se desagrega
    assert out[2013]["total"] == 100.0


def test_merge_normalizes_casing_and_scopes_detail_to_tipo_years():
    # 'valor' y 'sqm' difieren en casing de una etiqueta → misma clave canónica; el display
    # canónico lo fija el archivo de valor tasado. 2013 (mes) solo aporta total, no detalle.
    valor = {
        2025: {"total": 5127690356.0,
               "by_typology": {"Comerciales y Oficinas": 5127690356.0}},
        2013: {"total": 8732229958.0, "by_typology": None},
    }
    sqm = {2025: {"total": 343086.0, "by_typology": {"comerciales y oficinas": 343086.0}}}
    licencias = {2025: {"total": 149.0, "by_typology": {"Comerciales y Oficinas": 149.0}}}
    construcciones = {2025: {"total": 113.0, "by_typology": {"Comerciales y Oficinas": 113.0}}}

    merged = merge_metrics({"valor_tasado": valor, "sqm": sqm,
                            "licencias": licencias, "construcciones": construcciones})
    assert merged["typology_years"] == [2025]        # 2013 (mes) no aporta detalle
    assert set(merged["years"]) == {2013, 2025}
    row = merged["by_typology"][2025]["Comerciales y Oficinas"]  # display del archivo de valor
    assert row == {"licencias": 149.0, "construcciones": 113.0,
                   "sqm": 343086.0, "valor_tasado": 5127690356.0}
    # totales por métrica cubren todos los años, incluido el viejo (solo total)
    assert merged["totals"][2013]["valor_tasado"] == 8732229958.0
    assert merged["totals"][2025]["sqm"] == 343086.0


def test_merge_requires_some_detail_year():
    # Si ningún año trae detalle (todo mes/total), no hay typology_years pero sí totales.
    only_totals = {2013: {"total": 100.0, "by_typology": None}}
    merged = merge_metrics({"valor_tasado": only_totals})
    assert merged["typology_years"] == []
    assert merged["totals"][2013]["valor_tasado"] == 100.0
