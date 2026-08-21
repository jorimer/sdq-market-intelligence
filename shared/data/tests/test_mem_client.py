"""El anexo del MEM: solo diciembre, nada por posición, y el porcentaje que el emisor muestra.

Las tres reglas de este parser salen de defectos posibles y baratos de cometer:

* las columnas de un anexo que no es de diciembre son ACUMULADAS del año en curso, así que
  cuatro meses se publicarían con el rótulo de doce;
* el anexo cambia entre ediciones —de 2021 a 2024 se agregaron hojas y el bloque financiero
  se corrió una fila—, y un parser por índice devuelve el número de al lado sin fallar;
* las celdas están guardadas como fracción y mostradas como porcentaje.
"""
from pathlib import Path

import pytest

from shared.data.mem_client import (FILAS, MEMUnavailable, parse_anexo, series_anuales)

FIXTURE = Path(__file__).resolve().parents[1] / "fixtures" / "mem_anexo_diciembre_2024.xlsx"


def _libro(cabecera, etiquetas):
    """Un anexo mínimo con la forma real: cabecera de años en f6, etiquetas en la columna 2."""
    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EDE's"
    for i, texto in enumerate(cabecera):
        ws.cell(6, 3 + i).value = texto
    for fila, (etiqueta, valores) in enumerate(etiquetas.items(), start=20):
        ws.cell(fila, 2).value = etiqueta
        for i, v in enumerate(valores):
            ws.cell(fila, 3 + i).value = v
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


TODAS = {e: (0.5, 0.6) for e in FILAS}


class TestSoloDiciembre:
    def test_una_ventana_acumulada_no_es_un_ano(self):
        """`Ene24-Abr24` son cuatro meses. Publicarlos como el año no es una aproximación."""
        with pytest.raises(MEMUnavailable, match="año completo"):
            parse_anexo(_libro(["Ene24-Abr24", "Ene23-Abr23"], TODAS))

    def test_el_ano_completo_entra(self):
        d = parse_anexo(_libro(["Ene24-Dic24", "Ene23-Dic23"], TODAS))
        assert set(d["cri"]) == {2024, 2023}

    def test_se_ignora_la_columna_que_no_es_un_ano(self):
        """Entre los años reales hay columnas de comparación; no son observaciones."""
        d = parse_anexo(_libro(["Ene24-Dic24", "Comparación 2023 - 2024", "Ene22-Dic22"],
                               {e: (0.5, 999.0, 0.6) for e in FILAS}))
        assert set(d["cri"]) == {2024, 2022}
        assert 999.0 not in d["cri"].values() and 99900.0 not in d["cri"].values()


class TestNadaPorPosicion:
    def test_una_etiqueta_ausente_levanta_en_vez_de_servir_a_medias(self):
        faltando = {e: (0.5,) for e in list(FILAS)[:2]}
        with pytest.raises(MEMUnavailable, match="etiqueta"):
            parse_anexo(_libro(["Ene24-Dic24"], faltando))

    def test_la_fila_se_encuentra_aunque_se_mueva(self):
        """Es lo que protege contra la edición siguiente del anexo."""
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EDE's"
        ws.cell(6, 3).value = "Ene24-Dic24"
        for fila, etiqueta in zip((90, 150, 210), FILAS):
            ws.cell(fila, 2).value = etiqueta
            ws.cell(fila, 3).value = 0.42
        buf = io.BytesIO()
        wb.save(buf)
        d = parse_anexo(buf.getvalue())
        assert all(s[2024] == 42.0 for s in d.values())

    def test_sin_la_hoja_de_las_distribuidoras_no_se_inventa_otra(self):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.title = "Otra cosa"
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(MEMUnavailable, match="hoja"):
            parse_anexo(buf.getvalue())


def test_el_valor_publicado_es_el_PORCENTAJE_que_muestra_el_emisor():
    """La celda guarda 0,3763 y el informe del MEM dice 37,6%. La ley escribe sus metas en
    esa magnitud, así que servir la fracción invertiría el veredicto por un factor de cien —
    el mismo error que el GINI obligó a declarar como transformación."""
    d = parse_anexo(_libro(["Ene24-Dic24"], {e: (0.3763,) for e in FILAS}))
    assert d["perdidas"][2024] == pytest.approx(37.63, abs=0.01)


class TestContraElAnexoReal:
    """La fixture es un recorte del anexo de diciembre de 2024 publicado por el MEM."""

    @staticmethod
    @pytest.fixture(scope="class")
    def real():
        return parse_anexo(FIXTURE.read_bytes())

    def test_trae_los_tres_anos_del_anexo(self, real):
        assert set(real["cri"]) == {2022, 2023, 2024}

    def test_las_cifras_son_las_del_emisor(self, real):
        assert real["cri"][2024] == pytest.approx(59.45, abs=0.01)
        assert real["perdidas"][2024] == pytest.approx(37.64, abs=0.01)
        assert real["cobranzas"][2024] == pytest.approx(95.32, abs=0.01)

    def test_las_tres_series_estan_completas(self, real):
        assert set(real) == set(FILAS.values())
        assert all(len(s) == 3 for s in real.values())


class TestUnionDeEdiciones:
    def test_manda_la_edicion_mas_nueva(self):
        """El emisor revisa sus cifras; la edición reciente es la que sostiene hoy."""
        viejo = _libro(["Ene22-Dic22"], {e: (0.60,) for e in FILAS})
        nuevo = _libro(["Ene22-Dic22"], {e: (0.63,) for e in FILAS})
        s = series_anuales([2022, 2024],
                           listar=lambda a: {2022: "v", 2024: "n"},
                           bajar=lambda u: viejo if u == "v" else nuevo)
        assert s["cri"] == [("2022", 63.0)]

    def test_sin_anexo_alcanzable_no_devuelve_serie_vacia(self):
        """Una serie vacía se leería como «el sector no tiene dato». Es otra cosa."""
        with pytest.raises(MEMUnavailable, match="anexo de diciembre"):
            series_anuales([2024], listar=lambda a: {}, bajar=lambda u: b"")

    def test_las_series_salen_ascendentes(self):
        a = _libro(["Ene24-Dic24", "Ene23-Dic23"], {e: (0.59, 0.61) for e in FILAS})
        s = series_anuales([2024], listar=lambda x: {2024: "u"}, bajar=lambda u: a)
        assert [p for p, _ in s["cri"]] == ["2023", "2024"]
