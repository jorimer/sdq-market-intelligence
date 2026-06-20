"""Tests for the Hacienda Estado de Operaciones connector (fiscal pulse)."""
import io

import openpyxl
import pytest

from shared.data.hacienda_client import (
    SHEET,
    HaciendaClient,
    HaciendaError,
    build_records,
    discover_eo_url,
    parse_eo_xlsx,
)

_MONTHS = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
           "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
# (GFS code, label, monthly value) — a minimal but structurally faithful slice.
_LINES = [
    ("1", "I. Ingresos", 1000.0),
    ("11", "Impuestos", 900.0),
    ("111", "Impuestos sobre el ingreso", 500.0),
    ("2", "II. Gastos", 800.0),
    ("21", "Remuneración a los empleados", 300.0),
    ("24", "Intereses", 100.0),
    ("31", "III. Inversión neto en activos no financieros", 150.0),
]


def _toy_workbook(years=("2024", "2025"), lines=None) -> bytes:
    """Build an Excel mirroring the EO sheet: year row + month row + GFS code/label rows."""
    lines = lines if lines is not None else _LINES
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Ministerio de Hacienda"])
    ws.append(["Estado de Operaciones del Sector Público No Financiero"])
    # year row + month row, both starting at column index 2 (after code + label)
    year_row = [None, None]
    month_row = [None, None]
    for y in years:
        year_row += [int(y)] * 12
        month_row += list(_MONTHS)
    ws.append(year_row)
    ws.append(month_row)
    for code, label, val in lines:
        ws.append([code, label] + [val] * (12 * len(years)))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_maps_periods_and_curated_lines():
    groups = parse_eo_xlsx(_toy_workbook())
    assert "ingresos" in groups and "gastos" in groups and "inversion" in groups
    # 2 years × 12 months
    assert sorted(groups["ingresos"])[0] == "2024-01"
    assert sorted(groups["ingresos"])[-1] == "2025-12"
    assert len(groups["ingresos"]) == 24
    assert groups["ingresos"]["2024-03"] == 1000.0
    assert groups["imp_ingresos"]["2025-01"] == 500.0


def test_period_columns_ignores_stray_year_caption():
    """A caption row with scattered years must NOT beat the aligned year header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Cobertura 2000 2005 2010 2015 2020 2024 2025", 2000, 2005, 2010, 2015, 2020, 2024, 2025])
    year_row, month_row = [None, None], [None, None]
    for y in ("2024", "2025"):
        year_row += [int(y)] * 12
        month_row += list(_MONTHS)
    ws.append(year_row)
    ws.append(month_row)
    for code, label, val in _LINES:
        ws.append([code, label] + [val] * 24)
    buf = io.BytesIO()
    wb.save(buf)
    groups = parse_eo_xlsx(buf.getvalue())
    assert sorted(groups["ingresos"])[0] == "2024-01"     # aligned row won, not the caption
    assert sorted(groups["ingresos"])[-1] == "2025-12"


def test_parse_computes_balances():
    groups = parse_eo_xlsx(_toy_workbook())
    # resultado_operativo = ingresos − gastos; balance_global = − inversión
    assert groups["resultado_operativo"]["2024-01"] == pytest.approx(1000 - 800)
    assert groups["balance_global"]["2024-01"] == pytest.approx(1000 - 800 - 150)


def test_parse_fails_closed_without_headline_lines():
    lines = [ln for ln in _LINES if ln[0] != "2"]  # drop Gastos
    with pytest.raises(HaciendaError, match="Ingresos/Gastos"):
        parse_eo_xlsx(_toy_workbook(lines=lines))


def test_parse_fails_closed_without_sheet():
    wb = openpyxl.Workbook()
    wb.active.title = "Otra"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(HaciendaError, match="ausente"):
        parse_eo_xlsx(buf.getvalue())


def test_build_records_emits_lines_and_computed():
    recs = build_records(parse_eo_xlsx(_toy_workbook()))
    dims = {r.dimension for r in recs}
    assert {"ingresos", "gastos", "resultado_operativo", "balance_global"}.issubset(dims)
    assert all(r.series == "fiscal" and r.lineage.source == "Hacienda" for r in recs)


def test_discover_eo_url_picks_central_mensual_latest():
    html = """
      <a href="https://www.hacienda.gob.do/wp-content/uploads/2024/04/Gobierno-central-presupuestario-estado-de-operaciones-mensual-RD-y-US-9.xlsx">old</a>
      <a href="https://www.hacienda.gob.do/wp-content/uploads/2025/07/Gobierno-central-presupuestario-estado-de-operaciones-mensual-RD-y-US-10.xlsx">new</a>
      <a href="https://www.hacienda.gob.do/wp-content/uploads/2025/07/Gobierno-central-presupuestario-estado-de-operaciones-trimestral-RD-y-US-6.xlsx">qtr</a>
    """
    url = discover_eo_url(html)
    assert url is not None and "mensual" in url and "/2025/07/" in url  # latest mensual wins
    assert "trimestral" not in url


def test_fixture_mode_real_snapshot():
    recs = HaciendaClient(mode="fixture").fetch()
    assert recs, "fixture vacío"
    dims = {r.dimension for r in recs}
    assert {"ingresos", "gastos", "resultado_operativo", "balance_global"}.issubset(dims)
    # the fiscal identity holds in the committed real data
    by = {}
    for r in recs:
        by.setdefault(r.period, {})[r.dimension] = r.value
    p = max(by)
    assert by[p]["resultado_operativo"] == pytest.approx(by[p]["ingresos"] - by[p]["gastos"], abs=0.1)
    one = HaciendaClient(mode="fixture").fetch(series="fiscal", period=p)
    assert one and all(r.period == p for r in one)
