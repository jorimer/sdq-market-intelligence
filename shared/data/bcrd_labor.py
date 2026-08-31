"""ENCFT del Banco Central — informalidad laboral desde la FUENTE PRIMARIA.

Sustituye el raspado de la landing de la ONE, que dejó de funcionar cuando el portal
quedó detrás de un desafío de Cloudflare (403 desde cualquier cliente que no sea un
navegador, también desde producción). La reacción intuitiva sería pelear con el
desafío; la correcta era preguntarse **quién produce el dato**. La ONE no lo produce:
lo republica. La ENCFT es del BCRD, que publica sus tablas en el mismo CDN público del
que este repo ya baja otras series —sin token, sin navegador y sin User-Agent fingido.

El cambio deja el dato MEJOR de lo que estaba:

* fuente primaria en vez de republicación (un intermediario menos);
* cobertura 2014-2026 en vez de 2004-2024;
* la tasa viene **publicada por el emisor**, no calculada por nosotros.

Se toman las ventanas de AÑO CALENDARIO (``I YYYY - IV YYYY``) de la hoja de promedios
móviles: son las únicas comparables con una serie anual. Usar una ventana móvil
cualquiera y estamparla con un año sería inventar una correspondencia.

Nota de alcance: acá SOLO viaja la informalidad. El BCRD publica los deciles de ingreso
como CONTEOS de perceptores, no como montos, así que el ingreso laboral promedio —que
la ONE sí calculaba— no tiene sustituto en esta fuente. Declararlo es preferible a
rellenarlo con algo parecido.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger("sdq.data.bcrd_labor")

CDN = ("https://cdn.bancentral.gov.do/documents/estadisticas/"
       "mercado-de-trabajo/documents/")
INDICATORS_FILE = "00_Indicadores.xlsx"
INDICATORS_URL = CDN + INDICATORS_FILE

SOURCE = "BCRD"
LICENSE = "estadísticas públicas del Banco Central de la República Dominicana — uso con cita"
SHEET = "Promedio 4 Trimestres"

# Fila del indicador. "Ocupación Informal" es la informalidad DEL EMPLEO (la que la ONE
# republicaba como «tasa de informalidad en el empleo»); "Sector Informal" mide otra cosa
# —el peso del sector— y corre ~6 puntos más abajo. Confundirlas cambiaría el nivel de la
# serie sin que nada lo advirtiera, así que la etiqueta se exige explícita.
INFORMALITY_LABEL = "ocupacion informal"

# Ventana de año calendario: "I 2015 - IV 2015". El resto son móviles y no son anuales.
_CALENDAR_WINDOW = re.compile(r"^\s*I\s+((?:19|20)\d{2})\s*-\s*IV\s+((?:19|20)\d{2})\s*$")
_HEADER_SCAN_ROWS = 12


class BcrdLaborUnavailable(RuntimeError):
    """No se pudo leer la serie. Lleva el motivo, no solo el hecho."""


def _norm(s: object) -> str:
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return " ".join(t.casefold().split())


def parse_informality(content: bytes,
                      label: str = INFORMALITY_LABEL) -> List[Tuple[int, float]]:
    """Parsea ``00_Indicadores.xlsx`` → ``[(año, tasa_de_informalidad)]`` ascendente.

    Localiza la columna por el ENCABEZADO de ventana (no por posición) y la fila por su
    etiqueta, igual que el resto de los parsers de planilla del repo: así un cambio de
    layout deja la serie vacía —visible— en vez de leer la fila equivocada en silencio.
    Pura (sin red), para poder fijarla en tests."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if SHEET not in wb.sheetnames:
            raise BcrdLaborUnavailable(
                f"el libro no trae la hoja '{SHEET}' (hojas: {wb.sheetnames})")
        rows = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]
    finally:
        wb.close()

    # 1) columnas de año calendario, desde la fila de encabezado.
    year_by_col: dict = {}
    for r in rows[:_HEADER_SCAN_ROWS]:
        for ci, cell in enumerate(r):
            m = _CALENDAR_WINDOW.match(cell) if isinstance(cell, str) else None
            if m and m.group(1) == m.group(2):
                year_by_col[ci] = int(m.group(1))
        if year_by_col:
            break
    if not year_by_col:
        raise BcrdLaborUnavailable(
            "no se encontró ninguna ventana de año calendario ('I YYYY - IV YYYY') "
            "en el encabezado (¿cambió el layout del BCRD?)")

    # 2) la fila del indicador, por etiqueta exacta normalizada.
    target = _norm(label)
    data_row: Optional[list] = None
    for r in rows:
        if r and _norm(r[0]) == target:
            data_row = r
            break
    if data_row is None:
        raise BcrdLaborUnavailable(
            f"no se encontró la fila '{label}' en la hoja '{SHEET}'")

    out: List[Tuple[int, float]] = []
    for ci, year in sorted(year_by_col.items(), key=lambda kv: kv[1]):
        v = data_row[ci] if ci < len(data_row) else None
        if isinstance(v, (int, float)) and 0 < float(v) <= 100:
            out.append((year, round(float(v), 2)))
    if not out:
        raise BcrdLaborUnavailable(
            f"la fila '{label}' no trae ningún valor en rango 0-100")
    return out


SHEET_TRIMESTRAL = "Indicadores"
_ROMANOS = {"I": 1, "II": 2, "III": 3, "IV": 4}


def _fila_de_trimestres(rows: List[list]) -> Optional[int]:
    """La fila cuyo contenido son numerales romanos de trimestre, buscada por CONTENIDO.

    No por índice: si el BCRD agrega una fila de título arriba, una posición fija leería
    otra cosa en silencio. Es la misma regla que el resto de los parsers de este archivo.
    """
    for i, r in enumerate(rows[:_HEADER_SCAN_ROWS]):
        romanos = sum(1 for c in r[1:] if _trimestre(c) is not None)
        if romanos >= 4:
            return i
    return None


def _trimestre(cell: object) -> Optional[int]:
    """`'III'` → 3. Tolera la marca de nota al pie que el BCRD pega al trimestre en curso
    (`'I 1/'`): el número del trimestre no cambia porque el dato sea preliminar."""
    if not isinstance(cell, str):
        return None
    return _ROMANOS.get(cell.strip().split()[0].upper() if cell.strip() else "")


_NOTA_AL_PIE = re.compile(r"\s*\d+/\s*$")


def _sin_nota(txt: str) -> str:
    """Quita la marca de nota al pie que el BCRD pega al final de algunas etiquetas."""
    return _NOTA_AL_PIE.sub("", txt or "").strip()


SHEET_PRECISION = "Precisión Estadística Indicador"

#: De la etiqueta de la hoja «Indicadores» a la de la hoja de PRECISIÓN. Son dos vocabularios
#: del mismo emisor para la misma serie —«SU1: Tasa de Desocupación» allá,
#: «Tasa de desocupación (SU1)» acá— y se declara el puente en vez de emparejarlos por
#: parecido: un emparejamiento laxo pegaría la precisión de una serie a los valores de otra,
#: y el resultado se vería perfectamente normal.
PRECISION_POR_ETIQUETA = {
    "SU1: Tasa de Desocupación": "Tasa de desocupación (SU1)",
    "SU2: Desocupación y Subocupación": "Desocupación y Subocupación (SU2)",
    "SU3: Desocupación y Fuerza de Trabajo Potencial":
        "Desocupación y Fuerza de Trabajo Potencial (SU3)",
    "SU4: Desocupación + Subocupación + Fuerza de Trabajo Potencial":
        "Desocupación + Subocupación + Fuerza de Trabajo Potencial (SU4)",
    "Ocupación Informal": "Ocupación Informal",
    "Ocupación Formal": "Ocupación Formal",
    "Tasa de Ocupación": "Tasa de ocupación",
    "Tasa de subocupación por horas": "Tasa de subocupación por horas",
    "Tasa de Inactividad": "Tasa de inactividad",
}


def _anio(cell: object) -> Optional[int]:
    """El año de una celda, venga como número o como texto.

    Existe porque la hoja de precisión mezcla los dos tipos en la MISMA columna. Un parser
    que solo acepte el numérico deja de propagar el año en los bloques donde vino como texto
    y devuelve una serie vacía o —peor— con el año del bloque anterior."""
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        return int(cell) if 2000 < cell < 2100 else None
    if isinstance(cell, str):
        t = cell.strip()
        if t.isdigit() and 2000 < int(t) < 2100:
            return int(t)
    return None


#: Qué series servidas llevan su precisión. Son las que un informe cita, no las nueve: la
#: precisión existe para juzgar una AFIRMACIÓN, y las que no se afirman no la necesitan.
_PRECISION_DE_LAS_SERVIDAS = {
    "unemployment_rate_trimestral": "SU1: Tasa de Desocupación",
    "underutilization_su4_trimestral":
        "SU4: Desocupación + Subocupación + Fuerza de Trabajo Potencial",
    "informality_rate_trimestral": "Ocupación Informal",
    "underemployment_rate_trimestral": "Tasa de subocupación por horas",
}


def parse_precision(content: bytes, label: str) -> List[Tuple[str, Dict[str, Optional[float]]]]:
    """La PRECISIÓN de cada estimación trimestral: `[('YYYY-Qn', {...})]`.

    **Por qué importa.** La ENCFT es una ENCUESTA: cada cifra que publicamos es una
    estimación con error de muestreo, y el BCRD publica el error estándar, el intervalo de
    confianza al 95% y el coeficiente de variación de cada una — en dos hojas del mismo libro
    que ya descargamos. Servíamos las estimaciones sin ninguna medida de precisión.

    Para qué sirve en un informe: una diferencia menor que los intervalos NO es una
    diferencia. Sin el intervalo, el modelo (y el lector) leen como señal lo que puede ser
    ruido de muestreo, y eso es exactamente lo que la doctrina de esta casa llama ordenar lo
    que no es comparable.

    El coeficiente de variación es el resumen operativo: por convención estadística, hasta
    15% la estimación es fiable, entre 15 y 25 se usa con cautela y por encima de 25 no se
    publica. Se sirve el número, no el veredicto: el umbral lo aplica quien lee.
    """
    import openpyxl

    etiqueta = PRECISION_POR_ETIQUETA.get(label, label)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if SHEET_PRECISION not in wb.sheetnames:
            raise BcrdLaborUnavailable(
                f"el libro no trae la hoja '{SHEET_PRECISION}' (hojas: {wb.sheetnames})")
        rows = [list(r) for r in wb[SHEET_PRECISION].iter_rows(values_only=True)]
    finally:
        wb.close()

    objetivo = _norm(_sin_nota(etiqueta))
    out: List[Tuple[str, Dict[str, Optional[float]]]] = []
    # Indicador y año vienen en celdas COMBINADAS: aparecen una vez y valen hasta el próximo.
    # Se propagan explícitamente; leerlos fila por fila daría un `None` en casi todas.
    indicador, anio, dentro = "", None, False
    for fila in rows:
        if not fila:
            continue
        if isinstance(fila[0], str) and fila[0].strip():
            indicador = _norm(_sin_nota(fila[0]))
            dentro = indicador == objetivo
        # El año viene como NÚMERO en unas filas y como TEXTO en otras: la misma columna de
        # la misma hoja mezcla los dos tipos. Leer solo el numérico dejaba `anio` en None
        # durante bloques enteros y la serie salía vacía sin error — el parser «funcionaba».
        y = _anio(fila[1]) if len(fila) > 1 else None
        if y is not None:
            anio = y
        if not dentro or anio is None or len(fila) < 8:
            continue
        q = _trimestre(fila[2])
        if q is None or not isinstance(fila[3], (int, float)):
            continue
        out.append((f"{anio}-Q{q}", {
            "estimacion": round(float(fila[3]), 4),
            "error_estandar": round(float(fila[4]), 4) if isinstance(fila[4], (int, float)) else None,
            "ic95_inferior": round(float(fila[5]), 4) if isinstance(fila[5], (int, float)) else None,
            "ic95_superior": round(float(fila[6]), 4) if isinstance(fila[6], (int, float)) else None,
            "coeficiente_de_variacion_pct": round(float(fila[7]), 4) if isinstance(fila[7], (int, float)) else None,
        }))
    if not out:
        raise BcrdLaborUnavailable(
            f"no se encontró '{etiqueta}' en la hoja '{SHEET_PRECISION}'")
    return out


def parse_trimestral(content: bytes, label: str) -> List[Tuple[str, float]]:
    """``00_Indicadores.xlsx`` hoja «Indicadores» → ``[('YYYY-Qn', valor)]`` ascendente.

    **Por qué existe, además de la serie anual.** `parse_informality` lee la hoja «Promedio 4
    Trimestres», que son las ventanas anuales que el PROPIO BCRD calcula, y esa serie sostiene
    los indicadores de la END, que son anuales. Pero el crédito se mide por trimestre: para
    leer el deterioro de una cartera contra el mercado laboral hace falta la misma cadencia,
    y está publicada en otra hoja del mismo libro que ya descargamos.

    Las dos series conviven a propósito. La anual NO se deriva de ésta ni al revés: son del
    emisor, con su propia definición de ventana, y promediar cuatro trimestres para
    reproducir la anual daría un número que el BCRD no publicó.

    El año va en una fila con celdas combinadas (aparece una vez y se propaga) y el trimestre
    en la fila de abajo. El trimestre se lee del ROMANO, nunca de la posición: el libro
    arranca en III-2014, así que numerar por orden etiquetaría III como I.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if SHEET_TRIMESTRAL not in wb.sheetnames:
            raise BcrdLaborUnavailable(
                f"el libro no trae la hoja '{SHEET_TRIMESTRAL}' (hojas: {wb.sheetnames})")
        rows = [list(r) for r in wb[SHEET_TRIMESTRAL].iter_rows(values_only=True)]
    finally:
        wb.close()

    fila_q = _fila_de_trimestres(rows)
    if fila_q is None or fila_q == 0:
        raise BcrdLaborUnavailable(
            "no se encontró la fila de trimestres (I/II/III/IV) en la hoja "
            f"'{SHEET_TRIMESTRAL}' (¿cambió el layout del BCRD?)")
    fila_anio = rows[fila_q - 1]

    # El año se propaga hacia la derecha desde la celda combinada que lo declara.
    periodo_por_col: dict = {}
    anio: Optional[int] = None
    for ci, c in enumerate(rows[fila_q]):
        if ci < len(fila_anio) and isinstance(fila_anio[ci], (int, float)) \
                and 2000 < fila_anio[ci] < 2100:
            anio = int(fila_anio[ci])
        q = _trimestre(c)
        if q is not None and anio is not None:
            periodo_por_col[ci] = f"{anio}-Q{q}"

    if not periodo_por_col:
        raise BcrdLaborUnavailable(
            f"la hoja '{SHEET_TRIMESTRAL}' no produjo ningún período (año + trimestre)")

    # La hoja trimestral pega la MARCA DE NOTA AL PIE a la etiqueta («SU1: Tasa de
    # Desocupación 4/») y la anual no. Se quita la marca y se compara EXACTO — no por
    # prefijo: en esta hoja «Ocupados Informales» (un conteo, en millones) convive con
    # «Ocupación Informal» (una tasa), y un prefijo laxo puede tomar la fila equivocada.
    objetivo = _norm(_sin_nota(label))
    fila = next((r for r in rows
                 if r and isinstance(r[0], str) and _norm(_sin_nota(r[0])) == objetivo), None)
    if fila is None:
        raise BcrdLaborUnavailable(
            f"no se encontró la fila '{label}' en la hoja '{SHEET_TRIMESTRAL}'")

    out: List[Tuple[str, float]] = []
    for ci, periodo in sorted(periodo_por_col.items()):
        v = fila[ci] if ci < len(fila) else None
        if isinstance(v, (int, float)) and 0 < float(v) <= 100:
            out.append((periodo, round(float(v), 2)))
    if not out:
        raise BcrdLaborUnavailable(
            f"la fila '{label}' no trae ningún valor en rango 0-100 en la hoja trimestral")
    return out


def fetch_bcrd_informality() -> List[Tuple[int, float]]:  # pragma: no cover - network I/O
    """Live: descarga el libro de indicadores del CDN del BCRD y lo parsea."""
    import httpx

    try:
        r = httpx.get(INDICATORS_URL, timeout=120, follow_redirects=True)
        r.raise_for_status()
    except httpx.HTTPError as e:
        raise BcrdLaborUnavailable(
            f"no se pudo descargar {INDICATORS_FILE} del CDN del BCRD "
            f"({type(e).__name__}: {e})")
    return parse_informality(r.content)


# ─────────────────────────────────────────────────────────────────────────────
# Desocupación y brechas de género
#
# El mismo libro trae, además de la informalidad, la tasa de desocupación y —en hojas
# aparte— el corte por sexo. Los tres indicadores que la END 1-12 fija sobre este mercado
# salen de acá, pero NO con la misma calidad de evidencia, y la diferencia hay que
# declararla porque decide qué se puede publicar:
#
#   · La tasa de desocupación TOTAL viene publicada por el emisor como promedio de cuatro
#     trimestres de año calendario. Es el dato del BCRD, sin cuenta nuestra.
#   · Las hojas por sexo son TRIMESTRALES: el emisor no publica su promedio anual. El
#     promedio lo calculamos acá, y por eso va declarado en el nombre de la función y en
#     la procedencia — no se puede servir como «cifra del BCRD» algo que el BCRD no
#     publicó.
# ─────────────────────────────────────────────────────────────────────────────

#: Fila de la tasa de desocupación. Es "SU1" y no "Tasa de Desocupación (abiertos con
#: iniciadores)", que corre unas décimas más arriba y mide otra población. La etiqueta se
#: exige explícita por la misma razón que la informalidad: dos filas parecidas que miden
#: cosas distintas.
UNEMPLOYMENT_LABEL = "SU1: Tasa de Desocupación"
EMPLOYMENT_RATE_LABEL = "Tasa de Ocupación"

SHEET_MALE = "Masculino"
SHEET_FEMALE = "Femenino"

#: Trimestres que tiene que traer un año para que su promedio sea comparable con los
#: demás. El libro arranca en III-2014 (dos trimestres) y el año en curso viene incompleto:
#: promediarlos igual produciría un punto que no es un año y que nadie podría reproducir.
_TRIMESTRES_POR_ANIO = 4


def _anual_por_sexo(rows: List[list], label: str) -> dict:
    """`{año: promedio de sus cuatro trimestres}` de una hoja por sexo.

    Las columnas se atribuyen por el año del encabezado, que aparece UNA vez y se
    propaga a los trimestres que le siguen (celdas combinadas). Los años incompletos se
    descartan en vez de promediarse con lo que haya.
    """
    encabezado = None
    for r in rows[:40]:
        if r and _norm(r[0]) in ("indicador", "condicion"):
            if any(isinstance(c, (int, float)) and 2000 < c < 2100 for c in r[1:]):
                encabezado = r
                break
    if encabezado is None:
        raise BcrdLaborUnavailable(
            "no se encontró el encabezado de años en la hoja por sexo "
            "(¿cambió el layout del BCRD?)")

    anio_por_col: dict = {}
    actual: Optional[int] = None
    for ci, c in enumerate(encabezado):
        if isinstance(c, (int, float)) and 2000 < c < 2100:
            actual = int(c)
        if ci and actual is not None:
            anio_por_col[ci] = actual

    objetivo = _norm(label)
    fila = next((r for r in rows if r and _norm(r[0]).startswith(objetivo)), None)
    if fila is None:
        raise BcrdLaborUnavailable(f"no se encontró la fila '{label}' en la hoja por sexo")

    por_anio: dict = {}
    for ci, anio in anio_por_col.items():
        v = fila[ci] if ci < len(fila) else None
        if isinstance(v, (int, float)) and 0 < float(v) <= 100:
            por_anio.setdefault(anio, []).append(float(v))
    return {a: sum(vs) / len(vs) for a, vs in por_anio.items()
            if len(vs) == _TRIMESTRES_POR_ANIO}


def parse_gender_ratio(content: bytes, label: str) -> List[Tuple[int, float]]:
    """`[(año, tasa femenina / tasa masculina)]` para una fila del libro.

    **Es una razón, no una diferencia**: así la define la END —«tasa femenina/tasa
    masculina»— y por eso el indicador de ocupación mejora subiendo hacia 1 y el de
    desocupación mejora bajando hacia 1.

    **Se divide el promedio anual de cada sexo, no se promedian las razones
    trimestrales.** Las dos cuentas dan distinto y ninguna es «la» correcta: se elige
    ésta porque es la que reproduce cualquiera que tome las tasas anuales publicadas de
    cada sexo, que es la forma en que un tercero verificaría la cifra.

    **Por qué una razón sobrevive al cambio de encuesta y un nivel no.** La ENFT y la
    ENCFT dan niveles de desocupación que diferen casi al doble, así que comparar el
    nivel de 2024 contra una meta fijada sobre ENFT no significa nada. En una razón entre
    dos poblaciones medidas por la MISMA encuesta, el cambio de metodología se cancela en
    buena parte del numerador y del denominador. Es lo que vuelve evaluables a las
    brechas de género cuando el nivel no lo es.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        faltan = [h for h in (SHEET_FEMALE, SHEET_MALE) if h not in wb.sheetnames]
        if faltan:
            raise BcrdLaborUnavailable(
                f"el libro no trae {faltan} (hojas: {wb.sheetnames})")
        fem = _anual_por_sexo([list(r) for r in wb[SHEET_FEMALE].iter_rows(values_only=True)],
                              label)
        mas = _anual_por_sexo([list(r) for r in wb[SHEET_MALE].iter_rows(values_only=True)],
                              label)
    finally:
        wb.close()

    out: List[Tuple[int, float]] = []
    for anio in sorted(set(fem) & set(mas)):
        divisor = mas[anio]
        if divisor <= 0:          # una tasa masculina nula haría explotar la razón
            continue
        out.append((anio, round(fem[anio] / divisor, 3)))
    if not out:
        raise BcrdLaborUnavailable(
            f"no hay ningún año con los cuatro trimestres en ambos sexos para '{label}'")
    return out


#: Hoja del corte por región y fila de la desocupación AMPLIADA. El libro publica cuatro
#: medidas de subutilización (SU1 a SU4) y la ley nombra la ampliada, que es SU2: desocupados
#: abiertos MÁS la fuerza de trabajo potencial. Atar SU1 —la abierta— mediría otra población
#: y correría unos siete puntos por debajo.
SHEET_REGIONS = "Regiones"
BROAD_UNEMPLOYMENT_LABEL = "SU2"

#: Las cuatro columnas de dominio, sin `Total País`. Meter el total dentro de lo que se
#: ordena produciría una «brecha» contra un promedio, que es otra definición.
_COLS_REGION = (2, 6)


def parse_regional_gap(content: bytes) -> List[Tuple[int, float]]:
    """`[(año, brecha entre la región peor y la mejor)]` en puntos porcentuales.

    **Por qué máximo menos mínimo y no otra cosa.** La ley dice «brecha regional» sin
    definirla, y las tres lecturas posibles dan resultados muy distintos contra la línea base
    de 6,4 que fija para 2010. Contrastadas contra la encuesta vigente en ese año:

        máximo − mínimo          5,90   Δ  7,8%
        máximo − promedio        2,43   Δ 62,1%
        razón máximo/mínimo      1,54   Δ 76,0%

    La primera es la única que se acerca, y la distancia entre las tres es lo que vuelve
    concluyente la elección: no se eligió por conveniencia sino porque las otras dos quedan
    fuera por un factor.

    **La serie que se devuelve es de la ENCFT y arranca en 2015**, así que NO alcanza el año
    de la línea base. El contraste de arriba se hizo contra la ENFT, que es otra encuesta:
    sirve para identificar la definición, no para verificar la serie. Esa salvedad viaja al
    informe.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if SHEET_REGIONS not in wb.sheetnames:
            raise BcrdLaborUnavailable(
                f"el libro no trae la hoja '{SHEET_REGIONS}' (hojas: {wb.sheetnames})")
        filas = [list(r) for r in wb[SHEET_REGIONS].iter_rows(max_col=6, values_only=True)]
    finally:
        wb.close()

    out: List[Tuple[int, float]] = []
    anio: Optional[int] = None
    for fila in filas:
        cabeza = fila[0]
        if isinstance(cabeza, (int, float)) and 1990 < float(cabeza) < 2100:
            anio = int(cabeza)
            continue
        if not isinstance(cabeza, str):
            continue
        etiqueta = cabeza.strip()
        if re.fullmatch(r"(19|20)\d{2}", etiqueta):
            anio = int(etiqueta)
            continue
        if anio is None or not etiqueta.upper().startswith(BROAD_UNEMPLOYMENT_LABEL):
            continue
        tasas = [v for v in fila[_COLS_REGION[0]:_COLS_REGION[1]]
                 if isinstance(v, (int, float))]
        # Una región ausente cambiaría el máximo o el mínimo sin avisar, y la brecha
        # saldría más chica de lo que es. Se exige el panel completo.
        if len(tasas) == _COLS_REGION[1] - _COLS_REGION[0]:
            out.append((anio, round(max(tasas) - min(tasas), 2)))
        anio = None
    if not out:
        raise BcrdLaborUnavailable(
            f"no se encontró ninguna fila '{BROAD_UNEMPLOYMENT_LABEL}' con las cuatro "
            f"regiones en la hoja '{SHEET_REGIONS}'")
    return sorted(out)


def fetch_bcrd_labor_market() -> dict:  # pragma: no cover - network I/O
    """Live: baja el libro UNA vez y devuelve las tres series que la END necesita.

    Una sola descarga para los tres: el libro pesa ~400 KB y bajarlo tres veces sería
    pagar tres veces por el mismo archivo.
    """
    import httpx

    try:
        r = httpx.get(INDICATORS_URL, timeout=120, follow_redirects=True)
        r.raise_for_status()
    except httpx.HTTPError as e:
        raise BcrdLaborUnavailable(
            f"no se pudo descargar {INDICATORS_FILE} del CDN del BCRD "
            f"({type(e).__name__}: {e})")
    return {
        "informality_rate": parse_informality(r.content),
        "unemployment_rate": parse_informality(r.content, label=UNEMPLOYMENT_LABEL),
        "employment_gender_ratio": parse_gender_ratio(r.content, EMPLOYMENT_RATE_LABEL),
        "unemployment_gender_ratio": parse_gender_ratio(r.content, UNEMPLOYMENT_LABEL),
        "regional_unemployment_gap": parse_regional_gap(r.content),
        # TRIMESTRALES, de la hoja «Indicadores» del MISMO libro. No es una segunda
        # descarga ni una derivación de la anual: son series propias del emisor con su
        # cadencia real, y hacen falta porque el crédito se mide por trimestre.
        "unemployment_rate_trimestral": parse_trimestral(r.content, UNEMPLOYMENT_LABEL),
        "informality_rate_trimestral": parse_trimestral(r.content, "Ocupación Informal"),
        "employment_rate_trimestral": parse_trimestral(r.content, "Tasa de Ocupación"),
        # LAS CUATRO MEDIDAS DE SUBUTILIZACIÓN, no solo la angosta. El BCRD publica SU1 a
        # SU4 en la misma fila del mismo trimestre y nosotros citábamos SU1: 4,95% contra
        # 10,55% de SU4 al primer trimestre de 2026 — menos de la mitad de la holgura real.
        #
        # Para capacidad de pago la diferencia es exactamente la que importa: SU1 solo ve a
        # quien busca y no encuentra, mientras SU2 suma al subocupado por horas —que tiene
        # empleo e ingreso INSUFICIENTE— y SU3 a la fuerza de trabajo potencial. Un hogar
        # endeudado con menos horas de las que quiere no aparece en SU1 y sí en su mora.
        #
        # SU1 no se retira: es la que la END ata como «desocupación abierta» y la que se
        # compara internacionalmente. Conviven, cada una con su nombre.
        "underutilization_su2_trimestral": parse_trimestral(
            r.content, "SU2: Desocupación y Subocupación"),
        "underutilization_su3_trimestral": parse_trimestral(
            r.content, "SU3: Desocupación y Fuerza de Trabajo Potencial"),
        "underutilization_su4_trimestral": parse_trimestral(
            r.content, "SU4: Desocupación + Subocupación + Fuerza de Trabajo Potencial"),
        # Subocupación por horas: insuficiencia de INGRESO en gente que sí trabaja, que es
        # la población de crédito de consumo que ningún indicador de desempleo alcanza.
        "underemployment_rate_trimestral": parse_trimestral(
            r.content, "Tasa de subocupación por horas"),
        # Inactividad: quién directamente no está en la fuerza de trabajo. Acota el
        # denominador de todo lo anterior.
        "inactivity_rate_trimestral": parse_trimestral(r.content, "Tasa de Inactividad"),
        # Ocupación FORMAL, el complemento explícito de la informal. Se sirve junto porque
        # «formal» e «informal» no siempre suman cien —hay categorías fuera de ambas— y
        # derivar una de la otra fabricaría el resto.
        "formality_rate_trimestral": parse_trimestral(r.content, "Ocupación Formal"),
        # LA PRECISIÓN de cada estimación. La ENCFT es una ENCUESTA: cada cifra que
        # publicamos tiene error de muestreo, y el BCRD publica el intervalo de confianza y
        # el coeficiente de variación de todas — en otra hoja del MISMO libro. Servíamos las
        # estimaciones desnudas, y una diferencia menor que los intervalos no es una
        # diferencia. Se emiten aparte y no mezcladas con el valor: son otra magnitud.
        "precision_trimestral": {
            clave: parse_precision(r.content, etiqueta)
            for clave, etiqueta in _PRECISION_DE_LAS_SERVIDAS.items()
        },
    }
