"""SISDOM — pobreza monetaria e indigencia por ZONA de residencia (cuadro 03 3 003a).

Es el corte que los indicadores 2.3 y 2.6 de la Ley 1-12 piden y que el panel por región no
tiene. La ley mide pobreza RURAL; el conjunto abierto que ya se ingiere está desagregado por
las diez regiones de desarrollo y no por zona, así que ninguna combinación de sus filas
produce la cifra rural.

**Dos vocabularios para lo mismo.** El emisor llama «indigencia monetaria» a lo que la ley
llama pobreza extrema, y «pobreza monetaria» a la línea general. La correspondencia se
comprobó contra el oráculo: la indigencia rural de 2010 da 16,76 contra los 16,9 que fija el
indicador 2.3.

**El cuadro tiene un QUIEBRE de metodología y lo declara publicando 2016 dos veces**, una por
cada encuesta, la segunda marcada con asterisco. Este módulo no lo disimula: sirve la serie
con la metodología vigente y expone el año de solape para que quien ate la serie pueda medir
el tamaño del salto en vez de suponerlo. Encadenar los dos tramos sin declararlo sería el
defecto que este repositorio ya pagó con la informalidad laboral.

Nada se lee por posición. La zona y la línea salen de sus ETIQUETAS en la cabecera, porque el
cuadro gana columnas entre ediciones.
"""
from __future__ import annotations

import logging
import re
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger("sdq.data.sisdom_pobreza")

HOJA = "03 3 003a"
SOURCE = "SISDOM · MEPyD"

#: Fila de la cabecera donde vive cada rótulo. Se buscan por CONTENIDO dentro de una ventana
#: de filas, no por número fijo: el cuadro mueve su encabezado entre ediciones.
_FILAS_CABECERA = range(6, 12)

#: Cómo nombra el emisor cada línea de pobreza, y cómo la nombra la ley.
LINEAS = {"indigencia": "extrema", "pobreza": "general"}

#: El asterisco marca la metodología nueva. Se conserva para poder distinguir los dos tramos.
_RX_ANIO = re.compile(r"^(\d{4})(\*?)$")


class SISDOMZonaError(RuntimeError):
    """No se pudo leer el cuadro. NUNCA se degrada a «no hay dato»."""


def _norm(v: object) -> str:
    return " ".join(str(v or "").split()).lower()


#: Las zonas que rotula el cuadro. Se declaran para poder EXIGIR que la pedida esté entre
#: ellas: pedir una zona que el emisor no publica tiene que fallar, no devolver la de al lado.
ZONAS = ("nacional", "urbana", "rural")


def _columnas_de_zona(ws, zona: str = "rural") -> Dict[str, int]:
    """`{linea: columna}` de una zona, leído de las etiquetas de la cabecera.

    La fila de zonas rotula solo la primera columna de cada bloque, así que la pertenencia se
    arrastra hacia la derecha hasta la zona siguiente. Leerlo por posición fija devolvería la
    columna urbana el día que el emisor agregue un desglose, y las dos series se parecen lo
    suficiente como para que nadie lo note.
    """
    if zona not in ZONAS:
        raise SISDOMZonaError(f"zona '{zona}' desconocida; el cuadro publica {ZONAS}")
    fila_zona = fila_linea = None
    for f in _FILAS_CABECERA:
        vals = [_norm(ws.cell(f, c).value) for c in range(1, ws.max_column + 1)]
        if any(v == zona for v in vals):
            fila_zona = f
        if any(v.startswith("indigencia") for v in vals):
            fila_linea = f
    if fila_zona is None or fila_linea is None:
        raise SISDOMZonaError(
            "el cuadro no trae las cabeceras de zona y de línea donde se esperaban; el "
            "emisor cambió el layout y leerlo por posición serviría otra columna")

    zona_actual = ""
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        z = _norm(ws.cell(fila_zona, c).value)
        if z in ZONAS or z.startswith("regiones"):
            zona_actual = z
        if zona_actual != zona:
            continue
        etiqueta = _norm(ws.cell(fila_linea, c).value)
        for clave in LINEAS:
            if etiqueta.startswith(clave):
                out.setdefault(clave, c)
    if set(out) != set(LINEAS):
        raise SISDOMZonaError(
            f"no se ubicaron las dos líneas de la zona {zona}; se halló {sorted(out)}")
    return out


def parse_zona(contenido: bytes,
               zona: str = "rural") -> Tuple[Dict[str, List[Tuple[str, float]]], Dict]:
    """`({linea: [(año, valor)]}, quiebre)` de una zona del cuadro, solo filas anuales.

    Las filas de marzo y septiembre son cortes de la encuesta, no el año. Publicar una de
    ellas como el año sería el mismo error que tomar un anexo mensual del sector eléctrico por
    el ejercicio completo.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(__import__("io").BytesIO(contenido), data_only=True)
    except Exception as e:                                   # archivo corrupto
        raise SISDOMZonaError(f"no se pudo abrir el libro ({type(e).__name__}: {e})")
    hojas = [n for n in wb.sheetnames if n.strip() == HOJA]
    if not hojas:
        raise SISDOMZonaError(f"el libro no trae la hoja «{HOJA}»")
    ws = wb[hojas[0]]
    cols = _columnas_de_zona(ws, zona)

    crudo: Dict[str, Dict[str, float]] = {k: {} for k in LINEAS}
    anio = ""
    for f in range(1, ws.max_row + 1):
        etiqueta_anio = str(ws.cell(f, 1).value or "").strip()
        if _RX_ANIO.match(etiqueta_anio):
            anio = etiqueta_anio
        if _norm(ws.cell(f, 2).value) != "año" or not anio:
            continue
        for linea, c in cols.items():
            v = ws.cell(f, c).value
            if isinstance(v, (int, float)):
                crudo[linea][anio] = float(v)

    # ── El quiebre ──────────────────────────────────────────────────────────────────────
    # El emisor publica el año de solape dos veces. Se sirve la metodología VIGENTE y se
    # expone el salto medido, que es lo que permite a quien ate la serie decidir con una
    # cifra en vez de con una intuición.
    solape = sorted({a[:4] for a in crudo["indigencia"] if a.endswith("*")}
                    & {a for a in crudo["indigencia"] if not a.endswith("*")})
    salto: Dict[str, float] = {}
    for linea in LINEAS:
        if not solape:
            continue
        viejo, nuevo = crudo[linea].get(solape[0]), crudo[linea].get(solape[0] + "*")
        if viejo and nuevo:
            salto[linea] = round(abs(nuevo - viejo) / viejo * 100, 1)
    quiebre: Dict[str, object] = {"anio_de_solape": solape[0] if solape else None,
                                  "salto_pct": salto}

    series: Dict[str, List[Tuple[str, float]]] = {}
    for linea, vals in crudo.items():
        limpio: Dict[str, float] = {}
        for a, v in vals.items():
            base = a.rstrip("*")
            # Con las dos metodologías presentes para el mismo año manda la NUEVA: es la que
            # el emisor sostiene hoy, igual que con las ediciones del anexo eléctrico.
            if a.endswith("*") or base not in limpio:
                limpio[base] = v
        series[linea] = sorted(limpio.items())
    return series, quiebre


def parse_zona_rural(contenido: bytes) -> Tuple[Dict[str, List[Tuple[str, float]]], Dict]:
    """Compatibilidad: la zona rural es la que ya se ingería cuando el cuadro tenía un solo
    lector. Se conserva el nombre para no tocar a quien lo llama."""
    return parse_zona(contenido, "rural")


def fetch_zona(zona: str = "rural"
               ) -> Tuple[Dict[str, List[Tuple[str, float]]], Dict]:  # pragma: no cover
    """Live: baja el libro del emisor y devuelve la zona pedida."""
    from shared.data.sisdom_common import fetch_book

    return parse_zona(fetch_book("pobreza"), zona)


def fetch_zona_rural() -> Tuple[Dict[str, List[Tuple[str, float]]], Dict]:  # pragma: no cover
    from shared.data.sisdom_common import fetch_book

    return parse_zona_rural(fetch_book("pobreza"))
