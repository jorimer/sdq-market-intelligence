"""Superintendencia de Pensiones (SIPEN) — connector for Dominican pension data.

SIPEN data has two natures, and this connector exposes both behind the platform's
``Record`` contract:

  * **System / national** series (the pension system as a whole): rentabilidad,
    comisiones, afiliados, cotizantes, patrimonio del fondo. These behave like
    :class:`MacroSeries` — one value per period.
  * **Entity** series (per AFP — the ~7 fund administrators): rentabilidad,
    comisiones, patrimonio gestionado, afiliados. Carried as ``Record`` with the
    AFP slug in ``dimension`` (the ONE/region pattern), so one spine serves both.

Four real publication channels exist (the integration plan ``tasks/PLAN_PENSIONES_SIPEN.md``):
  A) CKAN ``datos.gob.do`` (afiliados/cotizantes, CSV/JSON)            — live: Fase 1
  B) Estadística Previsional XLSX (valor cuota, rentabilidad, cartera) — live: Fase 1
  C) Boletines Trimestrales (PDF) → digest IA                          — live: Fase 1
  D) Estados financieros AFP + accionistas/capital pagado             — live: Fase 2

The portal ``sipen.gob.do`` blocks non-browser clients (403/473), so the platform
doctrine of fixture-first is doubly right here. F0 ships a small **real, cited**
sample (``sipen.json`` + ``sipen_entities.json``); the live wiring per channel
lands in later phases. Missing values stay ``None`` — never interpolated.
"""
import logging
from datetime import date
from typing import Dict, List, Optional, Tuple

from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage

logger = logging.getLogger("sdq.data.sipen")

# ── AFP catalog ────────────────────────────────────────────────────────────
# The fund administrators currently in the Dominican system (ADAFP members +
# Atlántico). Slugs are stable keys used as ``Record.dimension`` and as the
# ``PensionEntity.slug``. Display names match SIPEN/ADAFP usage.
AFPS: List[Tuple[str, str]] = [
    ("afp_popular", "AFP Popular"),
    ("afp_crecer", "AFP Crecer"),
    ("afp_reservas", "AFP Reservas"),
    ("afp_siembra", "AFP Siembra"),
    ("afp_romana", "AFP Romana"),
    ("afp_jmmb_bdi", "AFP JMMB BDI"),
    ("afp_atlantico", "AFP Atlántico"),
]


def afp_catalog() -> List[Tuple[str, str]]:
    """``[(slug, display_name)]`` for seeding the AFP peer set."""
    return list(AFPS)


class SIPENClient(FixtureBackedClient):
    """SIPEN pension statistics. ``fixture`` reads the cited sample; ``live`` per channel.

    ``fetch`` returns the **system/national** series (standard fixture shape).
    Per-AFP series come from :meth:`fetch_entities` (``dimension`` = AFP slug).
    """

    source = "SIPEN"
    license = "datos públicos SIPEN — sistema dominicano de pensiones (uso con cita)"
    license_ok = True
    fixture_file = "sipen.json"
    entities_fixture_file = "sipen_entities.json"
    live_phase = "Fase 1 (CKAN datos.gob.do + XLSX Estadística Previsional)"

    # ── Entity (per-AFP) series ───────────────────────────────────────────
    def fetch_entities(
        self, series: Optional[str] = None, period: Optional[str] = None
    ) -> List[Record]:
        """Per-AFP observations as ``Record``\\ s (``dimension`` = AFP slug).

        Fixture shape::

            {"<afp_slug>": {"name": "AFP …",
                            "series": {"<metric>": {"unit": "%",
                                                    "observations": {"2025-02": 9.59}}}}}
        """
        self.check_license()
        if self.mode == "live":
            raise NotImplementedError(
                f"Conector {self.source} entidades en modo 'live' pendiente ({self.live_phase})"
            )
        fixture = self._load_fixture(self.entities_fixture_file)
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today())
        out: List[Record] = []
        for slug, entity in fixture.items():
            for s_name, s in (entity.get("series") or {}).items():
                if series and s_name != series:
                    continue
                unit = s.get("unit")
                for p, v in (s.get("observations") or {}).items():
                    if period and p != period:
                        continue
                    out.append(Record(
                        series=s_name, period=p,
                        value=None if v is None else float(v),
                        lineage=lineage, unit=unit, dimension=slug,
                    ))
        return out

    def entity_names(self) -> Dict[str, str]:
        """``{slug: display_name}`` from the catalog (+ any extra in the fixture)."""
        names = {slug: name for slug, name in AFPS}
        fixture = self._load_fixture(self.entities_fixture_file)
        for slug, entity in fixture.items():
            if entity.get("name"):
                names.setdefault(slug, entity["name"])
        return names


sipen_client = SIPENClient()


# ── Live system series (CKAN datos.gob.do) ─────────────────────────────────────
# SIPEN publishes three machine-readable national series on the government open-data
# portal (CKAN), monthly 2003-…: afiliados, cotizantes, salario mínimo cotizable. The
# resource is an XLSX with columns ``Mes | Año | <value>``. We resolve the current XLSX
# URL per dataset via the CKAN API (resilient to URL changes) and parse it to Records.
# This moves the national pulse off the fixture sample onto the full live series.
_CKAN_PACKAGE = "https://datos.gob.do/api/3/action/package_show"
_CKAN_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0 Safari/537.36",
}
# CKAN dataset id → (series_code, unit). The value column is the 3rd column.
_CKAN_DATASETS: List[Tuple[str, str, str]] = [
    ("estadisticas-sobre-afiliados", "sipen.afiliados.total", "personas"),
    ("distribucion-de-cotizantes", "sipen.cotizantes.total", "personas"),
    ("salario-minimo-cotizable", "sipen.salario_minimo_cotizable", "RD$"),
]
_MONTHS: Dict[str, str] = {
    m: f"{i:02d}" for i, m in enumerate(
        ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto",
         "septiembre", "octubre", "noviembre", "diciembre"], 1)
}


def parse_mes_ano_xlsx(content: bytes) -> List[Tuple[str, float]]:
    """Parse a SIPEN CKAN ``Mes | Año | valor`` workbook → ``[(period 'YYYY-MM', value)]``.

    Spanish month names → ``MM``; rows without a recognized month / numeric year / numeric
    value are skipped (never guessed). Pure (no network) — the unit under test."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        out: List[Tuple[str, float]] = []
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue
            mes, ano, val = row[0], row[1], row[2]
            month = _MONTHS.get(str(mes).strip().lower()) if mes is not None else None
            if month and str(ano).strip().isdigit() and isinstance(val, (int, float)):
                out.append((f"{int(ano)}-{month}", float(val)))
        return out
    finally:
        wb.close()


# ── Live rentabilidad (XLSX "Estadística Previsional") ──────────────────────────
# SIPEN publishes the system & per-fund return series as an XLSX in its "Estadística
# Previsional" section. Discovery is label-based (robust to the timestamped filenames):
# each download card carries aria-label="Descargar documento <Title>", so we pick the
# "Rentabilidad" .xlsx by its label, not its position. The workbook has two sheets:
#   * "Por SISTEMA": Mes | Promedio CCI | Promedio Sistema (trailing-12m nominal, fraction)
#   * "Por FONDO":   Mes | <one column per AFP> (same metric; discontinued AFPs are 0)
# Values are fractions (0.0775 = 7.75%); we carry them as percent (×100, 2 dp) to match
# the cited fixture sample exactly (verified: XLSX 2025-02/2025-04 == fixture figures).
_PREVISIONAL_PAGE = (
    "https://sipen.gob.do/estadisticas/estadistica-previsional/estadisticas-previsionales"
)
# Site header token (lowercased, footnote digits stripped) → AFP slug. Only the 7 live
# administrators map; discontinued funds (Camino/Caribálico/León/Porvenir) and the
# non-AFP fondos (Solidaridad Social, Reparto, INABIMA) have no slug → skipped.
_RENT_AFP_BY_HEADER: Dict[str, str] = {
    "atlántico": "afp_atlantico",
    "crecer": "afp_crecer",
    "jmmb-bdi": "afp_jmmb_bdi",
    "popular": "afp_popular",
    "reservas": "afp_reservas",
    "romana": "afp_romana",
    "siembra": "afp_siembra",
}


def previsional_xlsx_links(html: str) -> Dict[str, str]:
    """``{label_lower: url}`` of the labelled .xlsx downloads on the Estadística
    Previsional page (e.g. ``{"rentabilidad": ".../…_1781293450.xlsx", …}``).

    Keys on the card's ``aria-label="Descargar documento <Title>"`` — stable against the
    timestamped, otherwise-undistinguishable filenames. Pure (no network)."""
    import re

    pat = re.compile(
        r'href="(https://sipen\.gob\.do/descarga/[^"]+\.xlsx)"[^>]*'
        r'aria-label="Descargar documento ([^"]+)"',
        re.IGNORECASE,
    )
    return {label.strip().lower(): url for url, label in pat.findall(html)}


def _rent_pct(v) -> Optional[float]:
    """Fraction → percent (2 dp); non-numeric / 0 (fund not operating) → None."""
    return round(float(v) * 100, 2) if isinstance(v, (int, float)) and v else None


def _norm_afp_header(cell) -> str:
    """Header cell → lookup key: lowercase, trimmed, trailing footnote digits removed
    (``'Camino1'`` → ``'camino'``, ``'JMMB-BDI'`` → ``'jmmb-bdi'``)."""
    import re

    return re.sub(r"\d+$", "", str(cell).strip().lower()) if cell is not None else ""


def parse_rentabilidad_xlsx(
    content: bytes,
) -> Tuple[List[Tuple[str, Optional[float], Optional[float]]], Dict[str, List[Tuple[str, float]]]]:
    """Parse the Rentabilidad workbook → ``(system, per_afp)``.

    ``system`` = ``[(period 'YYYY-MM', cci_pct|None, sistema_pct|None)]`` from "Por SISTEMA".
    ``per_afp`` = ``{afp_slug: [(period, pct)]}`` from "Por FONDO" (only the 7 live AFPs;
    0/blank cells skipped — never an invented 0% return). Pure (no network); the unit under
    test. A data row is any row whose 2nd column is a date."""
    import datetime
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        names = set(wb.sheetnames)

        system: List[Tuple[str, Optional[float], Optional[float]]] = []
        sys_sheet = "Por SISTEMA" if "Por SISTEMA" in names else wb.sheetnames[0]
        for row in wb[sys_sheet].iter_rows(values_only=True):
            d = row[1] if len(row) > 1 else None
            if isinstance(d, datetime.datetime):
                cci = _rent_pct(row[2]) if len(row) > 2 else None
                sdp = _rent_pct(row[3]) if len(row) > 3 else None
                if cci is not None or sdp is not None:
                    system.append((d.strftime("%Y-%m"), cci, sdp))

        per_afp: Dict[str, List[Tuple[str, float]]] = {}
        if "Por FONDO" in names:
            rows = list(wb["Por FONDO"].iter_rows(values_only=True))
            colmap: Dict[int, str] = {}
            for row in rows:  # locate the header row (≥3 known AFP names)
                hits = {
                    j: _RENT_AFP_BY_HEADER[_norm_afp_header(c)]
                    for j, c in enumerate(row)
                    if _norm_afp_header(c) in _RENT_AFP_BY_HEADER
                }
                if len(hits) >= 3:
                    colmap = hits
                    break
            if colmap:
                for row in rows:
                    d = row[1] if len(row) > 1 else None
                    if not isinstance(d, datetime.datetime):
                        continue
                    period = d.strftime("%Y-%m")
                    for j, slug in colmap.items():
                        v = _rent_pct(row[j]) if len(row) > j else None
                        if v is not None:
                            per_afp.setdefault(slug, []).append((period, v))
        return system, per_afp
    finally:
        wb.close()


def fetch_sipen_rentabilidad(  # pragma: no cover - network I/O
    period: Optional[str] = None,
) -> Tuple[List[Record], List[Record]]:
    """LIVE: the SIPEN return series (system + per-AFP) from the Estadística Previsional
    XLSX → ``(system_records, entity_records)``.

    System Records carry the same codes as the fixture (``sipen.rentabilidad.cci_nominal_anual``
    / ``…sdp_nominal_anual``); entity Records carry ``rentabilidad_nominal_anual`` with the AFP
    slug in ``dimension``. Best-effort: discovery/parse failure returns ``([], [])`` and the
    sync keeps the fixture floor. Runs in the sync; offline/tests stay on the fixture."""
    import httpx

    lineage = Lineage(
        source=SIPENClient.source, license=SIPENClient.license, fetched_at=date.today(),
        url=_PREVISIONAL_PAGE,
        note="Rentabilidad nominal (sistema + por AFP) — XLSX Estadística Previsional",
    )
    system_out: List[Record] = []
    entity_out: List[Record] = []
    with httpx.Client(timeout=60, headers=_CKAN_HEADERS, follow_redirects=True) as http:
        url = previsional_xlsx_links(http.get(_PREVISIONAL_PAGE).text).get("rentabilidad")
        if not url:
            logger.warning("[SIPEN] no se halló el XLSX de Rentabilidad en %s", _PREVISIONAL_PAGE)
            return [], []
        content = http.get(url).content
    system, per_afp = parse_rentabilidad_xlsx(content)
    for per, cci, sdp in system:
        if period and per != period:
            continue
        if cci is not None:
            system_out.append(Record(series="sipen.rentabilidad.cci_nominal_anual",
                                      period=per, value=cci, lineage=lineage, unit="%"))
        if sdp is not None:
            system_out.append(Record(series="sipen.rentabilidad.sdp_nominal_anual",
                                      period=per, value=sdp, lineage=lineage, unit="%"))
    for slug, obs in per_afp.items():
        for per, val in obs:
            if period and per != period:
                continue
            entity_out.append(Record(series="rentabilidad_nominal_anual", period=per,
                                     value=val, lineage=lineage, unit="%", dimension=slug))
    return system_out, entity_out


def fetch_sipen_ckan(period: Optional[str] = None) -> List[Record]:  # pragma: no cover - network I/O
    """LIVE: the three SIPEN national series from CKAN datos.gob.do → ``Record``\\ s.

    Best-effort per dataset (a failed one never aborts the others); returns whatever it
    got. Runs in the sync; offline/tests stay on the fixture."""
    import urllib.parse

    import httpx

    lineage = Lineage(
        source=SIPENClient.source, license=SIPENClient.license, fetched_at=date.today(),
        url="https://datos.gob.do/organization/superintendencia-de-pensiones-sipen",
        note="Series nacionales CKAN (afiliados/cotizantes/salario mínimo cotizable)",
    )
    out: List[Record] = []
    with httpx.Client(timeout=40, headers=_CKAN_HEADERS, follow_redirects=True) as http:
        for dataset_id, code, unit in _CKAN_DATASETS:
            try:
                res = http.get(_CKAN_PACKAGE, params={"id": dataset_id}).json()["result"]
                xr = next((r for r in res.get("resources", [])
                           if (r.get("format") or "").upper() == "XLSX"), None)
                if not xr or not xr.get("url"):
                    continue
                content = http.get(urllib.parse.quote(xr["url"], safe=":/")).content
                for per, val in parse_mes_ano_xlsx(content):
                    if period and per != period:
                        continue
                    out.append(Record(series=code, period=per, value=val,
                                      lineage=lineage, unit=unit))
            except Exception as e:  # noqa: BLE001 — best-effort per dataset
                logger.warning("[SIPEN] CKAN %s falló: %s", dataset_id, e)
    return out
