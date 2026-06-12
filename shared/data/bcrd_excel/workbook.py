"""Uniform workbook reader — one grid abstraction over ``.xlsx`` and ``.xls``.

``openpyxl`` reads only ``.xlsx``; the BCRD corpus is roughly half legacy ``.xls``
(needs ``xlrd``). This module normalizes both into a plain ``Grid`` of Python
values so inference/extraction never branch on the file format.

A ``structure_hash`` keys the spec cache: it fingerprints the *layout* (sheet
names, dimensions, and the text of non-numeric cells in the header region) but
NOT the numeric values, so a routine data refresh keeps the cached spec while a
genuine layout change busts it and triggers re-inference.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any, List

# rows/cols of the header region that the structure hash fingerprints
_HASH_ROWS = 40
_HASH_COLS = 15


@dataclass
class Grid:
    """One sheet as a dense row-major matrix of cell values (``None`` = empty)."""

    name: str
    rows: List[List[Any]]

    @property
    def nrows(self) -> int:
        return len(self.rows)

    @property
    def ncols(self) -> int:
        return max((len(r) for r in self.rows), default=0)

    def cell(self, r: int, c: int) -> Any:
        """Value at 0-based (row, col); ``None`` if out of range."""
        if 0 <= r < len(self.rows):
            row = self.rows[r]
            if 0 <= c < len(row):
                return row[c]
        return None


@dataclass
class Workbook:
    path: str
    grids: List[Grid]

    @property
    def sheet_names(self) -> List[str]:
        return [g.name for g in self.grids]

    def grid(self, name: str | None = None) -> Grid:
        """Sheet by name (first sheet if *name* is None/missing)."""
        if name:
            for g in self.grids:
                if g.name == name:
                    return g
        return self.grids[0]

    def structure_hash(self) -> str:
        """Layout fingerprint (sheet names + dims + header text, not numbers)."""
        h = hashlib.sha256()
        for g in self.grids:
            h.update(g.name.encode("utf-8", "ignore"))
            h.update(f"|{g.nrows}x{g.ncols}|".encode())
            for r in range(min(_HASH_ROWS, g.nrows)):
                for c in range(min(_HASH_COLS, g.ncols)):
                    v = g.cell(r, c)
                    if v is not None and not isinstance(v, (int, float)):
                        h.update(f"{r},{c}:{v}".encode("utf-8", "ignore"))
        return h.hexdigest()[:16]


def _load_xlsx(path: Path) -> List[Grid]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        grids = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            grids.append(Grid(name=name, rows=rows))
        return grids
    finally:
        wb.close()


def _load_xls(path: Path) -> List[Grid]:
    import xlrd

    wb = xlrd.open_workbook(path)
    grids = []
    for sh in wb.sheets():
        rows = []
        for i in range(sh.nrows):
            row = []
            for j in range(sh.ncols):
                v = sh.cell_value(i, j)
                row.append(None if v == "" else v)
            rows.append(row)
        grids.append(Grid(name=sh.name, rows=rows))
    return grids


def load_workbook(path: str | Path) -> Workbook:
    """Read an ``.xlsx`` or ``.xls`` file into a uniform :class:`Workbook`."""
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".xlsx":
        grids = _load_xlsx(p)
    elif suffix == ".xls":
        grids = _load_xls(p)
    else:
        raise ValueError(f"Formato no soportado: {p.suffix} ({p.name})")
    if not grids:
        raise ValueError(f"Workbook sin hojas: {p.name}")
    return Workbook(path=str(p), grids=grids)
