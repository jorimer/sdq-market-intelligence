"""Dirección General de Aduanas (DGA) connector — real customs trade data.

Public open data: the DGA publishes quarterly foreign-trade statistics by HS
chapter (exports + imports, FOB) on its institutional pages, as downloadable
"Data Cruda" Excel files (``/media/{hash}/...-dc.xlsx``). We scrape the landing
pages for those media links (the hash is opaque, like the ONE PDFs), download the
raw Excel and parse the chapter table → trade flows for ``trade_intel``.

Country-of-partner detail lives only inside a Power BI embed (no scriptable
export), so this connector is chapter-level — sufficient for the resilience
score (HHI/diversification by product + import dependency are share-based).

Missing values stay ``None``; nothing is interpolated or fabricated.
"""
import logging
import re
from html import unescape
from typing import Dict, List, Optional, Tuple

from shared.data.base_client import FixtureBackedClient

logger = logging.getLogger("sdq.data.dga")

BASE = "https://www.aduanas.gob.do"
EXPORTS_LANDING = f"{BASE}/estadisticas-institucionales/exportaciones/"
IMPORTS_LANDING = f"{BASE}/estadisticas-institucionales/importaciones/"
_HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP DGA connector)"}

# First month of a quarter → quarter number (the files are 3-month ranges).
_MONTH_Q = {"enero": 1, "abril": 2, "julio": 3, "octubre": 4}
_MEDIA_RE = re.compile(r"/media/[a-z0-9]+/[^\"'>]+?\.xlsx", re.IGNORECASE)


def _filename(url: str) -> str:
    return unescape(url).rsplit("/", 1)[-1].lower()


def _is_data_cruda(name: str) -> bool:
    """A raw-data (Data Cruda) chapter file — not the formatted summary or a PDF."""
    if not name.endswith(".xlsx") or "excel.xlsx" in name:
        return False
    return "cruda" in name or re.search(r"(?:[-_]|\d{4})dc(?:[-_]v\d+)?\.xlsx$|[-_]dc[-_]", name) is not None


def _parse_period(name: str) -> Optional[str]:
    """``exportaciones..._octubre_diciembre_2025dc.xlsx`` → ``"2025-Q4"`` (or None)."""
    m = re.search(r"(enero|abril|julio|octubre)", name)
    y = re.search(r"(20\d{2})", name)
    if not m or not y:
        return None
    return f"{y.group(1)}-Q{_MONTH_Q[m.group(1)]}"


def discover_dc_files(html: str, direction: str) -> Dict[str, str]:
    """``{period: url}`` for the Data-Cruda chapter files in *html*.

    Keeps one file per period; prefers a ``-v2`` revision when several exist.
    *direction* (``export``/``import``) filters by filename prefix.
    """
    want = "exportaciones" if direction == "export" else "importaciones"
    best: Dict[str, Tuple[int, str]] = {}
    for raw in _MEDIA_RE.findall(html):
        url = unescape(raw)
        name = _filename(url)
        if want not in name or not _is_data_cruda(name):
            continue
        period = _parse_period(name)
        if not period:
            continue
        rank = 2 if "v2" in name else 1  # prefer a revised file
        cur = best.get(period)
        if cur is None or rank > cur[0]:
            best[period] = (rank, BASE + url if url.startswith("/") else url)
    return {p: u for p, (_, u) in best.items()}


def parse_chapter_xlsx(content: bytes) -> List[Tuple[str, str, float]]:
    """Parse a Data-Cruda chapter workbook → ``[(code, description, fob_value)]``.

    Locates the header row (``Valor FOB``) and reads chapter rows until the data
    ends. Rows without a numeric value are skipped (missing ≠ zero)."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = None
    for i, r in enumerate(rows[:20]):
        joined = " ".join(str(c) for c in r if c is not None).lower()
        if "valor fob" in joined or ("código" in joined and "capítulo" in joined):
            header_idx = i
            break
    if header_idx is None:
        return []

    out: List[Tuple[str, str, float]] = []
    for r in rows[header_idx + 1:]:
        if not r or r[0] is None or len(r) < 3:
            continue
        code = str(r[0]).strip()
        desc = str(r[1]).strip() if r[1] is not None else ""
        if not re.fullmatch(r"\d{1,2}", code):  # HS chapters 01-99 only; skips Total/blank
            continue
        val = r[2]
        if not isinstance(val, (int, float)):
            continue
        out.append((code.zfill(2), desc, float(val)))
    return out


class DGAClient(FixtureBackedClient):
    source = "DGA"
    license = "Datos públicos DGA — estadísticas de comercio exterior"
    license_ok = True
    fixture_file = "dga.json"
    live_phase = "Fase 5"


dga_client = DGAClient()
