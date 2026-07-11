"""BCRD national accounts — value added by economic activity (sector backbone).

Feeds Eje 3 (`sector_intel`) with real, full-economy sector data derived from the
BCRD's *PIB por sectores de origen* (production / value-added approach):

  * ``sector_size``   — the sector's share of total Value Added (nominal, %).
  * ``sector_growth`` — the sector's real year-over-year growth (%).

for the ~17 *leaf* sectors that together sum to total Value Added (the whole
economy), replacing the thin 3-anchor fixture.

Source: ``cdn.bancentral.gov.do/.../pib_origen_2018.xlsx`` (base 2018, public —
no token, no IP allowlist, unlike the MacroVariables API). Workbook sheets:
``PIB$_Trim`` (nominal RD$ MM) and ``PIBK_Trim`` (real chained-volume). The current
file starts in 2018; the panel is extended back to **2007** with the BCRD's official
retropolado (``pib_origen_retro_2018_2007.xlsx``) as a HISTORICAL vintage — a coarser
16-activity split (no separate 'Servicios profesionales') whose 2018+ overlap matches
the current file exactly. See :func:`build_sector_records` for the vintage seam rule.

Hierarchy trap (handled): the table interleaves parent aggregates
("Industrias", "Servicios", "Valor Agregado"), leaf sectors, and sub-items
(e.g. "Industrias de Alimentos" under "Manufactura Local"). We extract **only a
verified leaf partition** whose nominal values sum to "Valor Agregado" — the sum
is asserted at parse time so a future structural change can't silently double
count (the SIB ``TODOS`` + children lesson, in national-accounts form).

Missing values stay ``None`` — never interpolated. The first year of the panel
has no prior year, so its ``sector_growth`` is ``None`` rather than fabricated.
``Record.dimension`` carries the sector slug (no ``country`` field), mirroring how
:mod:`shared.data.wdi_client` carries the ISO2 country.
"""
import logging
import re
import unicodedata
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage

logger = logging.getLogger("sdq.data.bcrd_sectors")

# Base-2018 "PIB por sectores de origen", trimestral. Public BCRD CDN.
PIB_ORIGEN_URL = (
    "https://cdn.bancentral.gov.do/documents/estadisticas/sector-real/"
    "documents/pib_origen_2018.xlsx"
)
SHEET_NOMINAL = "PIB$_Trim"   # value added at current prices (RD$ MM)
SHEET_REAL = "PIBK_Trim"      # value added at chained-volume (real) — for growth

# Retropolado oficial del BCRD (misma base 2018, empalme hacia atrás): serie larga
# 2007-2026. Es un VINTAGE MÁS GRUESO — 16 actividades: NO desagrega "Servicios
# profesionales" (lo mantiene dentro de "Otras Actividades de Servicios de Mercado").
# Verificado que, en el solapamiento 2018+, las 16 actividades coinciden EXACTO con el
# archivo vigente; solo el residual difiere por esa (des)agregación. Se usa como fuente
# HISTÓRICA (años previos al primer año del archivo vigente), rotulada retropolada.
PIB_ORIGEN_RETRO_URL = (
    "https://cdn.bancentral.gov.do/documents/estadisticas/sector-real/"
    "documents/pib_origen_retro_2018_2007.xlsx"
)

TOTAL_LABEL = "Valor Agregado"

# The leaf partition: (slug, BCRD label, display name). These 17 sectors sum to
# "Valor Agregado" (verified). Order roughly follows the BCRD table. Slugs
# ``turismo`` / ``zonas_francas`` / ``energia`` are kept from the prior anchors.
SECTORS: List[Tuple[str, str, str]] = [
    ("agropecuario", "Agropecuario", "Agropecuario"),
    ("mineria", "Explotación de Minas y Canteras", "Minería y Canteras"),
    ("manufactura_local", "Manufactura Local", "Manufactura Local"),
    ("zonas_francas", "Manufactura Zonas Francas", "Manufactura Zonas Francas"),
    ("construccion", "Construcción", "Construcción"),
    ("energia", "Energía y Agua", "Energía y Agua"),
    ("comercio", "Comercio", "Comercio"),
    ("turismo", "Hoteles, Bares y Restaurantes", "Turismo (Hoteles/Bares/Rest.)"),
    ("transporte", "Transporte y Almacenamiento", "Transporte y Almacenamiento"),
    ("comunicaciones", "Comunicaciones", "Comunicaciones"),
    ("financiero", "Intermediación Financiera, Seguros y Actividades Conexas",
     "Intermediación Financiera"),
    ("inmobiliario", "Actividades Inmobiliarias y de Alquiler", "Inmobiliario y Alquiler"),
    ("ensenanza", "Enseñanza", "Enseñanza"),
    ("salud", "Salud", "Salud"),
    ("administracion_publica",
     "Administración Pública y Defensa; Seguridad Social de Afiliación Obligatoria "
     "y Otros Servicios", "Administración Pública"),
    ("servicios_profesionales", "Servicios profesionales", "Servicios Profesionales"),
    ("otros_servicios", "Otras Actividades de Servicios de Mercado",
     "Otros Servicios de Mercado"),
]

VAR_SIZE = "sector_size"
VAR_GROWTH = "sector_growth"
UNIT_SIZE = "% del Valor Agregado"
UNIT_GROWTH = "% (crecimiento real interanual)"
SIZE_SUM_TOLERANCE = 0.5  # % deviation of Σ leaves vs Valor Agregado that fails closed


class BCRDSectorsError(RuntimeError):
    """The PIB-origen workbook was missing, unreadable or structurally changed."""


def _norm(label: object) -> str:
    """Normalize a label for tolerant matching: strip accents, casefold, collapse
    whitespace — so spacing / accent / casing variants of a renamed BCRD label
    still match (the SIB accent-insensitive lesson)."""
    if label is None:
        return ""
    s = unicodedata.normalize("NFKD", str(label))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.casefold().split())


# Normalized BCRD label → slug. Built via ``_norm`` so the keys match exactly how
# workbook rows are normalized at parse time (accent/case/spacing tolerant).
_LABEL_TO_SLUG: Dict[str, str] = {_norm(label): slug for slug, label, _name in SECTORS}


def _year_token(cell: object) -> Optional[int]:
    """Parse a year header cell (``2018`` int or ``"2021 (p)"`` string)."""
    if isinstance(cell, bool):
        return None
    if isinstance(cell, (int, float)) and float(cell).is_integer():
        y = int(cell)
        return y if 2000 <= y <= 2035 else None
    if isinstance(cell, str):
        m = re.match(r"\s*(\d{4})", cell)
        if m and 2000 <= int(m.group(1)) <= 2035:
            return int(m.group(1))
    return None


def _complete_year_columns(header_row: tuple) -> Dict[int, int]:
    """For one candidate header row, return ``{col → year}`` keeping only years
    with exactly 4 quarter columns (so a partial latest year is excluded)."""
    col_year: Dict[int, int] = {}
    current: Optional[int] = None
    for j, cell in enumerate(header_row):
        if j == 0:
            continue  # label column
        y = _year_token(cell)
        if y is not None:
            current = y
        if current is not None:
            col_year[j] = current
    counts: Dict[int, int] = {}
    for y in col_year.values():
        counts[y] = counts.get(y, 0) + 1
    return {j: y for j, y in col_year.items() if counts.get(y) == 4}


def _column_years(rows: List[tuple]) -> Dict[int, int]:
    """Map each data column index → its year (forward-filled across quarters).

    Chooses the header row that yields the *most complete years* (4-quarter
    groups), so a stray title/note row with a few sparse year tokens can't be
    mistaken for the real quarter header (M2 — fragility hardening).
    """
    best: Dict[int, int] = {}
    best_years = 0
    for r in rows[:15]:
        if sum(1 for c in r if _year_token(c) is not None) < 3:
            continue
        complete = _complete_year_columns(r)
        n_years = len(set(complete.values()))
        if n_years > best_years:
            best_years, best = n_years, complete
    if not best:
        raise BCRDSectorsError(
            "no se encontró una fila de años con trimestres completos en el libro PIB-origen"
        )
    return best


def _annual_by_sector(rows: List[tuple], col_year: Dict[int, int]) -> Dict[str, Dict[int, float]]:
    """Sum quarter columns into annual totals per tracked label, **first block only**.

    Each sheet stacks several blocks that repeat the same sector labels: the
    nominal sheet has a RD$-level block then a structure-share block; the real
    sheet has a volume-index block then growth-% and incidence blocks. Only the
    *first* block carries the quantity we want (the level / the volume index). The
    first "Valor Agregado" row closes that block — we take the first occurrence of
    each label and stop at the next "Valor Agregado". Summing across blocks would
    corrupt the real YoY growth (the size shares survive because the level
    dominates and cancels in the ratio, but growth does not).
    """
    wanted = set(_LABEL_TO_SLUG) | {_norm(TOTAL_LABEL)}
    total = _norm(TOTAL_LABEL)
    out: Dict[str, Dict[int, float]] = {}
    for r in rows:
        if not r or not r[0]:
            continue
        label = _norm(r[0])
        if label == total and total in out:
            break  # second "Valor Agregado" → next block; the first block is done
        if label not in wanted or label in out:
            continue  # first occurrence only
        bucket: Dict[int, float] = {}
        for j, year in col_year.items():
            if j < len(r) and isinstance(r[j], (int, float)) and not isinstance(r[j], bool):
                bucket[year] = bucket.get(year, 0.0) + float(r[j])
        out[label] = bucket
    return out


def _published_at() -> Optional[date]:  # pragma: no cover - network I/O
    """The CDN's ``Last-Modified`` for the workbook = when the BCRD published it.

    A real provenance signal (not fabricated). Returns ``None`` on any failure —
    a missing publication date never blocks ingestion.
    """
    import httpx
    try:
        resp = httpx.head(PIB_ORIGEN_URL, timeout=20, follow_redirects=True)
        lm = resp.headers.get("last-modified")
        if lm:
            return datetime.strptime(lm, "%a, %d %b %Y %H:%M:%S %Z").date()
    except Exception as e:  # noqa: BLE001 — provenance is best-effort
        logger.warning("[bcrd_sectors] no se pudo leer Last-Modified: %s", e)
    return None


def _read_sheet(path, sheet: str) -> List[tuple]:
    import openpyxl  # lazy: app boots without a workbook in hand
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise BCRDSectorsError(f"hoja '{sheet}' ausente en el libro PIB-origen")
    return list(wb[sheet].iter_rows(values_only=True))


def parse_origin_workbook(path) -> Tuple[Dict[str, Dict[int, float]], Dict[str, Dict[int, float]]]:
    """Return ``(nominal, real)`` annual maps ``{norm_label: {year: value}}``.

    Pure of network — operates on a local workbook path so it is unit-testable.
    """
    nominal_rows = _read_sheet(path, SHEET_NOMINAL)
    real_rows = _read_sheet(path, SHEET_REAL)
    nominal = _annual_by_sector(nominal_rows, _column_years(nominal_rows))
    real = _annual_by_sector(real_rows, _column_years(real_rows))
    return nominal, real


def _yoy(real: Dict[str, Dict[int, float]], label: str, year: int) -> Optional[float]:
    """Real YoY % for *label* at *year* from one vintage's chained-volume map, or
    ``None`` when either endpoint is missing. Kept within a single vintage so a
    seam between the current and the retropolated series never yields a
    cross-vintage ratio (which would be meaningless for a redefined activity)."""
    s = real.get(label, {})
    now, prev = s.get(year), s.get(year - 1)
    if now is None or prev in (None, 0):
        return None
    return round(100.0 * (now / prev - 1.0), 2)


def _partition_dev(nominal: Dict[str, Dict[int, float]], year: int, va: float) -> float:
    """Deviation % of Σ(leaf labels present) vs Valor Agregado for *year*. Labels
    absent in this vintage (e.g. 'Servicios profesionales' in the 16-activity
    retropolado) contribute 0 — the remaining leaves still close to the total."""
    leaf_sum = sum(nominal[lbl].get(year, 0.0) for lbl in _LABEL_TO_SLUG if lbl in nominal)
    return 100.0 * (leaf_sum / va - 1.0) if va else 0.0


def build_sector_records(
    nominal: Dict[str, Dict[int, float]],
    real: Dict[str, Dict[int, float]],
    *,
    historical: Optional[Tuple[Dict[str, Dict[int, float]], Dict[str, Dict[int, float]]]] = None,
    url: Optional[str] = PIB_ORIGEN_URL,
    license_: str = "datos oficiales BCRD — uso público con cita",
    published_at: Optional[date] = None,
) -> List[Record]:
    """Compute ``sector_size`` + ``sector_growth`` records from annual maps.

    Pure function (no I/O): ``size = 100·nominal_sector / Valor Agregado``;
    ``growth = real YoY %`` (``None`` when no prior year).

    *nominal*/*real* are the **current** vintage (authoritative, most recent
    disaggregation — must carry all 17 leaves). *historical*, if given, is the
    ``(nominal, real)`` of the BCRD **retropolado** (a longer but coarser vintage,
    16 activities) used for years *before* the current vintage begins. Per year the
    size/partition come from one vintage (current from its first year on, historical
    before that); growth is always computed WITHIN a single vintage (current
    preferred, else historical) so the seam never produces a cross-vintage ratio.
    Leaves the historical vintage lacks (``servicios_profesionales``) simply have no
    pre-seam history — an honest gap, never fabricated.

    **Anti double-count guard (hard):** the leaf partition must sum to "Valor
    Agregado" each year (per vintage). If it deviates beyond ``SIZE_SUM_TOLERANCE``
    — a parent aggregate leaking in or a leaf renamed out — this *raises*
    :class:`BCRDSectorsError` rather than persisting distorted shares. The sync
    catches it and surfaces a visible error (the SIB ``TODOS`` lesson, fail-closed).
    """
    total = nominal.get(_norm(TOTAL_LABEL), {})
    if not total:
        raise BCRDSectorsError(f"fila '{TOTAL_LABEL}' ausente: no se puede normalizar el tamaño")

    missing = [lbl for lbl in _LABEL_TO_SLUG if lbl not in nominal]
    if missing:
        # A renamed/removed leaf breaks the current partition → fail closed, don't
        # degrade. (Only the current vintage must be complete; the retropolado is
        # allowed to omit the finer 'Servicios profesionales' split.)
        raise BCRDSectorsError(
            f"etiquetas hoja ausentes en el libro (¿renombre del BCRD?): {sorted(missing)}"
        )

    nom_hist, real_hist = historical or ({}, {})
    total_hist = nom_hist.get(_norm(TOTAL_LABEL), {})
    cutover = min(total) if total else None  # first year of the current vintage
    note = ("PIB por sectores de origen (base 2018); tamaño=share del VAB, "
            "crecimiento=real interanual")
    if total_hist:
        note += ("; historia previa a %s = serie oficial BCRD retropolada (16 actividades: "
                 "'servicios_profesionales' sin split → arranca en %s; 'otros_servicios' "
                 "previo incluye servicios profesionales)" % (cutover, cutover))
    lineage = Lineage(
        source="BCRD", license=license_, fetched_at=date.today(), url=url,
        published_at=published_at, note=note,
    )

    all_years = sorted(set(total) | set(total_hist))
    out: List[Record] = []
    for year in all_years:
        use_hist = cutover is not None and year < cutover
        nom = nom_hist if use_hist else nominal
        va = (total_hist if use_hist else total).get(year)
        if not va:
            continue
        dev = _partition_dev(nom, year, va)
        if abs(dev) > SIZE_SUM_TOLERANCE:
            raise BCRDSectorsError(
                f"{year}: Σ hojas={sum(nom[lbl].get(year, 0.0) for lbl in _LABEL_TO_SLUG if lbl in nom):.0f} "
                f"vs Valor Agregado={va:.0f} (dif {dev:+.2f}%) — posible doble conteo "
                f"o cambio de estructura del BCRD"
            )
        for norm_label, slug in _LABEL_TO_SLUG.items():
            nomv = nom.get(norm_label, {}).get(year)
            size = round(100.0 * nomv / va, 3) if nomv is not None else None
            out.append(Record(
                series=VAR_SIZE, period=str(year), value=size,
                lineage=lineage, unit=UNIT_SIZE, dimension=slug,
            ))
            # Growth within a single vintage: current if it has both endpoints, else
            # the retropolado — never current[Y]/historical[Y-1] (a redefined activity
            # like 'otros_servicios' would give a spurious jump).
            growth = _yoy(real, norm_label, year)
            if growth is None:
                growth = _yoy(real_hist, norm_label, year)
            out.append(Record(
                series=VAR_GROWTH, period=str(year), value=growth,
                lineage=lineage, unit=UNIT_GROWTH, dimension=slug,
            ))
    return out


def _filter(records: List[Record], series: Optional[str], period: Optional[str]) -> List[Record]:
    if series:
        records = [r for r in records if r.series == series]
    if period:
        records = [r for r in records if r.period == period]
    return records


class BCRDSectorsClient(FixtureBackedClient):
    """Sector value-added from the BCRD PIB-origen workbook (Eje 3 backbone)."""

    source = "BCRD"
    license = "datos oficiales BCRD — uso público con cita"
    license_ok = True
    fixture_file = "bcrd_sectors.json"
    live_phase = "Fase 4 (Eje 3 · sectorial)"

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        self.check_license()
        if self.mode == "live":
            return self._fetch_live(series, period)
        return self._fetch_fixture(series, period)

    # ── Live ──────────────────────────────────────────────────────
    def _fetch_live(self, series: Optional[str], period: Optional[str]) -> List[Record]:  # pragma: no cover - network I/O
        from shared.data.bcrd_excel.download import fetch_excel
        path = fetch_excel(PIB_ORIGEN_URL)
        nominal, real = parse_origin_workbook(path)
        # Historia larga (2007+) desde el retropolado oficial: extiende el panel ~11
        # años atrás. Best-effort — si el retro falla (CDN cambia el nombre, red), se
        # ingiere solo el vigente (2018+) sin tumbar la sync.
        historical = None
        try:
            hist_path = fetch_excel(PIB_ORIGEN_RETRO_URL)
            historical = parse_origin_workbook(hist_path)
        except Exception as e:  # noqa: BLE001 — la historia extra nunca bloquea la ingesta
            logger.warning("[bcrd_sectors] retropolado no disponible, solo serie vigente: %s", e)
        records = build_sector_records(
            nominal, real, historical=historical, published_at=_published_at())
        return _filter(records, series, period)

    # ── Fixture (offline / tests) ─────────────────────────────────
    def _fetch_fixture(self, series: Optional[str], period: Optional[str]) -> List[Record]:
        """Fixture shape: ``{"<slug>": {"<var>": {"<year>": value}}}``."""
        fixture = self._load_fixture(self.fixture_file)
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today())
        out: List[Record] = []
        for slug, vars_ in fixture.items():
            for var, obs in vars_.items():
                unit = UNIT_SIZE if var == VAR_SIZE else UNIT_GROWTH if var == VAR_GROWTH else None
                for yr, val in obs.items():
                    out.append(Record(
                        series=var, period=str(yr),
                        value=None if val is None else float(val),
                        lineage=lineage, unit=unit, dimension=slug,
                    ))
        return _filter(out, series, period)


def sector_catalog() -> List[Tuple[str, str]]:
    """``[(slug, display_name)]`` for seeding the Sector table."""
    return [(slug, name) for slug, _label, name in SECTORS]


bcrd_sectors_client = BCRDSectorsClient()
