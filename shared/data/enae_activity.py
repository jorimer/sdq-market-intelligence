"""ENAE structural-financial activity (ONE) — real per-sector economics for the IAI.

Pulls the ONE *Encuesta Nacional de Actividad Económica* (ENAE) pre-tabulated
``.xlsx`` tables ("según sector económico", 2015-2022): income, costs+expenses,
profit and profitability — the real structural variables that today's IAI fakes
with a flat rubric in its business/talent dimensions. The ENAE covers **9 survey
sectors** (a partial cut of the economy, NOT the 17 national-accounts slugs);
:mod:`shared.data.sector_crosswalk` maps ENAE sector ↔ BCRD-17 slugs. This
connector emits values **per ENAE sector** (the real resolution) — the per-slug
derivation (and the electricity+water → energía combine) is a later phase.
``Record.dimension`` carries the ENAE sector key, mirroring how
:mod:`shared.data.encft_employment` carries the activity branch.

Fail-closed guards (no silent partial/mis-parsed panel):
  * any of the 9 ENAE sectors missing from a table → the ONE changed it → raise.
  * a sector-looking row that doesn't map → nomenclature changed → raise.
  * for the additive level tables (income/costs/profit) Σ sectors must equal the
    table's own ``Total`` row each year → raise on deviation.
  * cross-table check: published ``rentabilidad`` ≈ ``utilidad / ingresos`` (it is
    a return-on-revenue margin) → catches a column mis-parse across files → raise.
Missing values stay ``None`` — never interpolated.
"""
import io
import logging
import re
import urllib.parse
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from shared.data._text import find_media_xlsx, norm
from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage
from shared.data.sector_crosswalk import ENAE_KEYS, map_enae_label

logger = logging.getLogger("sdq.data.enae_activity")

# ONE ENAE landing (Umbraco); the tabulado links are rotating /media/<hash>/… paths
# → scraped, not hard-coded (the media hash rotates between revisions).
ENAE_LANDING = (
    "https://www.one.gob.do/datos-y-estadisticas/temas/estadisticas-economicas/"
    "estadisticas-empresariales/encuesta-nacional-de-actividad-economica-enae/"
)
_HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP ENAE activity connector)"}

LICENSE = "datos oficiales ONE — uso público con cita"
SUM_TOLERANCE_PCT = 0.5     # Σ sectors vs the Total row that fails closed (level tables)
MARGIN_TOLERANCE = 0.01     # |rentabilidad − utilidad/ingresos| absolute, fails closed

# Each variable: its landing-page filename fragment, whether the table carries a
# national ``Total`` row (additive levels do; the rentabilidad ratio does not), and
# the published unit. The "según sector económico" fragment disambiguates from the
# sibling "según componentes" tables.
VAR_INGRESOS = "ingresos"
VAR_COSTOS = "costos_gastos"
VAR_UTILIDAD = "utilidad"
VAR_RENTABILIDAD = "rentabilidad"

VARIABLES: Dict[str, Dict[str, object]] = {
    VAR_INGRESOS: {
        "fragment": "ingresos-totales-de-las-empresas-por-ano-segun-sector-economico",
        "additive": True, "unit": "millones RD$"},
    VAR_COSTOS: {
        "fragment": "costos-y-gastos-totales-de-las-empresas-por-ano-segun-sector-economico",
        "additive": True, "unit": "millones RD$"},
    VAR_UTILIDAD: {
        "fragment": "utilidad-del-ejercicio-por-ano-segun-sector-economico",
        "additive": True, "unit": "millones RD$"},
    VAR_RENTABILIDAD: {
        "fragment": "rentabilidad-por-ano-segun-sector-economico",
        "additive": False, "unit": "ratio (utilidad/ingresos)"},
}


class EnaeError(RuntimeError):
    """The ENAE tabulados were missing, unreadable or structurally changed."""


def _year_columns(rows: List[tuple]) -> Tuple[Dict[int, int], int]:
    """``({year: column}, header_row_index)`` from the year-header row.

    These single-value tables have one column per year; the year token sits at that
    column. Picks the row with the most year tokens (so a title row can't be mistaken
    for it) and returns its index so the caller skips it.
    """
    best: Dict[int, int] = {}
    best_idx = -1
    for idx, r in enumerate(rows[:12]):
        cur: Dict[int, int] = {}
        for ci, c in enumerate(r):
            if ci == 0:
                continue  # sector label column
            y: Optional[int] = None
            if isinstance(c, (int, float)) and not isinstance(c, bool) and float(c).is_integer():
                y = int(c)
            elif isinstance(c, str):
                m = re.match(r"\s*(\d{4})", c)
                y = int(m.group(1)) if m else None
            if y is not None and 2000 <= y <= 2035:
                cur.setdefault(y, ci)
        if len(cur) > len(best):
            best, best_idx = cur, idx
    if not best:
        raise EnaeError("no se encontró la fila de años en el tabulado ENAE")
    return best, best_idx


def parse_sector_table(content: bytes) -> Tuple[Dict[str, Dict[int, float]], Dict[int, float]]:
    """Parse one ENAE tabulado → ``({sector_key: {year: value}}, {year: total})``.

    ``total`` is empty when the table has no ``Total`` row (the rentabilidad ratio).
    Pure of network — operates on bytes, unit-testable. Rows that don't map to one of
    the 9 ENAE sectors but carry year values raise (nomenclature change); note/title
    rows (no values) are skipped.
    """
    import openpyxl  # lazy: app boots without a workbook in hand

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    year_col, header_idx = _year_columns(rows)
    sectors: Dict[str, Dict[int, float]] = {}
    total: Dict[int, float] = {}
    unmapped: List[str] = []

    for i, r in enumerate(rows):
        if i == header_idx or not r or not isinstance(r[0], str) or not r[0].strip():
            continue
        label = r[0]
        values: Dict[int, float] = {}
        for year, col in year_col.items():
            v = r[col] if col < len(r) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                values[year] = round(float(v), 6)
        if norm(label) == "total":
            if values:
                total = values
            continue
        key = map_enae_label(label)
        if key is None:
            if values:  # numbers we couldn't map → the ONE renamed a sector
                unmapped.append(label.strip())
            continue     # otherwise a title/note row → skip
        sectors[key] = values

    if unmapped:
        raise EnaeError(
            f"sectores ENAE sin mapear (¿cambió la nomenclatura de la ONE?): {sorted(unmapped)}"
        )
    return sectors, total


def _check_present(var: str, sectors: Dict[str, Dict[int, float]]) -> None:
    missing = [k for k in ENAE_KEYS if k not in sectors]
    if missing:
        raise EnaeError(f"{var}: sectores ENAE ausentes en el tabulado: {sorted(missing)}")


def _check_partition(var: str, sectors: Dict[str, Dict[int, float]], total: Dict[int, float]) -> None:
    """Σ sectors == Total each year (additive level tables only)."""
    if not total:
        raise EnaeError(f"{var}: fila 'Total' ausente — no se puede verificar la partición")
    sector_years = {y for obs in sectors.values() for y in obs}
    uncovered = sorted(sector_years - set(total))
    if uncovered:
        raise EnaeError(f"{var}: años con dato de sector sin fila 'Total': {uncovered}")
    for year, tot in sorted(total.items()):
        leaf = sum(sectors[k].get(year, 0.0) for k in sectors)
        if not tot:
            if leaf:
                raise EnaeError(f"{var} {year}: Total=0 con Σ sectores={leaf:.0f} — inconsistente")
            continue
        dev = 100.0 * (leaf / tot - 1.0)
        if abs(dev) > SUM_TOLERANCE_PCT:
            raise EnaeError(
                f"{var} {year}: Σ sectores={leaf:.0f} vs Total={tot:.0f} (dif {dev:+.2f}%) "
                "— posible sector faltante o cambio de estructura de la ONE"
            )


def _check_margin(parsed: Dict[str, Dict[str, Dict[int, float]]]) -> None:
    """Cross-table: published rentabilidad ≈ utilidad / ingresos (a return-on-revenue
    margin). A mismatch means a column was mis-parsed in one of the three files."""
    ren = parsed.get(VAR_RENTABILIDAD, {})
    uti = parsed.get(VAR_UTILIDAD, {})
    ing = parsed.get(VAR_INGRESOS, {})
    for key in ENAE_KEYS:
        for year, r in ren.get(key, {}).items():
            u, i = uti.get(key, {}).get(year), ing.get(key, {}).get(year)
            if u is None or not i:
                continue
            if abs(r - u / i) > MARGIN_TOLERANCE:
                raise EnaeError(
                    f"{key} {year}: rentabilidad publicada {r:.4f} ≠ utilidad/ingresos "
                    f"{u / i:.4f} — columnas mal parseadas"
                )


def build_records(
    parsed: Dict[str, Dict[str, Dict[int, float]]],
    totals: Dict[str, Dict[int, float]],
    *,
    urls: Optional[Dict[str, str]] = None,
    published_at: Optional[date] = None,
    license_: str = LICENSE,
) -> List[Record]:
    """Validate every table, then emit one Record per (variable, sector, year).

    ``parsed`` = ``{variable: {sector_key: {year: value}}}``; ``totals`` =
    ``{variable: {year: total}}``. Pure (no I/O).
    """
    for var in VARIABLES:
        if var not in parsed:
            raise EnaeError(f"falta el tabulado de '{var}'")
        _check_present(var, parsed[var])
        if VARIABLES[var]["additive"]:
            _check_partition(var, parsed[var], totals.get(var, {}))
    _check_margin(parsed)

    out: List[Record] = []
    for var, sectors in parsed.items():
        unit = str(VARIABLES[var]["unit"])
        lineage = Lineage(
            source="ONE", license=license_, fetched_at=date.today(),
            url=(urls or {}).get(var), published_at=published_at,
            note="Encuesta Nacional de Actividad Económica (ENAE), según sector económico",
        )
        for key in ENAE_KEYS:
            for year, value in sorted(sectors[key].items()):
                out.append(Record(
                    series=var, period=str(year), value=value,
                    lineage=lineage, unit=unit, dimension=key,
                ))
    return out


def _filter(records: List[Record], series: Optional[str], period: Optional[str]) -> List[Record]:
    if series:
        records = [r for r in records if r.series == series]
    if period:
        records = [r for r in records if r.period == period]
    return records


class EnaeActivityClient(FixtureBackedClient):
    """ONE ENAE structural-financial activity by survey sector (income/costs/profit)."""

    source = "ONE"
    license = LICENSE
    license_ok = True
    fixture_file = "enae_activity.json"
    live_phase = "Fase 4 (Eje 3 · ENAE estructural)"

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        self.check_license()
        if self.mode == "live":
            return self._fetch_live(series, period)
        return self._fetch_fixture(series, period)

    # ── Live ──────────────────────────────────────────────────────
    def _fetch_live(self, series: Optional[str], period: Optional[str]) -> List[Record]:  # pragma: no cover - network I/O
        import httpx

        resp = httpx.get(ENAE_LANDING, timeout=40, follow_redirects=True, headers=_HEADERS)
        resp.raise_for_status()
        fragments = {var: str(spec["fragment"]) for var, spec in VARIABLES.items()}
        paths = find_media_xlsx(resp.text, fragments)
        missing = [var for var in VARIABLES if var not in paths]
        if missing:
            raise EnaeError(
                f"no se encontraron los tabulados ENAE {missing} en la landing de la ONE "
                "(¿cambiaron los enlaces?)"
            )
        parsed: Dict[str, Dict[str, Dict[int, float]]] = {}
        totals: Dict[str, Dict[int, float]] = {}
        urls: Dict[str, str] = {}
        published: Optional[date] = None
        for var, path in paths.items():
            url = "https://www.one.gob.do" + urllib.parse.quote(path)
            urls[var] = url
            published = published or self._published_at(url)
            f = httpx.get(url, timeout=90, follow_redirects=True, headers=_HEADERS)
            f.raise_for_status()
            parsed[var], totals[var] = parse_sector_table(f.content)
        records = build_records(parsed, totals, urls=urls, published_at=published)
        return _filter(records, series, period)

    @staticmethod
    def _published_at(url: str) -> Optional[date]:  # pragma: no cover - network I/O
        """The CDN's ``Last-Modified`` for the workbook (best-effort provenance)."""
        import httpx
        try:
            resp = httpx.head(url, timeout=20, follow_redirects=True, headers=_HEADERS)
            lm = resp.headers.get("last-modified")
            if lm:
                return datetime.strptime(lm, "%a, %d %b %Y %H:%M:%S %Z").date()
        except Exception as e:  # noqa: BLE001 — provenance is best-effort
            logger.warning("[enae_activity] no se pudo leer Last-Modified: %s", e)
        return None

    # ── Fixture (offline / tests) ─────────────────────────────────
    def _fetch_fixture(self, series: Optional[str], period: Optional[str]) -> List[Record]:
        """Fixture shape: ``{"<sector_key>": {"<variable>": {"<year>": value}}}``."""
        fixture = self._load_fixture(self.fixture_file)
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today())
        out: List[Record] = []
        for key, vars_ in fixture.items():
            for var, obs in vars_.items():
                unit = str(VARIABLES.get(var, {}).get("unit") or "")
                for yr, val in obs.items():
                    out.append(Record(
                        series=var, period=str(yr),
                        value=None if val is None else float(val),
                        lineage=lineage, unit=unit, dimension=key,
                    ))
        return _filter(out, series, period)


enae_activity_client = EnaeActivityClient()
