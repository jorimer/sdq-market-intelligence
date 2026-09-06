"""CMF Chile — Reporte Mensual de Información Financiera del Sistema Bancario.

Indicadores del SISTEMA bancario chileno: morosidad de 90 días o más y su composición,
cartera deteriorada, provisiones por riesgo de crédito, rentabilidad, eficiencia operativa y
actividad crediticia. Todos calculados por el propio emisor y a nivel nacional; este conector
no agrega ni deriva nada.

**Por qué NO se usa la API de la CMF**, que existe, tiene credencial cargada y funciona. Se
midió lo que publica cada vía y la API publica menos, y peor, para este producto:

  · `api.cmfchile.cl/api-sbifv3` **no expone morosidad, cartera vencida, colocaciones ni
    provisiones por riesgo de crédito**. No es una omisión de la documentación: el propio
    emisor declara en su FAQ que «la API de CMF Bancos utiliza la misma información que
    tienen los sitios web de la institución» y que «al presente no hay una programación
    disponible» para ampliarla. Es un subconjunto declarado, no el catálogo completo.
  · Su adecuación de capital es la del **artículo 66 de la Ley General de Bancos** — el marco
    ANTERIOR a Basilea III—, mientras el portal publica la consolidada bajo Basilea III desde
    diciembre de 2020. Publicar el artículo 66 en 2026 sería citar un marco superado teniendo
    el vigente disponible.
  · Además cuesta cuota (10.000 peticiones mensuales) y **no declara la unidad** de las cifras
    de balance, que hay que deducir de otra fuente del emisor.

Estos archivos, en cambio, no piden credencial, no tienen límite de tasa, declaran su unidad
en la propia hoja y traen la fórmula contable de cada indicador al costado.

**El quiebre de 2022.** En enero de 2022 la CMF cambió el plan de cuentas (Compendio de
Normas Contables, códigos de 7 a 9 dígitos) y, en los archivos de balance, la ESCALA de
millones de pesos a pesos: para una misma cuenta, `26.995.865` en 2021-12 y
`26.458.735.676.131` en 2022-01. Seis órdenes de magnitud en el corte exacto de mes. Por eso
el lector exige que la hoja declare la versión del Compendio con la que se armó este mapeo y
FALLA si es otra: los mismos códigos bajo otro Compendio pueden medir otra cosa, y un
conector que siga sirviendo en silencio es el que publica la cifra cien veces mal.
"""
import logging
import re
from datetime import date
from typing import Dict, List, Optional, Tuple

from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage

logger = logging.getLogger("sdq.data.cmf")

#: Índice de los reportes financieros mensuales. El identificador de cada archivo
#: (`articles-NNNNNN_recurso_1.xlsx`) cambia todos los meses, así que hay que resolverlo acá.
INDICE_REPORTES = ("https://www.cmfchile.cl/portal/estadisticas/617/"
                   "w3-propertyvalue-28910.html")

#: El reporte se busca por su TÍTULO, que es lo que el emisor mantiene estable, nunca por el
#: identificador ni por el orden en la página. Mismo criterio que en SECMCA, y por la misma
#: razón: el nombre del archivo es un accidente de publicación.
REPORTE_MENSUAL = "Reporte Mensual de Información Financiera del Sistema Bancario"

#: La hoja del sistema. El libro trae 31 hojas; las otras 30 son por entidad o por anexo, y
#: este boletín solo publica agregados nacionales de los países que no son RD.
HOJA_SISTEMA = "Indicadores Sistema"

#: Versión del Compendio de Normas Contables con la que se construyó `INDICADORES`. Si la
#: hoja declara otra, el conector se detiene: ver el quiebre de 2022 en el encabezado.
CNCB_ESPERADO = "2022"

_MESES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
          "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
          "noviembre": 11, "diciembre": 12}

#: Lo que el emisor escribe donde no hay dato. Va a `None`. La planilla usa además `0` para
#: ceros REALES —«Adeudado por bancos» de un banco sin esa cartera— así que confundirlos
#: publicaría un cero que el emisor no afirmó.
_AUSENTES = {"---", "--", "-", "n/a", "na", "s/i", ""}


class CmfError(RuntimeError):
    """La página o la planilla de la CMF no vinieron en la forma esperada."""


def _norm(texto: object) -> str:
    """Colapsa espacios y baja a minúsculas. Sin esto, un doble espacio en un rótulo del
    emisor —los hay— rompe la coincidencia y el indicador desaparece sin ruido."""
    return " ".join(str(texto or "").split()).casefold()


def _norm_seccion(texto: object) -> str:
    """Como `_norm`, pero además saca la marca de nota al pie del final del título de sección
    («Rentabilidad Promedio (1)»): el emisor renumera las notas entre ediciones."""
    return re.sub(r"\s*\(\d+\)\s*$", "", " ".join(str(texto or "").split())).casefold()


def _norm_formula(texto: object) -> str:
    """La fórmula sin espacios. El emisor escribe `(85700.00.00+ 85800.00.00+ 859...)` con
    espacios irregulares dentro del paréntesis."""
    return re.sub(r"\s+", "", str(texto or ""))


#: (sección, fórmula) → clave de la serie.
#:
#: La FÓRMULA identifica cada fila, no el rótulo ni la posición: es lo más estable que la
#: planilla publica, porque nombra las cuentas contables que se dividen. Un rótulo se
#: reescribe y una fila se corre; el código dice qué mide.
#:
#: Pero la fórmula SOLA no alcanza, y esto se comprobó sobre el archivo real: `50000.00.00`
#: es a la vez la variación mensual y la de 12 meses de las colocaciones. Sin la sección, una
#: de las dos pisaría a la otra y publicaríamos una variación mensual rotulada como anual.
INDICADORES: Dict[Tuple[str, str], str] = {
    (_norm_seccion("Actividad variación 12 meses"),
     _norm_formula("50000.00.00")): "colocaciones_var_12m",

    (_norm_seccion("Rentabilidad Promedio (1)"),
     _norm_formula("59000.00.00 / 30000.00.00")): "roae",
    (_norm_seccion("Rentabilidad Promedio (1)"),
     _norm_formula("59000.00.00 / 10000.00.00")): "roaa",

    (_norm_seccion("Eficiencia operativa"),
     _norm_formula("(56000.00.00 / 55000.00.00)")): "eficiencia_operativa",

    (_norm_seccion("Provisiones constituidas por riesgo de crédito de colocaciones a costo "
                   "amortizado (2)"),
     _norm_formula("(14315.01.00+14325.01.00+14900.00.00) / 50500.00.00")):
        "provisiones_colocaciones",

    (_norm_seccion("Cartera con morosidad de 90 días o más"),
     _norm_formula("(85700.00.00+85800.00.00+85900.00.00) / 50000.00.00")):
        "mora_90_colocaciones",
    (_norm_seccion("Cartera con morosidad de 90 días o más"),
     _norm_formula("85720.00.00 / 14500.00.00")): "mora_90_comercial",
    (_norm_seccion("Cartera con morosidad de 90 días o más"),
     _norm_formula("(85730.00.00+85740.00.00) / (14600.00.00+14800.00.00)")):
        "mora_90_personas",
    # Consumo es 14800 y vivienda 14600, no al revés. Se verificó contra las cuatro secciones
    # de esta hoja y contra el archivo de mora individual, que coinciden; y la economía lo
    # confirma: el índice de provisiones da 8,1% para consumo y 0,65% para vivienda.
    (_norm_seccion("Cartera con morosidad de 90 días o más"),
     _norm_formula("85740.00.00 / 14800.00.00")): "mora_90_consumo",
    (_norm_seccion("Cartera con morosidad de 90 días o más"),
     _norm_formula("85730.00.00 / 14600.00.00")): "mora_90_vivienda",

    (_norm_seccion("Cartera deteriorada de colocaciones a costo amortizado"),
     _norm_formula("81100.00.00 / 50500.00.00")): "cartera_deteriorada",
}

#: Qué mide cada clave, para que el número no viaje sin su sujeto.
NOMBRES: Dict[str, str] = {
    "colocaciones_var_12m": "Colocaciones — variación en 12 meses",
    "roae": "Rentabilidad sobre patrimonio promedio, después de impuestos (ROAE)",
    "roaa": "Rentabilidad sobre activos promedio, después de impuestos (ROAA)",
    "eficiencia_operativa": "Gastos operacionales sobre ingresos operacionales",
    "provisiones_colocaciones": "Provisiones por riesgo de crédito sobre colocaciones",
    "mora_90_colocaciones": "Cartera con morosidad de 90 días o más sobre colocaciones",
    "mora_90_comercial": "Morosidad de 90 días o más — cartera comercial",
    "mora_90_personas": "Morosidad de 90 días o más — cartera de personas",
    "mora_90_consumo": "Morosidad de 90 días o más — cartera de consumo",
    "mora_90_vivienda": "Morosidad de 90 días o más — cartera de vivienda",
    "cartera_deteriorada": "Cartera deteriorada sobre colocaciones a costo amortizado",
}


def descubrir_ediciones(html: str, titulo: str = REPORTE_MENSUAL) -> Dict[str, str]:
    """Las ediciones publicadas de un reporte: ``{"2026-07": "articles-113057_recurso_1.xlsx"}``.

    La página agrupa cada reporte bajo un ``<h2>`` con su título y cuelga de él un enlace por
    mes, cuyo ``aria-label`` dice «Descargar Julio 2026 (xlsx…)». De ahí sale el período: es
    la única marca de mes que traen los enlaces.

    **Falla si el título coincide con más de un grupo**, y no es hipotético: la página trae
    trece reportes y TRES títulos están duplicados —cartera vencida, provisiones y Basilea
    III aparecen dos veces cada uno—. Elegir el primero sería elegir en silencio.
    """
    cabeceras = [
        (m.start(), " ".join(re.sub("<[^>]+>", "", m.group(1)).split()))
        for m in re.finditer(
            r'<h2[^>]*id="estadisticas_filtros_grupo_\d+_Label"[^>]*>(.*?)</h2>', html, re.S)
    ]
    if not cabeceras:
        raise CmfError(f"ningún reporte enlazado en {INDICE_REPORTES}: cambió la página")

    import html as _html
    buscado = _norm(titulo)
    cortes = [c[0] for c in cabeceras] + [len(html)]
    elegidos = [k for k, (_, t) in enumerate(cabeceras)
                if _norm(_html.unescape(t)) == buscado]
    if not elegidos:
        disponibles = [_html.unescape(t) for _, t in cabeceras]
        raise CmfError(f"no hay un reporte titulado {titulo!r}; hay {disponibles}")
    if len(elegidos) > 1:
        raise CmfError(
            f"{len(elegidos)} reportes se titulan {titulo!r} en la página: no se puede elegir "
            "sin criterio, y elegir el primero sería elegir en silencio")

    bloque = html[cabeceras[elegidos[0]][0]:cortes[elegidos[0] + 1]]
    fuera: Dict[str, str] = {}
    for m in re.finditer(
            r'<a\s+href="(articles-\d+_recurso_1\.xlsx)[^"]*"[^>]*'
            r'aria-label="Descargar\s+([^("]+?)\s*\(', bloque):
        periodo = _periodo_de_etiqueta(m.group(2))
        if periodo:
            fuera.setdefault(periodo, m.group(1))
    return fuera


def _periodo_de_etiqueta(etiqueta: str) -> Optional[str]:
    """«Julio 2026» → ``"2026-07"``. El emisor alterna mayúscula y minúscula en el mes
    ('Julio 2026' y 'diciembre 2013' conviven en la misma lista), así que se compara en
    minúsculas."""
    partes = _norm(etiqueta).split()
    if len(partes) < 2:
        return None
    mes = _MESES.get(partes[0])
    if not mes or not partes[-1].isdigit():
        return None
    return f"{int(partes[-1]):04d}-{mes:02d}"


def parse_valor(celda: object) -> Optional[float]:
    """El valor de una celda, o ``None`` si el emisor declara que no hay dato.

    Nunca devuelve 0.0 por un ausente: en un indicador de morosidad el cero es una afirmación
    fuerte —«ninguna cartera en mora»— y falsa.
    """
    if celda is None:
        return None
    if isinstance(celda, bool):
        return None
    if isinstance(celda, (int, float)):
        return float(celda)
    texto = str(celda).strip()
    if _norm(texto) in _AUSENTES:
        return None
    # La CMF escribe con coma decimal en algunas planillas y con punto en otras, dentro del
    # mismo emisor. Se toleran las dos y se rechaza cualquier otra cosa en vez de elegir una
    # por defecto: leer «15066018,00» como separador de miles da tres órdenes de magnitud.
    limpio = texto.replace("%", "").strip()
    if limpio.count(",") == 1 and "." not in limpio:
        limpio = limpio.replace(",", ".")
    try:
        return float(limpio)
    except ValueError:
        return None


def _exigir_encabezado(filas: List[List]) -> str:
    """Comprueba que la hoja siga declarando su ESCALA y su Compendio, y devuelve la versión.

    Es el corazón del conector, no una validación de trámite. La hoja dice «INDICADORES (en
    %)» y «Códigos según CNCB versión 2022»; si mañana pasa a fracción, o a otro Compendio,
    los mismos códigos pueden medir otra cosa y las mismas cifras significar otra magnitud.
    Preferimos detenernos a publicar una cifra cien veces mal, que es lo que ya nos pasó.
    """
    cabecera = " ".join(_norm(c) for fila in filas[:12] for c in fila if c is not None)
    if "(en %)" not in cabecera and "en %" not in cabecera:
        raise CmfError(
            "la hoja dejó de declarar que sus indicadores están en porcentaje: sin esa "
            "declaración no se sabe si 2,34 son 2,34% o 234%, y adivinarlo no es una opción")
    m = re.search(r"cncb\s+versi[oó]n\s+(\d{4})", cabecera)
    if not m:
        raise CmfError(
            "la hoja dejó de declarar la versión del Compendio de Normas Contables: el mapeo "
            "de este conector está atado a códigos contables que solo significan lo mismo "
            "dentro de una versión")
    if m.group(1) != CNCB_ESPERADO:
        raise CmfError(
            f"la hoja declara el Compendio versión {m.group(1)} y este conector mapeó la "
            f"{CNCB_ESPERADO}. En el cambio de 2022 las mismas cuentas cambiaron de escala y "
            "de longitud de código: hay que revisar el mapeo antes de volver a publicar")
    return m.group(1)


def _periodos_de(filas: List[List]) -> Tuple[int, List[str]]:
    """La fila de cortes y los períodos que declara, como ``YYYY-MM-DD``.

    El emisor pone tres columnas —el mismo mes del año anterior, el mes previo y el corriente—
    y las escribe como fechas, no como texto.
    """
    for i, fila in enumerate(filas[:14]):
        cortes = [(j, c) for j, c in enumerate(fila) if hasattr(c, "year")]
        if len(cortes) >= 2:
            return i, [f"{c.year:04d}-{c.month:02d}-{c.day:02d}" for _, c in cortes]
    raise CmfError("no se encontró la fila de períodos de la hoja del sistema")


def _columnas_de_periodo(filas: List[List], fila_cortes: int) -> List[int]:
    return [j for j, c in enumerate(filas[fila_cortes]) if hasattr(c, "year")]


def leer_indicadores_sistema(filas: List[List]) -> Dict[str, object]:
    """Lee la hoja `Indicadores Sistema` y devuelve períodos y series.

    Forma: ``{"cncb": "2022", "periodos": [...], "series": {clave: [valores]}}``.

    **Falla si falta algún indicador declarado en `INDICADORES`.** Un conector al que le
    desaparece una entrada no falla: sirve de menos, y el eje que la consumía se queda sin
    dato sin que nadie se entere. Ese modo de fallo ya costó un mapa entero en este repo.
    """
    cncb = _exigir_encabezado(filas)
    fila_cortes, periodos = _periodos_de(filas)
    columnas = _columnas_de_periodo(filas, fila_cortes)

    series: Dict[str, List[Optional[float]]] = {}
    seccion = ""
    for fila in filas[fila_cortes + 1:]:
        etiqueta = fila[1] if len(fila) > 1 else None
        formula = fila[8] if len(fila) > 8 else None
        valores = [parse_valor(fila[j]) if j < len(fila) else None for j in columnas]
        if etiqueta and not formula and not any(v is not None for v in valores):
            seccion = _norm_seccion(etiqueta)
            continue
        if not formula:
            continue
        clave = INDICADORES.get((seccion, _norm_formula(formula)))
        if clave:
            series[clave] = valores

    faltantes = sorted(set(INDICADORES.values()) - set(series))
    if faltantes:
        raise CmfError(
            f"la planilla no trajo {faltantes}: o el emisor cambió la fórmula de esos "
            "indicadores, o el lector dejó de reconocer su sección. No se publica un "
            "subconjunto en silencio")
    return {"cncb": cncb, "periodos": periodos, "series": series}


# ── El conector ───────────────────────────────────────────────────────────────────────
LICENSE = ("CMF Chile — reportes financieros publicados para descarga en cmfchile.cl. El "
           "emisor autoriza el uso y la publicación con mención de la fuente MÁS un enlace; "
           "no es una licencia abierta.")


class CMFClient(FixtureBackedClient):
    """Indicadores del sistema bancario chileno, del reporte mensual de la CMF."""

    source = "CMF Chile"
    license = LICENSE
    license_ok = True
    fixture_file = "cmf_chile.json"
    live_phase = "boletín regional (T-CL-1)"

    #: El marco vigente, no el del artículo 66 que publica la API. Aun así estas cifras NO se
    #: comparan en nivel contra otros países: cada supervisor define su cartera y su corte de
    #: mora. Lo hace cumplir el guard de no-comparabilidad del boletín.
    NORMA_CONTABLE = "CMF Chile — Compendio de Normas Contables 2022"
    COMPARABLE_ENTRE_PAISES: set = set()

    #: El emisor lo advierte en el propio archivo, así que viaja con el dato y no como nota
    #: al pie de un documento que nadie relee.
    NOTA_PROVISORIA = ("Información provisoria: el emisor declara que puede ser modificada "
                       "en cualquier momento.")

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        self.check_license()
        datos = self._fetch_live() if self.mode == "live" else self._fetch_fixture()
        return _filtrar(datos, series, period)

    # ── Live ──────────────────────────────────────────────────────
    def _fetch_live(self, edicion: Optional[str] = None) -> List[Record]:  # pragma: no cover
        import httpx
        import openpyxl

        cab = {"User-Agent": "Mozilla/5.0 (compatible; sdq-mip/1.0)"}
        # `follow_redirects` no es cosmético: el portal responde 302 y sin seguirlo se baja un
        # HTML de 261 bytes que openpyxl rechaza con un error que no nombra la causa.
        resp = httpx.get(INDICE_REPORTES, timeout=120, follow_redirects=True, headers=cab)
        resp.raise_for_status()
        ediciones = descubrir_ediciones(resp.text)
        if not ediciones:
            raise CmfError(f"ninguna edición de «{REPORTE_MENSUAL}» en {INDICE_REPORTES}")
        elegida = edicion or max(ediciones)
        if elegida not in ediciones:
            raise CmfError(f"no está publicada la edición {elegida}; hay {max(ediciones)}")

        url = f"https://www.cmfchile.cl/portal/estadisticas/617/{ediciones[elegida]}"
        bruto = httpx.get(url, timeout=300, follow_redirects=True, headers=cab)
        bruto.raise_for_status()
        import io
        libro = openpyxl.load_workbook(io.BytesIO(bruto.content), read_only=True,
                                       data_only=True)
        if HOJA_SISTEMA not in libro.sheetnames:
            raise CmfError(f"el libro no trae la hoja «{HOJA_SISTEMA}»: "
                           f"tiene {libro.sheetnames}")
        filas = [list(f) for f in libro[HOJA_SISTEMA].iter_rows(max_col=9, values_only=True)]
        return self._records_de(leer_indicadores_sistema(filas), url)

    # ── Fixture (offline) ─────────────────────────────────────────
    def _fetch_fixture(self) -> List[Record]:
        fixture = self._load_fixture(self.fixture_file)
        return self._records_de(fixture, fixture.get("url", ""))

    def _records_de(self, datos: Dict, url: str) -> List[Record]:
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today(),
                          url=url, note=self.NOTA_PROVISORIA)
        periodos = list(datos["periodos"])
        fuera: List[Record] = []
        for clave, valores in sorted(datos["series"].items()):
            for periodo, valor in zip(periodos, valores):
                fuera.append(Record(
                    series=clave,
                    period=periodo,
                    value=valor,
                    lineage=lineage,
                    unit="%",
                    dimension="CHL",
                    reason=None if valor is not None else "el emisor no publica el dato",
                ))
        return fuera


def _filtrar(records: List[Record], series: Optional[str],
             period: Optional[str]) -> List[Record]:
    if series:
        records = [r for r in records if r.series == series]
    if period:
        records = [r for r in records if r.period == period]
    return records
