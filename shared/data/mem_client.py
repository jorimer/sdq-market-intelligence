"""MEM — los tres índices comerciales del sector eléctrico, del anexo del Informe de Desempeño.

El Ministerio de Energía y Minas publica cada mes un **anexo XLSX** junto a su Informe de
Desempeño de las Empresas Eléctricas Estatales. Su hoja ``EDE's`` trae, agregado para las tres
distribuidoras, exactamente las tres magnitudes que la Ley 1-12 fija como indicadores 3.27,
3.28 y 3.29 — y el CRI viene con el mismo nombre que usa el legislador, «Índice de
Recuperación de Efectivo».

**Solo se leen los anexos de DICIEMBRE, y no es una comodidad.** Las columnas del anexo son
ventanas ACUMULADAS del año en curso: la de abril dice `Ene24-Abr24`. Tomar esa cifra como el
año sería publicar cuatro meses con el rótulo de doce, y en un indicador estacional eso no es
una aproximación, es otro número. La de diciembre dice `Ene24-Dic24` y es el año completo.
Cada anexo de diciembre trae además los DOS años anteriores, así que dos archivos cubren seis
años.

**Nada se lee por posición.** El año sale de la cabecera y la fila sale de su ETIQUETA. Un
anexo cambia de una edición a otra —entre 2021 y 2024 se agregaron hojas y el bloque
financiero se corrió una fila— y un parser por índice habría seguido devolviendo el número de
al lado sin fallar. Si una etiqueta no aparece, esto levanta excepción en vez de servir un
diccionario incompleto.

**El subsidio del Estado (indicador 3.30) SÍ se sirve, desde 2026-08-24.** Estaba declarado
acá como brecha porque «esa hoja no tiene cabecera de años». Tenía razón el diagnóstico y
estaba mal la conclusión: la hoja no tiene años porque sus columnas son MESES —enero a
diciembre más un acumulado—, y el año no hay que inferirlo, viene del anexo que se abrió.
Como solo se leen los de diciembre, ese acumulado es el año completo.

Y la fila no se toma por posición, que era el otro miedo. Se toma por etiqueta y se
comprueba con las cuentas del propio cuadro: el acumulado tiene que ser la suma de los doce
meses, y el bloque del TOTAL tiene que ser la suma de los bloques de las tres distribuidoras.
La etiqueta «Aportes del gobierno» aparece cinco veces en el anexo de 2020; la identidad dice
cuál es la que agrega, sin que haya que saberlo de antemano.
"""
from __future__ import annotations

import logging
import re
from typing import Dict, List, Sequence, Tuple

import httpx

logger = logging.getLogger("sdq.data.mem")

SOURCE = "MEM · Informe de Desempeño (anexo)"
BASE_CATEGORIA = "https://mem.gob.do/category/sector-electrico/informe-de-desempeno"
HOJA = "EDE's"
_TIMEOUT = 120
_UA = "Mozilla/5.0 (SDQMIP research; +https://sdqconsulting.com.do)"

#: Etiqueta en el anexo → clave de serie. Se toma la variante **Año Móvil** porque es la que
#: el emisor calcula sobre doce meses; en el anexo de diciembre coincide con la acumulada, y
#: preferirla deja el parser correcto si algún día se leyera otro mes.
FILAS = {
    "CRI - Año Movil (%)": "cri",
    "Pérdidas - Año Movil (%)": "perdidas",
    "Cobranzas - Año Movil (%)": "cobranzas",
}

#: `Ene24-Dic24`. Los dos años tienen que coincidir: una ventana `Ene24-Abr24` no es un año y
#: no entra. Es el guard que separa «doce meses» de «lo que va del año».
_RX_ANIO = re.compile(r"^\s*Ene(\d{2})\s*-\s*Dic(\d{2})\s*$", re.I)

_RX_XLSX_DIC = re.compile(r'href="([^"]*[Dd]iciembre[^"]*\.xlsx)"')


class MEMUnavailable(RuntimeError):
    """No se pudo obtener o leer el anexo. NUNCA se degrada a «no hay dato»."""


def anexos_de_diciembre(anios: List[int]) -> Dict[int, str]:
    """URL del anexo de diciembre por año de informe, descubierta del sitio del emisor.

    Se descubre en vez de fijarse: la ruta del archivo lleva el año y el mes en que se SUBIÓ,
    no el que mide (el de diciembre de 2024 vive bajo `/2025/03/`), así que componerla a mano
    la erraría cada vez que el emisor publica tarde.
    """
    out: Dict[int, str] = {}
    for a in anios:
        url = f"{BASE_CATEGORIA}/{a}-informe-de-desempeno/"
        try:
            r = httpx.get(url, timeout=_TIMEOUT, follow_redirects=True,
                          headers={"User-Agent": _UA})
            r.raise_for_status()
        except httpx.HTTPError as e:
            logger.warning("MEM: no se pudo listar el año %s (%s)", a, e)
            continue
        halladas = _RX_XLSX_DIC.findall(r.text)
        if halladas:
            out[a] = halladas[0]
    return out


def _columnas_de_anio(ws) -> Dict[int, int]:
    """Columna → año, leído de la cabecera. Solo entran las ventanas de DOCE meses."""
    cols: Dict[int, int] = {}
    for fila in (5, 6, 7):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(fila, c).value
            if not isinstance(v, str):
                continue
            m = _RX_ANIO.match(v)
            if m and m.group(1) == m.group(2):
                cols[c] = 2000 + int(m.group(1))
    return cols


def parse_anexo(contenido: bytes) -> Dict[str, Dict[int, float]]:
    """`{serie: {año: porcentaje}}` del anexo de un diciembre.

    El emisor guarda estas celdas como fracción y las MUESTRA como porcentaje. Se publica el
    porcentaje —que es lo que se lee en su informe y la magnitud en que la ley escribió sus
    metas— y no se convierte nada más: cualquier otro ajuste dejaría de cuadrar contra la
    fuente el día que alguien coteje.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(_io(contenido), data_only=True, read_only=True)
    except Exception as e:                                   # archivo corrupto o no-xlsx
        raise MEMUnavailable(f"no se pudo abrir el anexo ({type(e).__name__}: {e})")
    if HOJA not in wb.sheetnames:
        raise MEMUnavailable(f"el anexo no trae la hoja «{HOJA}»; hojas: {wb.sheetnames}")
    ws = wb[HOJA]

    cols = _columnas_de_anio(ws)
    if not cols:
        raise MEMUnavailable(
            "ninguna columna del anexo declara un año completo `EneNN-DicNN`. Puede ser un "
            "anexo de otro mes: sus columnas son acumuladas y no son el año.")

    etiquetas = {e: None for e in FILAS}
    for fila in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
        for celda in fila:
            v = celda.value
            if isinstance(v, str) and (t := " ".join(v.split())) in etiquetas:
                etiquetas[t] = celda.row

    if (faltan := [e for e, r in etiquetas.items() if r is None]):
        raise MEMUnavailable(
            f"el anexo no trae las filas {faltan}. Se busca por etiqueta a propósito: leerlas "
            f"por posición devolvería el número de al lado sin fallar.")

    out: Dict[str, Dict[int, float]] = {}
    for etiqueta, clave in FILAS.items():
        fila = etiquetas[etiqueta]
        serie: Dict[int, float] = {}
        for c, anio in sorted(cols.items()):
            v = ws.cell(fila, c).value
            if isinstance(v, (int, float)):
                serie[anio] = round(float(v) * 100.0, 4)
        out[clave] = serie
    return out


def _io(contenido: bytes):
    import io

    return io.BytesIO(contenido)


def descargar(url: str) -> bytes:  # pragma: no cover - I/O de red
    try:
        r = httpx.get(url, timeout=_TIMEOUT, follow_redirects=True,
                      headers={"User-Agent": _UA})
        r.raise_for_status()
        return r.content
    except httpx.HTTPError as e:
        raise MEMUnavailable(f"no se pudo descargar el anexo ({type(e).__name__}: {e})")


def series_anuales(anios: List[int], listar=None, bajar=None,
                   ) -> Dict[str, List[Tuple[str, float]]]:
    """`{serie: [(año, valor)]}` ascendente, uniendo los anexos de diciembre disponibles.

    Cada anexo trae tres años y los años se solapan entre archivos. Manda el archivo MÁS
    NUEVO: el emisor revisa sus cifras y la edición reciente es la que él sostiene hoy. La
    discrepancia se registra en vez de resolverse en silencio — un dato que cambia de valor
    entre ediciones es exactamente lo que un lector puede reprocharle a un informe viejo.

    `listar` y `bajar` se inyectan para que los tests ejerciten esta lógica sin red: la regla
    de precedencia entre ediciones es justo lo que hay que poder probar, y contra el sitio
    del emisor no se prueba nada.
    """
    listar = listar or anexos_de_diciembre
    bajar = bajar or descargar
    urls = listar(sorted(anios))
    if not urls:
        raise MEMUnavailable("el emisor no publicó ningún anexo de diciembre alcanzable")
    acumulado: Dict[str, Dict[int, float]] = {c: {} for c in FILAS.values()}
    for anio_informe in sorted(urls):                 # del más viejo al más nuevo
        datos = parse_anexo(bajar(urls[anio_informe]))
        for clave, serie in datos.items():
            for a, v in serie.items():
                previo = acumulado[clave].get(a)
                if previo is not None and abs(previo - v) > 0.05:
                    logger.info("MEM: %s %s revisado por el emisor: %.2f → %.2f",
                                clave, a, previo, v)
                acumulado[clave][a] = v
    return {c: [(str(a), v) for a, v in sorted(s.items())] for c, s in acumulado.items()}

# ── Indicador 3.30: el aporte del Gobierno a las distribuidoras ────────────────────────────

#: La hoja de resultados financieros. Es otra que la de los índices comerciales.
HOJA_FINANCIERA = "Anexo Res Financieros"

#: La etiqueta del aporte, EXACTA. En el mismo anexo vive «Aportes del Gobierno para
#: Inversión», que es otro concepto; la igualdad las separa y un `in` las confundiría.
FILA_APORTES = "APORTES DEL GOBIERNO"

#: Cuánto pueden fallar las identidades del cuadro para darse por cerradas.
TOLERANCIA_IDENTIDAD_PCT = 0.5


def _norm_texto(t: object) -> str:
    import unicodedata

    s = "".join(c for c in unicodedata.normalize("NFD", str(t if t is not None else ""))
                if unicodedata.category(c) != "Mn")
    return " ".join(s.upper().split())


def _rotulo_de(fila: Sequence[object]) -> str:
    """El rótulo de una fila: su PRIMER texto, no su primera celda.

    El cuadro tiene la etiqueta en la segunda columna y la primera vacía — leer `fila[0]`
    devolvía vacío para todas y el lector no encontraba ninguna fila. Buscar el primer texto
    sobrevive a que el emisor agregue o quite una columna de sangría.
    """
    for v in fila:
        if isinstance(v, str) and v.strip():
            return _norm_texto(v)
    return ""


def _numeros(fila: Sequence[object]) -> List[float]:
    return [float(v) for v in fila
            if isinstance(v, (int, float)) and not isinstance(v, bool)]


def acumulado_de(fila: Sequence[object]) -> float:
    """El acumulado del año de una fila mensual, comprobado contra la suma de los meses.

    La última columna numérica es el acumulado. No se toma por ser la última: se toma porque
    **es la suma de las doce anteriores**, y eso se verifica. Si el emisor agregara una
    columna al final —una variación, una proyección—, la identidad deja de cerrar y esto
    levanta en vez de publicar la columna de al lado.
    """
    nums = _numeros(fila)
    if len(nums) < 13:
        raise MEMUnavailable(
            f"la fila trae {len(nums)} números y el cuadro tiene doce meses más el "
            f"acumulado: no hay identidad que comprobar")
    meses, acumulado = nums[:-1], nums[-1]
    suma = sum(meses)
    if not suma or abs(acumulado - suma) / abs(suma) * 100.0 > TOLERANCIA_IDENTIDAD_PCT:
        raise MEMUnavailable(
            f"la última columna ({acumulado:,.2f}) no es la suma de los meses "
            f"({suma:,.2f}): o cambió el cuadro, o no es la fila que se cree")
    return acumulado


def aporte_del_gobierno(filas: Sequence[Sequence[object]]) -> float:
    """El aporte del Gobierno al TOTAL de las distribuidoras, en millones de US$.

    La etiqueta aparece varias veces —una por distribuidora, una por el total, y alguna en
    otro bloque del mismo anexo—, así que **cuál es el total lo dice la aritmética**: es la
    única cuyo acumulado equivale a la suma de otras tres. Comprobado en 2020, donde el total
    da 578,59 y las tres distribuidoras 22,39 + 15,46 + 21,61 = 59,46 en enero y cierran
    igual en el año.

    Elegir «la primera» habría funcionado hoy y sería una posición disfrazada de regla.
    """
    from itertools import combinations

    candidatos = [acumulado_de(f) for f in filas if _rotulo_de(f) == FILA_APORTES]
    if not candidatos:
        raise MEMUnavailable(
            f"el anexo no trae ninguna fila «{FILA_APORTES}»: o cambió la etiqueta, o se "
            f"está leyendo la hoja equivocada")
    totales = []
    for i, posible in enumerate(candidatos):
        resto = candidatos[:i] + candidatos[i + 1:]
        for tres in combinations(resto, 3):
            suma = sum(tres)
            if suma and abs(posible - suma) / abs(suma) * 100.0 <= TOLERANCIA_IDENTIDAD_PCT:
                totales.append(round(posible, 6))
                break
    unicos = sorted(set(totales))
    if len(unicos) == 1:
        return unicos[0]
    if not unicos:
        raise MEMUnavailable(
            f"ninguna de las {len(candidatos)} filas de aporte equivale a la suma de otras "
            f"tres: el cuadro dejó de tener el bloque de total sobre las tres distribuidoras")
    raise MEMUnavailable(
        f"{len(unicos)} filas cierran la identidad del total {unicos}: es ambiguo y elegir "
        f"una sería adivinar")


def subsidios_anuales(anios: List[int], listar=None,
                      bajar=None) -> List[Tuple[int, float]]:  # pragma: no cover - red
    """`[(año, millones de US$)]` del aporte del Gobierno, de los anexos de diciembre."""
    import pandas as pd

    listar = listar or anexos_de_diciembre
    bajar = bajar or descargar
    out: List[Tuple[int, float]] = []
    for anio, url in sorted(listar(anios).items()):
        try:
            df = pd.read_excel(_io(bajar(url)), sheet_name=HOJA_FINANCIERA, header=None)
            valor = aporte_del_gobierno([list(df.iloc[k]) for k in range(len(df))])
        except Exception as e:  # noqa: BLE001 — un año ilegible no se lleva a los demás
            logger.warning("[mem] 3.30 %s: %s", anio, e)
            continue
        out.append((anio, round(valor, 3)))
    if not out:
        raise MEMUnavailable(
            "ningún anexo de diciembre dio el aporte del Gobierno; no se degrada a «sin dato»")
    return out
