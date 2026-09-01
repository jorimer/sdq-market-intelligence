"""SISDOM — descubrimiento y descarga, común a todos sus libros.

El SISDOM publica un libro Excel por área temática y de ahí salen ya varias series del
IDM y de la END: el ingreso per cápita (*Pobreza y Distribución de Ingresos*), la
escolaridad promedio (*Educación*), la mortalidad infantil (*Salud*) y la población por
región de desarrollo (*Demográficos*). Descubrir el libro y bajarlo es idéntico en todos
los casos; lo único que cambia es el layout de la hoja. Esta plumbing vive acá para que no
haya dos copias que envejezcan por separado — que es exactamente cómo se rompe un conector
cuando el emisor cambia algo y solo se arregla una de las dos.

EL EMISOR SE MUDÓ (2026-09-01) — leer antes de tocar este módulo
---------------------------------------------------------------
El MEPyD fue absorbido por Hacienda y ``mepyd.gob.do`` **dejó de existir**: responde 301
a ``www.hacienda.gob.do``, donde el plugin *WP File Download* no está instalado. El
listado viejo (``/sisdom/areas-tematicas``) da 404 y la descarga por ``admin-ajax.php``
da 400. Es decir: el conector no fallaba con un dato viejo, fallaba entero — y la última
corrida sana en producción fue el 2026-08-31.

Lo que la mudanza confirmó es la advertencia que este mismo módulo llevaba escrita: había
DOS copias de esta plumbing —acá y en :mod:`shared.data.sisdom_income`— y se arregló una
sola cuando el END se recableó a Hacienda, así que income/escolaridad/salud/pobreza se
quedaron apuntando al host muerto. La copia se borró; lo vigila
``shared/data/tests/test_una_sola_plumbing_del_sisdom.py``.

Cómo se descubre ahora, y por qué en dos pasos
----------------------------------------------
Hacienda no publica un índice de archivos: publica una *publicación* por edición, y los
``.xlsx`` cuelgan de su página. Así que el descubrimiento es:

1. La **API REST del sitio** (``/wp-json/wp/v2/search``) responde qué publicaciones del
   SISDOM existen. De ahí sale la edición más nueva por el año de su título.
2. La **página de esa edición** trae un ancla por libro, con su título y su ``.xlsx``.

Ni el slug de la edición ni la URL del archivo se clavan: rotan con cada edición anual, y
clavarlos es precisamente lo que dejó al conector de la ONE roto el día que cambió la
página. Lo único declarado es el fragmento del título de cada libro, que es una decisión
nuestra.

**Un fragmento ambiguo NO elige**: si calza más de un libro, falla nombrando los dos. Un
desempate silencioso entregaría otra área temática con la misma pinta, y el parser de la
hoja fallaría después con un error que no habla de la causa.
"""
from __future__ import annotations

import io
import re
import unicodedata
from html import unescape
from typing import Any, List, Optional, Sequence, Tuple

#: Paso 1: qué ediciones del SISDOM publica el emisor. Es la API del propio sitio, no un
#: raspado: devuelve título y URL canónica de cada publicación.
LISTING_URL = ("https://www.hacienda.gob.do/wp-json/wp/v2/search"
               "?search=sisdom&subtype=publicacion&per_page=30")

#: El sello con el que estas series se persisten. Se mantiene "MEPyD" A PROPÓSITO aunque el
#: ministerio ya no exista: `source` es la CLAVE del registro de licencias
#: (`modules/social_dev/service.py`) y de la cuarentena de la Data API, así que renombrarlo
#: acá dejaría a las filas nuevas sin licencia resoluble — una serie sin licencia no se
#: sirve. El cambio de emisor a MHE es una migración propia, con su mapa de licencias y su
#: reescritura de las filas ya persistidas, no un efecto colateral de reparar el conector.
SOURCE = "MEPyD"
#: Quién publica HOY el archivo. Viaja aparte del sello para que el linaje pueda decir de
#: qué host salió el .xlsx sin tocar la clave del registro de licencias.
PUBLICA_HOY = "Ministerio de Hacienda y Economía (MHE) — absorbió al MEPyD"
LICENSE = "SISDOM (VAES/MEPyD) — indicadores sociales oficiales, uso público con cita"

HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP SISDOM connector)"}

# Regionalizaciones, en orden de preferencia. La de 685-00 queda FUERA a propósito: usa
# otro conjunto de regiones, y mezclarla produciría una serie incoherente en silencio.
REGIONALIZATIONS = ("345-22", "710-04")
BLOCK_HEADER = "regiones de desarrollo"

#: Ancla de descarga en la página de una edición: cualquier ``<a>`` cuyo href sea un libro.
_BOOK_ANCHOR = re.compile(r'<a\b[^>]*href="([^"]+\.xlsx?)"[^>]*>(.*?)</a>', re.S | re.I)
_TAGS = re.compile(r"<[^>]+>")


class SisdomUnavailable(RuntimeError):
    """El listado o un libro del SISDOM no se pudo leer. Lleva el motivo."""


def norm(s: object) -> str:
    """Clave insensible a acento/caso/espacios para calzar títulos y rótulos."""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return " ".join(t.casefold().split())


def edition_year(title: str) -> int:
    """Mayor año de 4 dígitos del título ('SISDOM 2024 – …') → 2024, o 0."""
    return max((int(y) for y in re.findall(r"(?:19|20)\d{2}", title)), default=0)


def discover_publication(payload: Sequence[Any]) -> Optional[Tuple[int, str]]:
    """``(año de la edición, URL de la publicación)`` de la edición MÁS NUEVA, o ``None``.

    Pura para poder fijarla en tests: recibe el JSON ya decodificado de la búsqueda. El
    desempate por año es determinista y el mismo que usaba el listado anterior.
    """
    best: Optional[Tuple[int, str]] = None
    for post in payload or ():
        if not isinstance(post, dict):
            continue
        url = post.get("url") or post.get("link")
        title = unescape(_TAGS.sub("", str(post.get("title") or "")))
        if not url or "sisdom" not in norm(title):
            continue
        year = edition_year(title)
        if best is None or year > best[0]:
            best = (year, str(url))
    return best


def discover_book_url(html_text: str, fragment: str) -> str:
    """URL del ``.xlsx`` cuyo título de ancla contiene *fragment*.

    Falla —nunca elige— si no calza ninguno o si calza más de uno: los nueve libros de una
    edición se llaman casi igual («Indicadores de Salud», «Indicadores de Educación»), y un
    desempate silencioso entrega otra área temática cuyo parser después falla con un error
    de layout que no dice nada de la causa real.
    """
    objetivo = norm(fragment)
    calzan = []
    for href, inner in _BOOK_ANCHOR.findall(html_text):
        title = unescape(_TAGS.sub("", inner)).strip()
        if objetivo and objetivo in norm(title):
            calzan.append((title, unescape(href)))
    if not calzan:
        raise SisdomUnavailable(
            f"no se encontró el libro '{fragment}' en la página de la edición "
            "(¿cambió el título o se despublicó?)")
    if len(calzan) > 1:
        raise SisdomUnavailable(
            f"el fragmento '{fragment}' calza {len(calzan)} libros de la misma edición "
            f"({', '.join(t for t, _u in calzan)}): se niega a elegir")
    return calzan[0][1]


def fetch_book(fragment: str) -> bytes:  # pragma: no cover - network I/O
    """El ``.xlsx`` del libro, de la edición vigente. Sin el año: la mayoría no lo necesita."""
    return fetch_book_con_edicion(fragment)[1]


def fetch_book_con_edicion(fragment: str) -> Tuple[int, bytes]:  # pragma: no cover - network I/O
    """``(año de la edición, .xlsx)``. El año importa cuando el consumidor une VARIAS
    ediciones y tiene que saber cuál le habló — es el caso de los indicadores de la END,
    donde una discrepancia entre ediciones se registra nombrando a cada una."""
    import httpx

    try:
        with httpx.Client(timeout=180, follow_redirects=True, headers=HEADERS) as client:
            listado = client.get(LISTING_URL)
            listado.raise_for_status()
            edicion = discover_publication(listado.json())
            if edicion is None:
                raise SisdomUnavailable(
                    f"el emisor no lista ninguna edición del SISDOM en {LISTING_URL} "
                    "(¿cambió la API del sitio o se despublicó el sistema?)")
            anio, url_edicion = edicion
            pagina = client.get(url_edicion)
            pagina.raise_for_status()
            book = client.get(discover_book_url(pagina.text, fragment))
            book.raise_for_status()
    except httpx.HTTPError as e:
        raise SisdomUnavailable(
            f"no se pudo bajar el libro del SISDOM ({type(e).__name__}: {e})")
    if not book.content.startswith(b"PK"):
        # El sitio devuelve 200 con HTML cuando el archivo ya no existe: sin este guard,
        # openpyxl fallaría con un error de zip que no dice nada sobre la causa real.
        raise SisdomUnavailable(
            f"la descarga del libro '{fragment}' no es un .xlsx "
            f"({len(book.content)} bytes, {book.headers.get('content-type')})")
    return anio, book.content


def open_sheet(content: bytes, sheet: str) -> List[list]:
    """Abre la hoja por nombre normalizado — los libros escriben ``'03 3 021 '`` con
    espacio final. Devuelve las filas como listas; cierra el libro siempre."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        name = next((n for n in wb.sheetnames if norm(n) == norm(sheet)), None)
        if name is None:
            raise SisdomUnavailable(
                f"el libro no trae la hoja '{sheet}' (hojas: {wb.sheetnames[:12]}…)")
        return [list(r) for r in wb[name].iter_rows(values_only=True)]
    finally:
        wb.close()
