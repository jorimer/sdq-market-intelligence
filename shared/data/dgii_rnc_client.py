"""DGII — Estadísticas de Contribuyentes por Actividad Económica (RNC · CIIU rev.3, subclase).

Conector de la **Fuente A** del padrón de contribuyentes activos, agregado a nivel de
subclase CIIU (docs/SPEC_GATE_HONESTIDAD_Y_FUENTES_DGII.md §B). Alimenta el motor de
research libre con el conteo real de participantes por vertical (comida rápida, hoteles,
supermercados, farmacias…) — dato que hoy ningún eje del catálogo aporta.

Hermano de ``dgii_client.py`` (pulso fiscal): misma fuente/licencia, mismo idiom de
descarga (página pública → ZIP → Excel), sin API en vivo. Se diferencia en la FORMA del
dato: no es una serie temporal ``Record`` (una dimensión), sino un agregado
multidimensional por subclase (activos + split por tipo de persona). Por eso expone
``fetch_aggregates()`` en vez del ``fetch()`` serie-temporal del contrato base.

Reglas duras (§B.3):
- **No se ingieren las 190k filas crudas**: se agrega a nivel subclase; el detalle por
  provincia/municipio no se necesita para este caso de uso (segunda pasada si hiciera falta).
- Cada cifra viaja con su lineage (fuente, URL, fecha de corte) y el **caveat obligatorio**
  (contribuyentes ≠ locales; CIIU auto-declarada) — §B.1.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
import urllib.parse
import zipfile
from dataclasses import dataclass
from datetime import date
from typing import Dict, List, Optional, Tuple

from shared.data.base_client import Record, SourceClient
from shared.data.lineage import Lineage

logger = logging.getLogger("sdq.data.dgii_rnc")

PAGE_URL = "https://dgii.gov.do/estadisticas/Paginas/Estadisticas-Contribuyentes.aspx"
LICENSE = "datos públicos DGII (Estadísticas de Contribuyentes) — uso con cita; CIIU rev.3"
SHEET = "Cantidad de contribuyentes"
NOTES_SHEET = "Notas"
_HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP DGII contribuyentes connector)"}

# Nombres de columna en la hoja (mapeo por header, robusto a reordenamiento de columnas).
COL_SUBCLASE = "Subclase"
COL_TIPO = "Tipo de persona"
COL_ESTATUS = "Estatus"
COL_CANT = "Cantidad"

# El caveat DEBE viajar con el dato (§B.1): no queda solo en el spec.
CAVEAT = (
    "«Cantidad de contribuyentes» no equivale a «cantidad de locales»: una cadena puede "
    "operar varias tiendas bajo un mismo RNC. Además, la clasificación CIIU es "
    "auto-declarada por el contribuyente al inscribirse, por lo que puede haber "
    "inconsistencias de auto-clasificación."
)


@dataclass
class _Acc:
    """Acumulador mutable interno de ``aggregate_rows`` (una subclase)."""

    desc: str
    total: int = 0
    fisica: int = 0
    juridica: int = 0


@dataclass(frozen=True)
class SubclaseContrib:
    """Conteo de contribuyentes activos de una subclase CIIU en un corte."""

    subclase: str                 # código CIIU, p.ej. "552118"
    descripcion: str              # etiqueta de la subclase (parte tras el guion)
    activos: int                  # total de contribuyentes activos
    activos_fisica: int           # de ellos, personas físicas
    activos_juridica: int         # de ellos, personas jurídicas


@dataclass(frozen=True)
class Vertical:
    """Crosswalk subclase(s) CIIU → vertical en lenguaje natural (§B.3). ``terms`` son
    sinónimos que ayudan al retrieval léxico a matchear la pregunta sin conocer el código."""

    key: str
    label: str
    terms: str
    codes: Tuple[str, ...] = ()          # subclases exactas
    prefix: Optional[str] = None         # o un prefijo de grupo (p.ej. "552")
    scope_note: str = ""                 # aclaración de alcance (p.ej. superset que engloba a otro)

    def matches(self, code: str) -> bool:
        if self.prefix is not None:
            return code.startswith(self.prefix)
        return code in self.codes

    @property
    def code_label(self) -> str:
        """Etiqueta de códigos para citar en el pasaje ('552118+552113' o '552xxx')."""
        return "+".join(self.codes) if self.codes else f"{self.prefix}xxx"


# Verticales verificadas contra el archivo real (§B.1, cifras confirmadas al dígito).
VERTICALS: Tuple[Vertical, ...] = (
    Vertical("comida_rapida", "comida rápida",
             "restaurantes de comida rápida, fast food, cadenas de comida rápida",
             codes=("552118", "552113")),
    Vertical("restaurantes_bares_cantinas", "restaurantes, bares y cantinas",
             "restaurantes, bares, cantinas, servicios de comida y bebida",
             prefix="552",
             scope_note="Universo completo del grupo CIIU 552 (incluye a los de comida rápida)"),
    Vertical("hoteles", "hoteles",
             "hoteles, hotelería, servicios de alojamiento", prefix="551"),
    Vertical("supermercados", "supermercados",
             "supermercados, comercio mayorista y minorista de alimentos",
             codes=("513992", "521120")),
    Vertical("farmacias", "farmacias",
             "farmacias, droguerías, venta de productos farmacéuticos",
             codes=("523110", "513310")),
)

_MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
_CORTE_RE = re.compile(r"actualizados?\s+al\s+(\d{1,2})\s+de\s+([a-zñ]+)\s+de\s+(\d{4})")


class DGIIRncError(RuntimeError):
    """El padrón de contribuyentes de DGII estaba ausente, ilegible o cambió de estructura."""


def _to_int(v) -> int:
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (TypeError, ValueError):
        return 0


def _norm(s: str) -> str:
    """minúsculas sin acentos, para parsear la fecha de corte."""
    return "".join(c for c in unicodedata.normalize("NFKD", (s or "").lower())
                   if not unicodedata.combining(c))


def aggregate_rows(header: Tuple, data_rows) -> List[SubclaseContrib]:
    """Agrega filas del padrón a nivel subclase. ``header`` es la fila de encabezados;
    ``data_rows`` un iterable de tuplas de valores. Solo cuenta ``Estatus == Activo``
    (defensivo: el archivo ya viene filtrado, pero no dependemos de eso)."""
    try:
        col = {name: header.index(name) for name in (COL_SUBCLASE, COL_TIPO, COL_ESTATUS, COL_CANT)}
    except ValueError as e:  # columna esperada ausente → estructura cambió
        raise DGIIRncError(f"columna esperada ausente en el padrón DGII: {e}") from e

    acc: Dict[str, _Acc] = {}
    for row in data_rows:
        sub = row[col[COL_SUBCLASE]]
        if not sub:
            continue
        estatus = str(row[col[COL_ESTATUS]] or "").strip().lower()
        if estatus and estatus != "activo":
            continue
        code, _, desc = str(sub).partition("-")
        code = code.strip()
        if not code:
            continue
        cant = _to_int(row[col[COL_CANT]])
        tipo = _norm(str(row[col[COL_TIPO]] or ""))
        a = acc.setdefault(code, _Acc(desc.strip()))
        a.total += cant
        if "fis" in tipo:            # "física" / "persona física" → físicas
            a.fisica += cant
        elif "jur" in tipo:          # "jurídica" / "persona jurídica" → jurídicas
            a.juridica += cant
    return [SubclaseContrib(code, a.desc, a.total, a.fisica, a.juridica)
            for code, a in sorted(acc.items())]


def aggregate_xlsx(xlsx_bytes: bytes) -> List[SubclaseContrib]:
    """Streaming (read_only) del xlsx del padrón → agregados por subclase. Evita cargar
    las 190k filas en memoria de una."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise DGIIRncError(f"el Excel de DGII no trae la hoja '{SHEET}' (hojas: {wb.sheetnames})")
    ws = wb[SHEET]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if header is None:
        raise DGIIRncError("la hoja de contribuyentes está vacía")
    return aggregate_rows(tuple(header), it)


def parse_corte(xlsx_bytes: bytes) -> Optional[date]:
    """Fecha de corte de la hoja 'Notas' ('Datos actualizados al 29 de enero de 2026')."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if NOTES_SHEET not in wb.sheetnames:
        return None
    for row in wb[NOTES_SHEET].iter_rows(values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            m = _CORTE_RE.search(_norm(cell))
            if m and m.group(2) in _MESES:
                return date(int(m.group(3)), _MESES[m.group(2)], int(m.group(1)))
    return None


class DGIIRncClient(SourceClient):
    """Padrón de contribuyentes activos de DGII, agregado por subclase CIIU.

    ``live``: descarga el ZIP público de la página de estadísticas y parsea el Excel.
    ``fixture``: lee un Excel de muestra committeado (tests offline). Emite agregados
    vía :meth:`fetch_aggregates` — no el ``fetch()`` serie-temporal del contrato base."""

    source = "DGII"
    license = LICENSE
    license_ok = True
    fixture_file = "dgii_contribuyentes_sample.xlsx"

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        # Este conector emite agregados multidimensionales, no series temporales.
        raise NotImplementedError(
            "DGIIRncClient emite agregados por subclase: usar fetch_aggregates()")

    def fetch_aggregates(self) -> Tuple[Optional[date], List[SubclaseContrib], Lineage]:
        """(corte, filas por subclase, lineage). ``corte`` puede ser ``None`` si la hoja
        de Notas cambió de formato — el llamador decide si eso es un error."""
        self.check_license()
        if self.mode == "live":
            xlsx_bytes, url = self._download_live()
        else:
            xlsx_bytes, url = self._load_fixture_xlsx(), None
        corte = parse_corte(xlsx_bytes)
        rows = aggregate_xlsx(xlsx_bytes)
        lineage = Lineage(source=self.source, license=self.license, published_at=corte,
                          fetched_at=date.today(), url=url or PAGE_URL, note=CAVEAT)
        return corte, rows, lineage

    # ── Live ──────────────────────────────────────────────────────
    def _download_live(self) -> Tuple[bytes, str]:  # pragma: no cover - network I/O
        import httpx
        page = httpx.get(PAGE_URL, timeout=40, follow_redirects=True, headers=_HEADERS)
        page.raise_for_status()
        # No hardcodeamos el nombre del ZIP: el CDN de DGII renombra sin aviso. Buscamos
        # el href .zip que mencione 'contribuyentes'.
        zips = re.findall(r'href="([^"]+\.zip)"', page.text, re.IGNORECASE)
        link = next((z for z in zips if "contribuyente" in _norm(z)), zips[0] if zips else None)
        if not link:
            raise DGIIRncError("no se encontró el ZIP de contribuyentes en la página de DGII")
        url = urllib.parse.urljoin(PAGE_URL, urllib.parse.quote(link, safe="/:%"))
        resp = httpx.get(url, timeout=120, follow_redirects=True, headers=_HEADERS)
        resp.raise_for_status()
        zf = zipfile.ZipFile(io.BytesIO(resp.content))
        name = next((n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls"))), None)
        if not name:
            raise DGIIRncError("el ZIP de DGII no contiene un Excel")
        return zf.read(name), url

    # ── Fixture (offline / tests) ─────────────────────────────────
    def _load_fixture_xlsx(self) -> bytes:
        from pathlib import Path
        path = Path(__file__).resolve().parent / "fixtures" / self.fixture_file
        if not path.exists():
            raise DGIIRncError(f"fixture ausente: {path.name}")
        return path.read_bytes()


dgii_rnc_client = DGIIRncClient()
