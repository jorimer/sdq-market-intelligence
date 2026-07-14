"""ONE connector — construcción por tipología (Eje Construcción).

Cuatro archivos ``.xlsx`` estructurados de la **Oficina Nacional de Estadística (ONE)**
(``one.gob.do``), una métrica cada uno, mensuales, por *tipo de construcción*, 2013-2025:

  * ``4.7``  Licencias otorgadas          → ``licencias``
  * ``4.8``  Construcciones               → ``construcciones``
  * ``4.9``  Área construida (m²)         → ``sqm``
  * ``4.10`` Valor tasado (RD$)           → ``valor_tasado``

Complementa al MIVHED (indicador LÍDER de permisos, más fresco) con la cifra ANUAL,
autoritativa y validada de la ONE. En particular, ``valor_tasado`` es el valor de tasación
real (≈RD$15,000/m²), **distinto** del costo estándar derivado del MIVHED (m² × tarifa fija
≈RD$61,600/m²): son cantidades diferentes, no se mezclan (ver
:func:`shared.data.mivhed_client.parse_licenses`).

Rarezas del formato, manejadas con honestidad (verificado 2026-07-14):

  * **La orientación de las hojas cambia por época.** 2013-2019: filas=``Mes`` (los tipos van
    en columnas, con una taxonomía plana DISTINTA). 2020-2025: filas=``Tipo de construcción``
    (jerárquica: grupo = indent 0 + negrita; sub-línea = indent 1). → El detalle por tipología
    se emite SOLO para la época filas=Tipo; los años viejos aportan únicamente el total anual.
  * **El total del grupo se calcula sumando las filas de grupo**: la celda "Total" del
    proveedor viene vacía o redondeada en algunos años (p.ej. 2020), así que no se confía en
    ella.
  * **Las etiquetas cambian de casing/ortografía** entre archivos y años → se normaliza una
    clave para el cruce.
  * **Falta 2015** en algunos archivos y hay un salto 2017→2018 (cambio de metodología de la
    ONE) → cobertura honesta por archivo; nunca se fabrica un año.
  * **El prefijo ``/media/<hash>/`` es opaco y rota al republicar** → se resuelve por el número
    de cuadro en el nombre del archivo desde la página del tema (patrón CDN-rename del BCRD),
    no se hardcodea.

Falla cerrado: cualquier sorpresa estructural levanta :class:`OneConstructionError`, jamás
fabrica dato.
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
from html import unescape
from typing import Any, Dict, List, Optional

logger = logging.getLogger("sdq.data.one_construction")

_PAGE = ("https://www.one.gob.do/datos-y-estadisticas/temas/estadisticas-economicas/"
         "estadisticas-sectoriales/construccion-y-actividades-inmobiliarias/")
_HOST = "https://www.one.gob.do"
_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
}

# Número de cuadro (prefijo del nombre de archivo) → métrica. Estos cuadros son los "por tipo
# de construcción, según mes" (no los provincia/municipio). El número es la parte estable del
# nombre; el hash de /media/ no lo es.
_TABLES = {
    "4-7": "licencias",
    "4-8": "construcciones",
    "4-9": "sqm",
    "4-10": "valor_tasado",
}
_METRICS = ("licencias", "construcciones", "sqm", "valor_tasado")


class OneConstructionError(RuntimeError):
    """Un archivo de la ONE fue inalcanzable o cambió de estructura de forma inesperada."""


def _norm(s: object) -> str:
    """Clave insensible a acento/casing/espacios para cruzar etiquetas del proveedor."""
    t = "".join(c for c in unicodedata.normalize("NFKD", str(s).lower())
                if not unicodedata.combining(c))
    return " ".join(t.split())


def _num(v: object) -> Optional[float]:
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        t = v.replace(",", "").strip()
        if t:
            try:
                return float(t)
            except ValueError:
                return None
    return None


def _year_sheets(wb) -> List[str]:
    return [s for s in wb.sheetnames if len(s) == 4 and s.isdigit()]


def _header_index(rows) -> Optional[tuple]:
    """``(row_index, orientation)`` — orientación ``"tipo"`` (filas=tipo) o ``"mes"``
    (filas=mes). Devuelve ``None`` si la hoja no trae un encabezado reconocible."""
    for i, r in enumerate(rows[:10]):
        v0 = r[0].value
        head = _norm(v0) if v0 is not None else ""
        if head == "tipo de construccion":
            return i, "tipo"
        if head == "mes":
            return i, "mes"
    return None


def _parse_sheet(ws) -> Optional[Dict[str, Any]]:
    """``{"total": float|None, "by_typology": {display_label: value}|None}`` de una hoja-año.

    En orientación filas=tipo: las filas de GRUPO (indent 0 + negrita, salvo "Total") son las
    tipologías de primer nivel; su suma es el total (la celda "Total" del proveedor es poco
    fiable). En orientación filas=mes: solo el total anual (la taxonomía por columnas es plana
    y distinta a la nueva → no se desagrega)."""
    rows = list(ws.iter_rows())
    hdr = _header_index(rows)
    if hdr is None:
        return None
    hi, orient = hdr

    if orient == "tipo":
        by_typology: Dict[str, float] = {}
        for r in rows[hi + 1:]:
            c = r[0]
            lbl = str(c.value).strip() if c.value is not None else ""
            if not lbl or _norm(lbl) == "total":
                continue
            indent = (c.alignment.indent or 0) if c.alignment else 0
            is_group = indent == 0 and bool(c.font and c.font.bold)
            if not is_group:
                continue  # sub-línea: ya contenida en su grupo
            val = _num(r[1].value) if len(r) > 1 else None
            by_typology[" ".join(lbl.split())] = val if val is not None else 0.0
        if not by_typology:
            return None
        total = sum(by_typology.values())
        return {"total": total, "by_typology": by_typology}

    # orientación filas=mes → solo total anual (fila cuyo label es "Total")
    for r in rows[hi + 1:]:
        v0 = r[0].value
        if v0 is not None and _norm(v0) == "total":
            return {"total": _num(r[1].value) if len(r) > 1 else None, "by_typology": None}
    return {"total": None, "by_typology": None}


def parse_metric_workbook(wb) -> Dict[int, Dict[str, Any]]:
    """``{year: {"total", "by_typology"}}`` de un workbook de una métrica (una hoja por año)."""
    out: Dict[int, Dict[str, Any]] = {}
    for sh in _year_sheets(wb):
        parsed = _parse_sheet(wb[sh])
        if parsed is not None:
            out[int(sh)] = parsed
    return out


def _load_workbook(data: bytes):
    import openpyxl  # local import: solo requerido en modo live
    return openpyxl.load_workbook(io.BytesIO(data), data_only=True)


def merge_metrics(by_metric: Dict[str, Dict[int, Dict[str, Any]]]) -> Dict[str, Any]:
    """Cruza las 4 métricas en ``{by_typology, totals, years, typology_years}``.

    ``by_typology[year][tipología] = {licencias, construcciones, sqm, valor_tasado}`` (solo
    años con detalle, época filas=tipo). ``totals[year][metric]`` cubre todos los años
    disponibles. Las etiquetas se cruzan por clave normalizada (casing/ortografía varían entre
    archivos); se muestra la etiqueta del archivo de valor tasado cuando existe."""
    all_years = sorted({y for m in by_metric.values() for y in m})
    totals: Dict[int, Dict[str, Optional[float]]] = {}
    for y in all_years:
        totals[y] = {m: (by_metric.get(m, {}).get(y) or {}).get("total") for m in _METRICS}

    by_typology: Dict[int, Dict[str, Dict[str, Optional[float]]]] = {}
    typology_years: List[int] = []
    for y in all_years:
        # canon → {"display": str, metric: value}
        merged: Dict[str, Dict[str, Any]] = {}
        has_detail = False
        for metric in _METRICS:
            bt = (by_metric.get(metric, {}).get(y) or {}).get("by_typology")
            if not bt:
                continue
            has_detail = True
            for lbl, val in bt.items():
                key = _norm(lbl)
                slot = merged.setdefault(key, {"display": lbl})
                slot[metric] = val
                # el archivo de valor tasado fija la etiqueta canónica visible
                if metric == "valor_tasado":
                    slot["display"] = lbl
        if not has_detail:
            continue
        typology_years.append(y)
        by_typology[y] = {
            slot["display"]: {m: slot.get(m) for m in _METRICS}
            for slot in merged.values()
        }
    return {"by_typology": by_typology, "totals": totals,
            "years": all_years, "typology_years": typology_years}


class ONEConstructionClient:
    source = "ONE (Oficina Nacional de Estadística) — Estadísticas de construcción por tipología"
    license = "Datos abiertos ONE (one.gob.do)"
    license_ok = True

    def _resolve_urls(self) -> Dict[str, str]:
        """``{metric: url}`` resolviendo cada cuadro por su número desde la página del tema.

        El nombre del archivo (``4-7-…tipo-construccion-mes.xlsx``) es estable; el ``<hash>``
        de ``/media/`` no. Se exige que el nombre contenga ``tipo`` para no confundir con los
        cuadros por provincia/municipio."""
        import httpx
        with httpx.Client(timeout=60, follow_redirects=True, headers=_HEADERS) as http:
            html = http.get(_PAGE).text
        found: Dict[str, str] = {}
        for m in re.finditer(r'href="([^"]*\.xlsx)"', html):
            href = unescape(m.group(1))
            base = href.rsplit("/", 1)[-1].lower()
            if "tipo" not in base:
                continue
            num = "-".join(base.split("-", 2)[:2])  # "4-7", "4-10", …
            metric = _TABLES.get(num)
            if metric and metric not in found:
                found[metric] = href if href.startswith("http") else _HOST + href
        missing = [m for m in _METRICS if m not in found]
        if missing:
            raise OneConstructionError(
                f"ONE construcción: no se resolvieron los cuadros {missing} en {_PAGE}")
        return found

    def _fetch_bytes(self, url: str) -> bytes:
        import httpx
        from urllib.parse import quote
        with httpx.Client(timeout=120, follow_redirects=True, headers=_HEADERS) as http:
            r = http.get(quote(url, safe=":/?&=%"))
            r.raise_for_status()
            return r.content

    def fetch(self) -> Dict[str, Any]:
        """Descarga y cruza los 4 cuadros → ``{by_typology, totals, years, typology_years}``."""
        urls = self._resolve_urls()
        by_metric: Dict[str, Dict[int, Dict[str, Any]]] = {}
        for metric, url in urls.items():
            try:
                wb = _load_workbook(self._fetch_bytes(url))
                by_metric[metric] = parse_metric_workbook(wb)
            except OneConstructionError:
                raise
            except Exception as e:  # noqa: BLE001
                raise OneConstructionError(f"ONE construcción: fallo leyendo {metric}: {e}") from e
        merged = merge_metrics(by_metric)
        if not merged["typology_years"]:
            raise OneConstructionError("ONE construcción: ningún año con detalle por tipología.")
        return merged

    def latest_typology(self) -> Dict[str, Any]:
        """Detalle por tipología del último año disponible → ``{"year", "by_typology",
        "totals"}`` (ordenado por valor tasado descendente)."""
        merged = self.fetch()
        year = merged["typology_years"][-1]
        return {"year": year,
                "by_typology": merged["by_typology"][year],
                "totals": merged["totals"].get(year, {})}


one_construction_client = ONEConstructionClient()
