"""Ministerio de Hacienda — fiscal pulse connector (Estado de Operaciones).

The consolidated government fiscal accounts (GFS / *Estado de Operaciones del
Sector Público*) for Eje 2 (macro): revenue, expenditure and the fiscal
result/deficit, MONTHLY back to 2000. Far richer than tax collection alone — it is
the whole fiscal position. Source: the *Gobierno Central Presupuestario · Estado de
Operaciones mensual* Excel on hacienda.gob.do/estadisticas-fiscales (public).

We emit a CURATED set of top-level GFS lines (by their classification code), not
all ~130 rows: total revenue and its main tax groups, total expenditure and its
main items, capital investment, plus two computed balances (operating result =
revenue − expense; overall balance = operating result − net investment, i.e. the
GFS net lending/borrowing, the headline deficit/surplus). ``Record.dimension`` carries
the line slug. Missing cells stay absent (never fabricated).
"""
import io
import logging
import re
from datetime import date, datetime
from typing import Dict, List, Optional

from shared.data._text import norm
from shared.data.base_client import FixtureBackedClient, Record
from shared.data.lineage import Lineage

logger = logging.getLogger("sdq.data.hacienda")

ESTADISTICAS_URL = "https://www.hacienda.gob.do/estadisticas-fiscales/"
# The EO-mensual workbook link (its wp-content path carries a rotating version
# suffix, so it is discovered by name, not hard-coded).
EO_FRAGMENT = "estado-de-operaciones-mensual"
EO_REQUIRE = "gobierno-central"   # disambiguate from the GGC-consolidated / quarterly files
SHEET = "EO Mensual Pág. Web"
LICENSE = "datos públicos Ministerio de Hacienda (Estadísticas Fiscales) — uso con cita"
VAR = "fiscal"
UNIT = "RD$ millones"
_HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/124.0 Safari/537.36"}

_MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
          "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

# Curated GFS classification codes → slug (the operations section; codes are unique
# there — the financing section repeats 32/33, which we don't curate).
CODE_TO_SLUG: Dict[str, str] = {
    "1": "ingresos",
    "11": "impuestos",
    "111": "imp_ingresos",
    "113": "imp_propiedad",
    "114": "imp_bienes_servicios",
    "115": "imp_comercio_exterior",
    "12": "contribuciones_sociales",
    "14": "otros_ingresos",
    "2": "gastos",
    "21": "remuneracion",
    "24": "intereses",
    "25": "subsidios",
    "31": "inversion",
}


class HaciendaError(RuntimeError):
    """The Estado de Operaciones workbook was missing, unreadable or restructured."""


def _code_str(cell: object) -> str:
    """Normalize a GFS code cell ('1', 11, 111.0) to a plain string code."""
    if isinstance(cell, bool) or cell is None:
        return ""
    if isinstance(cell, (int, float)) and float(cell).is_integer():
        return str(int(cell))
    return str(cell).strip()


def discover_eo_url(html: str) -> Optional[str]:
    """Resolve the EO-mensual workbook URL from the Estadísticas Fiscales page."""
    from html import unescape

    best: Optional[tuple] = None  # (recency_key, url)
    for raw in {unescape(m) for m in re.findall(r'https://[^"\']+?\.xlsx', html, re.I)}:
        n = norm(raw)
        if EO_FRAGMENT in n and EO_REQUIRE in n:
            # later wp-content/uploads/YYYY/MM path = more recent
            ym = re.search(r"/uploads/(\d{4})/(\d{2})/", raw)
            key = (int(ym.group(1)), int(ym.group(2))) if ym else (0, 0)
            if best is None or key > best[0]:
                best = (key, raw)
    return best[1] if best else None


def _is_year(c: object) -> bool:
    return isinstance(c, (int, float)) and not isinstance(c, bool) and 2000 <= c <= 2035


def _period_columns(rows: List[list]) -> Dict[int, str]:
    """Map each data column → ``"YYYY-MM"`` from the year row + month row headers.

    Fail-closed and robust: the month row is the first with ≥12 month names; the
    year row is the row above it that carries a year in the MOST of those month
    columns (the real header covers them all — a stray caption with a few scattered
    years can't win). The year is forward-filled across the month columns; a month
    column with no year, or non-monotonic years, raises (a misaligned header must
    not produce garbage periods silently).
    """
    month_row_idx = None
    for i, r in enumerate(rows[:20]):
        if sum(1 for c in r if isinstance(c, str) and norm(c) in _MESES) >= 12:
            month_row_idx = i
            break
    if month_row_idx is None:
        raise HaciendaError("no se encontró la fila de meses en el Estado de Operaciones")
    months = rows[month_row_idx]
    month_cols = {j: norm(c) for j, c in enumerate(months) if isinstance(c, str) and norm(c) in _MESES}

    # year row = the one above months covering the most of those month columns
    year_row, best = None, 0
    for r in rows[:month_row_idx]:
        cov = sum(1 for j in month_cols if j < len(r) and _is_year(r[j]))
        if cov > best:
            best, year_row = cov, r
    if year_row is None or best == 0:
        raise HaciendaError("no se encontró la fila de años alineada con los meses")

    out: Dict[int, str] = {}
    last_year: Optional[int] = None
    for j in range(len(months)):
        if j < len(year_row) and _is_year(year_row[j]):
            last_year = int(year_row[j])
        if j in month_cols:
            if last_year is None:
                raise HaciendaError("columna de mes sin año en el Estado de Operaciones")
            out[j] = f"{last_year}-{_MESES.index(month_cols[j]) + 1:02d}"

    years = [int(p[:4]) for p in out.values()]
    if any(years[i] > years[i + 1] for i in range(len(years) - 1)):
        raise HaciendaError("años no monótonos en el Estado de Operaciones (header desalineado)")
    return out


def parse_eo_xlsx(content: bytes) -> Dict[str, Dict[str, float]]:
    """Parse the EO-mensual sheet → ``{slug: {YYYY-MM: value}}`` for curated GFS lines.

    Pure of network. Fail-closed if the sheet, the period headers, or the two
    headline lines (ingresos/gastos) are missing. Adds computed balances.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if SHEET not in wb.sheetnames:
            raise HaciendaError(f"hoja '{SHEET}' ausente en el Estado de Operaciones")
        rows = [list(r) for r in wb[SHEET].iter_rows(values_only=True)]
    finally:
        wb.close()

    col_period = _period_columns(rows)
    out: Dict[str, Dict[str, float]] = {}
    seen: set = set()
    for r in rows:
        code = _code_str(r[0]) if r else ""
        slug = CODE_TO_SLUG.get(code)
        if slug is None or code in seen:
            continue
        seen.add(code)
        obs: Dict[str, float] = {}
        for j, period in col_period.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                obs[period] = round(float(v), 3)
        out[slug] = obs

    if "ingresos" not in out or "gastos" not in out:
        raise HaciendaError("renglones Ingresos/Gastos ausentes — ¿cambió el Estado de Operaciones?")

    # computed balances: operating result and overall balance (the deficit/surplus)
    ing, gas, inv = out["ingresos"], out["gastos"], out.get("inversion", {})
    common = sorted(set(ing) & set(gas))
    out["resultado_operativo"] = {p: round(ing[p] - gas[p], 3) for p in common}
    out["balance_global"] = {p: round(ing[p] - gas[p] - inv.get(p, 0.0), 3) for p in common}
    return out


def build_records(groups: Dict[str, Dict[str, float]], *,
                  url: Optional[str] = None, published_at: Optional[date] = None) -> List[Record]:
    """One Record per (fiscal line, month)."""
    lineage = Lineage(source="Hacienda", license=LICENSE, fetched_at=date.today(),
                      url=url, published_at=published_at,
                      note="Estado de Operaciones del Gobierno Central Presupuestario (mensual)")
    out: List[Record] = []
    for slug in list(CODE_TO_SLUG.values()) + ["resultado_operativo", "balance_global"]:
        for period in sorted(groups.get(slug, {})):
            out.append(Record(series=VAR, period=period, value=groups[slug][period],
                              lineage=lineage, unit=UNIT, dimension=slug))
    return out


def _filter(records: List[Record], series: Optional[str], period: Optional[str]) -> List[Record]:
    if series:
        records = [r for r in records if r.series == series]
    if period:
        records = [r for r in records if r.period == period]
    return records


class HaciendaClient(FixtureBackedClient):
    """Ministerio de Hacienda Estado de Operaciones (Eje 2 fiscal pulse)."""

    source = "Hacienda"
    license = LICENSE
    license_ok = True
    fixture_file = "hacienda_eo.json"
    live_phase = "Fase 4 (Eje 2 · pulso fiscal)"

    def fetch(self, series: Optional[str] = None, period: Optional[str] = None) -> List[Record]:
        self.check_license()
        if self.mode == "live":
            return self._fetch_live(series, period)
        return self._fetch_fixture(series, period)

    # ── Live ──────────────────────────────────────────────────────
    def _fetch_live(self, series: Optional[str], period: Optional[str]) -> List[Record]:  # pragma: no cover - network I/O
        import httpx

        page = httpx.get(ESTADISTICAS_URL, timeout=40, follow_redirects=True, headers=_HEADERS)
        page.raise_for_status()
        url = discover_eo_url(page.text)
        if not url:
            raise HaciendaError("no se encontró el Excel del Estado de Operaciones mensual en Hacienda")
        f = httpx.get(url, timeout=90, follow_redirects=True, headers=_HEADERS)
        f.raise_for_status()
        groups = parse_eo_xlsx(f.content)
        return _filter(build_records(groups, url=url, published_at=self._published_at(url)), series, period)

    @staticmethod
    def _published_at(url: str) -> Optional[date]:  # pragma: no cover - network I/O
        import httpx
        try:
            resp = httpx.head(url, timeout=20, follow_redirects=True, headers=_HEADERS)
            lm = resp.headers.get("last-modified")
            if lm:
                return datetime.strptime(lm, "%a, %d %b %Y %H:%M:%S %Z").date()
        except Exception as e:  # noqa: BLE001 — provenance best-effort
            logger.warning("[hacienda] no se pudo leer Last-Modified: %s", e)
        return None

    # ── Fixture (offline / tests) ─────────────────────────────────
    def _fetch_fixture(self, series: Optional[str], period: Optional[str]) -> List[Record]:
        """Fixture shape: ``{"<slug>": {"fiscal": {"<YYYY-MM>": value}}}``."""
        fixture = self._load_fixture(self.fixture_file)
        lineage = Lineage(source=self.source, license=self.license, fetched_at=date.today())
        out: List[Record] = []
        for slug, vars_ in fixture.items():
            for var, obs in vars_.items():
                for period_, val in obs.items():
                    out.append(Record(series=var, period=str(period_),
                                      value=None if val is None else float(val),
                                      lineage=lineage, unit=UNIT, dimension=slug))
        return _filter(out, series, period)


hacienda_client = HaciendaClient()
