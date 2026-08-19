"""INDOTEL connector — DR telecom sector indicators.

Public open data from INDOTEL's quarterly statistical-indicators bulletins
(``indotel.gob.do``), machine-readable XLSX. The bulletin's "Ind. Generales"
sheet carries the headline totals by CÓDIGO (line totals LTT, internet
subscriptions TSI, broadband TSIBA), and "RESUMEN" the annual revenue series.

NOTE (data reality, 2026-06): the latest bulletin PUBLISHED on the canonical page
is 2022-Q1; INDOTEL's public quarterly series appears frozen there. The connector
ingests the latest published bulletin honestly — freshness (quarterly cadence) will
reflect its true age, so the product stays cabled-but-not-publishable until a newer
bulletin appears. Never fabricated. Re-point ``LATEST_BULLETIN`` when INDOTEL
publishes a newer file (or wire the INDOTEL–ONE automated portal when it exists).
"""
import logging
from typing import Any, Dict, Optional

logger = logging.getLogger("sdq.data.indotel")

_HEADERS = {"User-Agent": "Mozilla/5.0 (SDQ-MIP)"}

# Latest published quarterly bulletin (XLSX) + the period it covers.
LATEST_BULLETIN = (
    "https://indotel.gob.do/wp-content/uploads/2022/10/"
    "indicadores-estadisticos-de-telecomunicaciones-trimestrales-enero-marzo-2022-res-026-21-web.xlsx"
)
LATEST_PERIOD = "2022-Q1"

# DR population — ONE, X Censo Nacional de Población y Vivienda 2022 (oficial).
# Denominador real de las tasas de penetración (no fabricado).
POP_2022 = 10_760_028


def parse_indicators(xlsx_bytes: bytes) -> Dict[str, Optional[float]]:
    """Extract the headline telecom totals from a bulletin XLSX (bytes).

    Reads "Ind. Generales" (CÓDIGO in col 3, TOTALES in col 5) for the latest month
    of the quarter — line totals (LTT.3er), internet subs (TSI.3er), broadband
    (TSIBA) — and "RESUMEN" for the annual revenue series. Missing → None (never
    fabricated)."""
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    codes: Dict[str, float] = {}
    if "Ind. Generales" in wb.sheetnames:
        for row in wb["Ind. Generales"].iter_rows(values_only=True):
            code = row[2] if len(row) > 2 else None
            total = row[4] if len(row) > 4 else None
            if code and isinstance(total, (int, float)):
                codes[str(code).strip()] = float(total)

    def _pick(*keys: str) -> Optional[float]:
        for k in keys:
            if k in codes:
                return codes[k]
        return None

    lines = _pick("LTT.3er", "LTT.2do", "LTT.1er")
    internet = _pick("TSI.3er", "TSI.2do", "TSI.1er")
    broadband = _pick("TSIBA")

    revenue_latest = revenue_prev = None
    if "RESUMEN" in wb.sheetnames:
        for row in wb["RESUMEN"].iter_rows(values_only=True):
            if row and row[0] and "ngresos Totales" in str(row[0]):
                nums = [c for c in row[1:] if isinstance(c, (int, float))]
                # El RESUMEN lista años completos seguidos del trimestre parcial en curso
                # (p.ej. 2019, 2020, 2021, Ene-Mar.22). Para nivel/crecimiento ANUAL se
                # descarta el parcial final y se toman los DOS últimos años completos
                # (robusto a cuántos años traiga el boletín, no índices fijos).
                complete = nums[:-1] if len(nums) >= 2 else nums
                if len(complete) >= 2:
                    revenue_latest, revenue_prev = complete[-1], complete[-2]
                elif complete:
                    revenue_latest = complete[-1]
                break

    return {
        "lines_total": lines,
        "internet_total": internet,
        "broadband_total": broadband,
        "revenue_latest": revenue_latest,
        "revenue_prev": revenue_prev,
    }


class INDOTELClient:
    source = "INDOTEL"
    license = "Datos Abiertos RD (INDOTEL)"
    license_ok = True

    def fetch_indicators(self) -> Dict[str, Any]:
        """Download the latest published bulletin and parse the headline indicators.

        Returns ``{period, **indicators}``. Raises on network failure (caller treats
        the sync as best-effort)."""
        import httpx

        with httpx.Client(timeout=120, follow_redirects=True, headers=_HEADERS) as http:
            content = http.get(LATEST_BULLETIN).content
        return {"period": LATEST_PERIOD, **parse_indicators(content)}


indotel_client = INDOTELClient()


# ── Vigilancia del boletín congelado ────────────────────────────────────────────
#
# El boletín público de INDOTEL se congeló en 2022-Q1 y el IDT pasó a servirse de ITU
# DataHub (ver ``itu_client``). Eso convirtió una brecha en una DECISIÓN — pero una decisión
# que, sin vigilancia, hay que volver a tomar a mano cada vez que alguien se pregunta si
# INDOTEL volvió. Esta sonda contesta esa pregunta sola: descubre el boletín más reciente
# publicado y lo compara con el que el conector conoce.
#
# NO ingiere ni promueve: propone. Si aparece uno más nuevo, entra al tablero de
# Inteligencia de Fuentes para que el dueño decida (el sistema propone, el dueño dispone).

_PAGINAS_BOLETIN = (
    "https://indotel.gob.do/estadisticas/indicadores-estadisticos-trimestrales/",
    "https://indotel.gob.do/estadisticas/",
)

# Un boletín trimestral se nombra por el rango de meses que cubre.
_TRIMESTRES = {
    ("enero", "marzo"): "Q1", ("abril", "junio"): "Q2",
    ("julio", "septiembre"): "Q3", ("octubre", "diciembre"): "Q4",
}


def _periodo_de_nombre(nombre: str) -> Optional[str]:
    """``…enero-marzo-2024…`` → ``"2024-Q1"``. ``None`` si el nombre no lo declara."""
    import re

    bajo = nombre.lower()
    for (desde, hasta), q in _TRIMESTRES.items():
        m = re.search(rf"{desde}[\-_ ]+{hasta}[\-_ ]+(20\d\d)", bajo)
        if m:
            return f"{m.group(1)}-{q}"
    return None


def descubrir_boletin_mas_reciente(timeout: int = 45) -> Dict[str, Any]:
    """Busca en las páginas oficiales el boletín trimestral más nuevo publicado.

    Devuelve ``{"periodo", "url", "hay_mas_nuevo", "conocido", "paginas_ok"}``. Nunca
    levanta: una sonda que rompe la operación deja de correr, y entonces no vigila nada.
    """
    import re

    import httpx

    encontrados: Dict[str, str] = {}
    paginas_ok = 0
    for pagina in _PAGINAS_BOLETIN:
        try:
            r = httpx.get(pagina, headers=_HEADERS, timeout=timeout, follow_redirects=True)
            if r.status_code != 200:
                continue
            paginas_ok += 1
            for url in re.findall(r'href="(https?://[^"]+\.xlsx?)"', r.text, re.I):
                periodo = _periodo_de_nombre(url)
                if periodo:
                    encontrados.setdefault(periodo, url)
        except Exception as e:  # noqa: BLE001 — la sonda informa, no rompe
            logger.warning("[INDOTEL] sonda no pudo leer %s: %s", pagina, e)

    mas_nuevo = max(encontrados) if encontrados else None
    return {
        "periodo": mas_nuevo,
        "url": encontrados.get(mas_nuevo) if mas_nuevo else None,
        "conocido": LATEST_PERIOD,
        "hay_mas_nuevo": bool(mas_nuevo and mas_nuevo > LATEST_PERIOD),
        "paginas_ok": paginas_ok,
        "periodos_vistos": sorted(encontrados),
    }
